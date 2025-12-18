from openpyxl import load_workbook

def norm(s):
    """Нормализация строки для надёжного сравнения."""
    if s is None:
        return ""
    return str(s).strip().lower()

# ====== НАСТРОЙКИ ======
file1_path = "file1.xlsx"   # первый файл
file2_path = "file2.xlsx"   # второй файл
out_path   = "file1_fixed.xlsx"  # результат

sheet1_name = None  # None = активный лист
sheet2_name = None  # None = активный лист

# В каких колонках что лежит:
file1_name_col = "A"  # имена в 1-м файле
file1_write_col = "B" # сюда пишем цифры

file2_name_col = "D"  # имя сервиса во 2-м файле
file2_value_col = "B" # отсюда берём цифры
# =======================

# --- читаем 2-й файл и строим словарь: имя -> значение ---
wb2 = load_workbook(file2_path, data_only=True)
ws2 = wb2[sheet2_name] if sheet2_name else wb2.active

map_name_to_value = {}
duplicates = set()

for row in range(1, ws2.max_row + 1):
    name = ws2[f"{file2_name_col}{row}"].value
    val  = ws2[f"{file2_value_col}{row}"].value

    key = norm(name)
    if not key:
        continue

    # если одно и то же имя встречается несколько раз — запомним как дубликат
    if key in map_name_to_value:
        duplicates.add(key)
    map_name_to_value[key] = val  # последнее значение победит

# --- открываем 1-й файл и заполняем колонку B ---
wb1 = load_workbook(file1_path)
ws1 = wb1[sheet1_name] if sheet1_name else wb1.active

filled = 0
missing = 0

for row in range(1, ws1.max_row + 1):
    name = ws1[f"{file1_name_col}{row}"].value
    key = norm(name)

    if not key:
        continue

    if key in map_name_to_value:
        ws1[f"{file1_write_col}{row}"].value = map_name_to_value[key]
        filled += 1
    else:
        # можно оставить как есть или очистить — я очищаю, чтобы было видно что не найдено
        ws1[f"{file1_write_col}{row}"].value = None
        missing += 1

wb1.save(out_path)

print("ГОТОВО")
print(f"Заполнено строк: {filled}")
print(f"Не найдено совпадений: {missing}")
if duplicates:
    print(f"ВНИМАНИЕ: дубликаты имён во 2-м файле (показаны первые 20): {list(duplicates)[:20]}")
