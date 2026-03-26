import os
import logging
from collections import defaultdict

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font
from openpyxl import load_workbook
from zabbix_utils import ZabbixAPI

load_dotenv()

ZABBIX_URL = os.getenv("ZABBIX_URL", "").rstrip("/")
ZABBIX_TOKEN = os.getenv("ZABBIX_TOKEN", "")

DB_FILE = os.getenv("DB_FILE")
ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "zabbix_report.xlsx")

BAN_SERVICE_IDS = [15473]
SKIP_EMPTY_SERVICE_ID = True
USE_LAST_AMBIGUOUS_MATCH = True

ZBX_CHUNK = 500

logger = logging.getLogger("zabbix_report")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
handler.setFormatter(
    logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S")
)
logger.handlers.clear()
logger.addHandler(handler)


def die(msg: str, code: int = 2):
    logger.error(msg)
    raise SystemExit(code)


def validate_env_and_files():
    if not ZABBIX_URL:
        die("ENV ZABBIX_URL пустой")
    if not ZABBIX_TOKEN:
        die("ENV ZABBIX_TOKEN пустой")

    if not DB_FILE:
        die("ENV DB_FILE пустой")
    if not os.path.isfile(DB_FILE):
        die(f"DB_FILE не найден: {DB_FILE}")

    if not ACTIVITY_FILE:
        die("ENV ACTIVITY_FILE пустой")
    if not os.path.isfile(ACTIVITY_FILE):
        die(f"ACTIVITY_FILE не найден: {ACTIVITY_FILE}")

    out_dir = os.path.dirname(OUTPUT_XLSX)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)


def build_ban_set(ban_list):
    if not isinstance(ban_list, (list, tuple, set)):
        die("BAN_SERVICE_IDS должен быть list / tuple / set")
    return {str(x).strip() for x in ban_list if str(x).strip()}


def clean_spaces(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def clean_dns(s: str) -> str:
    return (s or "").strip().lower()


def clean_ip_only_32(s: str) -> str:
    s = (s or "").strip()
    if s.endswith("/32"):
        s = s[:-3].strip()
    return s


def pick_primary_interface(interfaces):
    if not interfaces:
        return {}
    for it in interfaces:
        if str(it.get("main")) == "1":
            return it
    return interfaces[0]


def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i : i + n]


def fetch_hosts(api):
    raw = api.host.get(
        output=["hostid", "host", "name", "status"],
        selectInterfaces=["ip", "dns", "main"],
    )
    rows = []
    for h in raw or []:
        if int(h.get("status", 0)) != 0:
            continue
        iface = pick_primary_interface(h.get("interfaces") or [])
        rows.append(
            {
                "hostid": str(h.get("hostid", "")).strip(),
                "host": (h.get("host") or "").strip(),
                "ip": clean_ip_only_32(iface.get("ip")),
                "dns": clean_dns(iface.get("dns")),
            }
        )
    return pd.DataFrame(rows)


def fetch_items_triggers_counts(api, hostids):
    items_cnt = defaultdict(int)
    triggers_cnt = defaultdict(int)

    hostids = [str(x).strip() for x in hostids if str(x).strip()]
    if not hostids:
        return dict(items_cnt), dict(triggers_cnt)

    for part in chunks(hostids, ZBX_CHUNK):
        try:
            items = api.item.get(
                hostids=part,
                output=["itemid", "hostid", "status"],
            )
        except Exception as e:
            logger.error("item.get failed: %s", e)
            items = []

        for it in items or []:
            if str(it.get("status", "0")) != "0":
                continue
            hid = str(it.get("hostid", "")).strip()
            if hid:
                items_cnt[hid] += 1

    for part in chunks(hostids, ZBX_CHUNK):
        try:
            trigs = api.trigger.get(
                hostids=part,
                output=["triggerid", "status"],
                selectHosts=["hostid"],
            )
        except Exception as e:
            logger.error("trigger.get failed: %s", e)
            trigs = []

        for tr in trigs or []:
            if str(tr.get("status", "0")) != "0":
                continue
            for h in (tr.get("hosts") or []):
                hid = str(h.get("hostid", "")).strip()
                if hid:
                    triggers_cnt[hid] += 1

    return dict(items_cnt), dict(triggers_cnt)


def load_db(path):
    df = pd.read_excel(path, usecols="A,B,N,V", dtype=str, engine="openpyxl")
    df.columns = ["service", "service_id", "dns", "ip"]
    df = df.fillna("")

    df["service"] = df["service"].astype(str).map(clean_spaces)
    df["service_id"] = df["service_id"].astype(str).str.strip()
    df["ip"] = df["ip"].astype(str).map(clean_ip_only_32)
    df["dns"] = df["dns"].astype(str).map(clean_dns)

    return df


def load_activity_map(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    out = {}

    for row in ws.iter_rows(values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else None)
        if not code:
            continue

        if code in out:
            continue

        out[code] = {
            "service_name": clean_spaces(row[1] if len(row) > 1 else ""),
            "activity_code": clean_spaces(row[2] if len(row) > 2 else ""),
            "activity_name": clean_spaces(row[3] if len(row) > 3 else ""),
        }

    wb.close()
    logger.info("ACTIVITY загружено сервисов: %d", len(out))
    return out


def build_map(df, keys):
    tmp = df.copy()
    tmp["_k"] = list(map(tuple, tmp[keys].to_numpy()))
    counts = tmp["_k"].value_counts().to_dict()
    last = tmp.drop_duplicates("_k", keep="last")
    mp = {
        k: (r["service"], r["service_id"])
        for k, r in zip(last["_k"], last.to_dict("records"))
    }
    return mp, counts


def main():
    logger.info("Старт формирования отчета Zabbix (hosts/items/triggers)")

    validate_env_and_files()

    ban_set = build_ban_set(BAN_SERVICE_IDS)
    logger.info("BAN_SERVICE_IDS=%s", sorted(ban_set) if ban_set else "[]")
    logger.info("USE_LAST_AMBIGUOUS_MATCH=%s", USE_LAST_AMBIGUOUS_MATCH)

    api = ZabbixAPI(url=ZABBIX_URL)
    try:
        api.login(token=ZABBIX_TOKEN)

        df_hosts = fetch_hosts(api)
        total_active = len(df_hosts)

        items_cnt, triggers_cnt = fetch_items_triggers_counts(
            api, df_hosts["hostid"].tolist() if not df_hosts.empty else []
        )
    finally:
        try:
            api.logout()
        except Exception:
            pass

    if df_hosts.empty:
        die("В Zabbix не найдено активных хостов (status=0)")

    df_hosts["items"] = df_hosts["hostid"].map(lambda x: int(items_cnt.get(str(x), 0)))
    df_hosts["triggers"] = df_hosts["hostid"].map(
        lambda x: int(triggers_cnt.get(str(x), 0))
    )

    df_db = load_db(DB_FILE)
    activity_map = load_activity_map(ACTIVITY_FILE)

    map_both, dup_both = build_map(df_db, ["ip", "dns"])
    map_ip, dup_ip = build_map(df_db[df_db["ip"] != ""], ["ip"])
    map_dns, dup_dns = build_map(df_db[df_db["dns"] != ""], ["dns"])

    per_service = {}
    matched_hosts = 0
    ambiguous = 0
    ambiguous_skipped = 0
    banned_hits = 0
    skipped_empty = 0
    activity_miss = 0
    unaccounted = []

    def add_unacc(
        r,
        reason,
        detail,
        service="",
        service_id="",
        service_name="",
        activity_code="",
        activity_name="",
        amb=False,
    ):
        unaccounted.append(
            {
                "Хост": r.get("host", ""),
                "IP": r.get("ip", ""),
                "DNS": r.get("dns", ""),
                "Айтемы": int(r.get("items", 0)),
                "Триггеры": int(r.get("triggers", 0)),
                "service_from_db": service,
                "КОД": service_id,
                "Имя сервиса": service_name,
                "Код активности": activity_code,
                "Наименование активности": activity_name,
                "ambiguous": "yes" if amb else "",
                "reason": reason,
                "detail": detail,
            }
        )

    for r in df_hosts.to_dict("records"):
        ip = r.get("ip", "")
        dns = r.get("dns", "")
        svc_key = None
        amb = False
        matched_by = ""

        if ip and dns and (ip, dns) in map_both:
            svc_key = map_both[(ip, dns)]
            amb = dup_both.get((ip, dns), 0) > 1
            matched_by = "ip+dns"
        elif ip and (ip,) in map_ip:
            svc_key = map_ip[(ip,)]
            amb = dup_ip.get((ip,), 0) > 1
            matched_by = "ip"
        elif dns and (dns,) in map_dns:
            svc_key = map_dns[(dns,)]
            amb = dup_dns.get((dns,), 0) > 1
            matched_by = "dns"

        if not svc_key:
            add_unacc(r, "no_match_in_db", "no match by (ip,dns) / ip / dns")
            continue

        service, service_id = svc_key
        service_id = normalize_code(service_id)

        meta = activity_map.get(service_id, {})
        activity_service_name = meta.get("service_name", "")
        activity_code = meta.get("activity_code", "")
        activity_name = meta.get("activity_name", "")

        if amb and not USE_LAST_AMBIGUOUS_MATCH:
            ambiguous += 1
            ambiguous_skipped += 1
            add_unacc(
                r,
                "ambiguous_match",
                f"matched_by={matched_by} and DB has duplicates for key; skipped because USE_LAST_AMBIGUOUS_MATCH=False",
                service=service,
                service_id=service_id,
                service_name=activity_service_name,
                activity_code=activity_code,
                activity_name=activity_name,
                amb=True,
            )
            continue

        if SKIP_EMPTY_SERVICE_ID and not service_id:
            skipped_empty += 1
            add_unacc(
                r,
                "empty_service_id",
                "SKIP_EMPTY_SERVICE_ID=True and service_id empty",
                service=service,
                service_id=service_id,
                service_name=activity_service_name,
                activity_code=activity_code,
                activity_name=activity_name,
                amb=amb,
            )
            continue

        if service_id in ban_set:
            banned_hits += 1
            add_unacc(
                r,
                "banned_service_id",
                "service_id in BAN_SERVICE_IDS",
                service=service,
                service_id=service_id,
                service_name=activity_service_name,
                activity_code=activity_code,
                activity_name=activity_name,
                amb=amb,
            )
            continue

        if service_id not in activity_map:
            activity_miss += 1
            add_unacc(
                r,
                "activity_mapping_miss",
                "service_id not found in activity.xlsx",
                service=service,
                service_id=service_id,
                service_name=activity_service_name,
                activity_code=activity_code,
                activity_name=activity_name,
                amb=amb,
            )
            continue

        matched_hosts += 1

        if amb:
            ambiguous += 1
            add_unacc(
                r,
                "ambiguous_match",
                f"matched_by={matched_by} and DB has duplicates for key; using last match because USE_LAST_AMBIGUOUS_MATCH=True",
                service=service,
                service_id=service_id,
                service_name=activity_service_name,
                activity_code=activity_code,
                activity_name=activity_name,
                amb=True,
            )

        key = service_id
        if key not in per_service:
            per_service[key] = {
                "service_name": activity_service_name or service,
                "service_id": service_id,
                "activity_code": activity_code,
                "activity_name": activity_name,
                "hosts": 0,
                "items": 0,
                "triggers": 0,
            }

        per_service[key]["hosts"] += 1
        per_service[key]["items"] += int(r.get("items", 0))
        per_service[key]["triggers"] += int(r.get("triggers", 0))

    candidates = []
    for service_id, cnts in per_service.items():
        candidates.append(
            {
                "Имя сервиса": cnts.get("service_name", ""),
                "Код": cnts.get("service_id", ""),
                "Код активности": cnts.get("activity_code", ""),
                "Наименование активности": cnts.get("activity_name", ""),
                "Кол-во хостов": int(cnts["hosts"]),
                "Кол-во айтемов": int(cnts["items"]),
                "Кол-во триггеров": int(cnts["triggers"]),
            }
        )

    eligible_items_total = sum(x["Кол-во айтемов"] for x in candidates)

    for x in candidates:
        pct = (
            (x["Кол-во айтемов"] / eligible_items_total) * 100
            if eligible_items_total
            else 0.0
        )
        x["% потребления (items)"] = round(pct, 2)

    if candidates:
        df_report = pd.DataFrame(candidates).sort_values(
            ["Кол-во айтемов", "Кол-во хостов"],
            ascending=[False, False],
        )
    else:
        df_report = pd.DataFrame(
            columns=[
                "Имя сервиса",
                "Код",
                "Код активности",
                "Наименование активности",
                "Кол-во хостов",
                "Кол-во айтемов",
                "Кол-во триггеров",
                "% потребления (items)",
            ]
        )

    if unaccounted:
        df_unacc = pd.DataFrame(unaccounted).sort_values(
            ["reason", "Айтемы", "Триггеры"], ascending=[True, False, False]
        )
    else:
        df_unacc = pd.DataFrame(
            columns=[
                "Хост",
                "IP",
                "DNS",
                "Айтемы",
                "Триггеры",
                "service_from_db",
                "КОД",
                "Имя сервиса",
                "Код активности",
                "Наименование активности",
                "ambiguous",
                "reason",
                "detail",
            ]
        )

    installed_pct = (matched_hosts / total_active * 100) if total_active else 0.0

    logger.info("Активных хостов: %d", total_active)
    logger.info("Установлено/определено хостов: %d (%.2f%%)", matched_hosts, installed_pct)
    logger.info("Сомнительных совпадений: %d", ambiguous)
    logger.info("Сомнительных совпадений, исключенных из отчета: %d", ambiguous_skipped)
    logger.info("Скип по бан-листу (service_id): %d", banned_hits)
    logger.info("Скип пустых service_id: %d", skipped_empty)
    logger.info("Скип по отсутствию в activity.xlsx: %d", activity_miss)
    logger.info("Итого айтемов (eligible): %d", eligible_items_total)
    logger.info("Сервисов в отчёте: %d", len(df_report))
    logger.info("Unaccounted строк: %d", len(df_unacc))

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df_report.to_excel(writer, index=False, sheet_name="Отчет Zabbix")
        df_unacc.to_excel(writer, index=False, sheet_name="Unaccounted")

        wb = writer.book
        for sheet in ("Отчет Zabbix", "Unaccounted"):
            ws = wb[sheet]
            for c in ws[1]:
                c.font = Font(bold=True)


if __name__ == "__main__":
    main()