import os
import logging
import requests
import urllib3
import re

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)

load_dotenv()

SONAR_URL = os.getenv("SONAR_URL", "").rstrip("/")
SONAR2_URL = os.getenv("SONAR2_URL", "").rstrip("/")

SONAR_TOKEN = os.getenv("SONAR_TOKEN", "")
SONAR2_TOKEN = os.getenv("SONAR2_TOKEN", "")

OUT_FILE = os.getenv("OUT_FILE", "sonarQube_report.xlsx")
ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

BAN_SERVICE_IDS = {15473}

CODE_ANYWHERE_RE = re.compile(r"(?:^|[^0-9])(\d{2,})(?:[^0-9]|$)")


def clean(s):
    if s is None:
        return ""
    return " ".join(str(s).replace(",", " ").split())


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def split_service_name_code(prefix: str):
    p = (prefix or "").strip()
    if not p:
        return "", ""

    m = CODE_ANYWHERE_RE.search(p)
    if not m:
        return p, ""

    return p, m.group(1)


def make_session(token):
    s = requests.Session()
    s.auth = (token, "")
    s.headers.update({"Accept": "application/json"})
    return s


def sonar_get(s, url, path, params):
    r = s.get(f"{url}{path}", params=params, verify=False, timeout=60)
    r.raise_for_status()
    return r.json()


def get_projects(s, url):
    res = []
    p = 1
    while True:
        d = sonar_get(s, url, "/api/projects/search", {"p": p, "ps": 500})
        batch = d.get("components", []) or []
        res.extend(batch)
        total = int((d.get("paging") or {}).get("total") or 0)
        logger.info("projects page=%d fetched=%d total=%d", p, len(batch), total)
        if p * 500 >= total:
            break
        p += 1
    return res


def get_tasks(s, url, key, label):
    res = []
    p = 1
    while True:
        d = sonar_get(
            s,
            url,
            "/api/ce/activity",
            {
                "component": key,
                "status": "IN_PROGRESS,SUCCESS,FAILED,CANCELED",
                "p": p,
                "ps": 100,
            },
        )
        batch = d.get("tasks", []) or []
        res.extend(batch)
        total = int((d.get("paging") or {}).get("total") or 0)
        logger.info("[%s] tasks project=%s page=%d fetched=%d total=%d", label, key, p, len(batch), total)
        if p * 100 >= total:
            break
        p += 1
    return res


def measure(s, url, key, metric, branch=None, pr=None):
    params = {"component": key, "metricKeys": metric}
    if branch:
        params["branch"] = branch
    if pr:
        params["pullRequest"] = str(pr)

    d = sonar_get(s, url, "/api/measures/component", params)
    measures = (d.get("component", {}) or {}).get("measures") or [{}]
    m = measures[0] if measures else {}
    v = (m.get("period") or {}).get("value") if pr else m.get("value")

    try:
        return int(float(v))
    except Exception:
        return 0


def calc_project_tasks_lines(s, url, key, tasks, ncloc_cache, new_cache):
    tcnt = 0
    lines = 0

    for t in tasks:
        tcnt += 1
        pr = t.get("pullRequest")
        br = t.get("branch")

        if pr:
            ck = (key, pr)
            if ck not in new_cache:
                new_cache[ck] = measure(s, url, key, "new_lines", pr=pr)
            lines += new_cache[ck]
        else:
            b = br or "__main__"
            ck = (key, b)
            if ck not in ncloc_cache:
                ncloc_cache[ck] = measure(
                    s, url, key, "ncloc", branch=None if b == "__main__" else b
                )
            lines += ncloc_cache[ck]

    return tcnt, lines


def load_activity(path):
    if not path or not os.path.exists(path):
        raise RuntimeError(f"ACTIVITY_FILE не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    m = {}

    for r in ws.iter_rows(values_only=True):
        code = normalize_code(r[0] if len(r) > 0 else "")
        if not code:
            continue

        if code in m:
            continue

        m[code] = {
            "service_name": clean(r[1] if len(r) > 1 else ""),
            "activity_code": clean(r[2] if len(r) > 2 else ""),
            "activity_name": clean(r[3] if len(r) > 3 else ""),
        }

    wb.close()
    logger.info("activity loaded: %d", len(m))
    return m


def process_sonar(label, url, token, activity, acc, unacc):
    if not url or not token:
        logger.warning("[%s] skipped: url or token is empty", label)
        return

    logger.info("[%s] start processing url=%s", label, url)

    s = make_session(token)

    try:
        projects = get_projects(s, url)
    except Exception as e:
        logger.exception("[%s] failed to load projects: %s", label, e)
        return

    logger.info("[%s] projects found: %d", label, len(projects))

    ncloc_cache = {}
    new_cache = {}

    accounted_projects = 0
    unaccounted_projects = 0

    for idx, p in enumerate(projects, start=1):
        key = p.get("key")
        if not key:
            logger.warning("[%s] project without key skipped", label)
            continue

        prefix = key.split(":", 1)[0]
        svc_guess, code = split_service_name_code(prefix)

        logger.info("[%s] project %d/%d key=%s prefix=%s parsed_code=%s", label, idx, len(projects), key, prefix, code or "EMPTY")

        if not code:
            unaccounted_projects += 1
            logger.warning("[%s] unaccounted project=%s reason=no_code", label, key)
            unacc.append(
                {
                    "instance": label,
                    "project_key": key,
                    "prefix": prefix,
                    "svc_guess": svc_guess,
                    "code": "",
                    "service_name": "",
                    "activity_code": "",
                    "activity_name": "",
                    "tasks_total": 0,
                    "total_lines": 0,
                    "reason": "no_code",
                    "detail": "cannot parse service_id from project key prefix",
                }
            )
            continue

        if code in {str(x) for x in BAN_SERVICE_IDS}:
            unaccounted_projects += 1
            meta = activity.get(code, {})
            logger.warning("[%s] unaccounted project=%s code=%s reason=banned_service_id", label, key, code)
            unacc.append(
                {
                    "instance": label,
                    "project_key": key,
                    "prefix": prefix,
                    "svc_guess": svc_guess,
                    "code": code,
                    "service_name": meta.get("service_name", ""),
                    "activity_code": meta.get("activity_code", ""),
                    "activity_name": meta.get("activity_name", ""),
                    "tasks_total": 0,
                    "total_lines": 0,
                    "reason": "banned_service_id",
                    "detail": "code in BAN_SERVICE_IDS",
                }
            )
            continue

        meta = activity.get(code)
        if not meta:
            unaccounted_projects += 1
            logger.warning("[%s] unaccounted project=%s code=%s reason=activity_mapping_miss", label, key, code)
            unacc.append(
                {
                    "instance": label,
                    "project_key": key,
                    "prefix": prefix,
                    "svc_guess": svc_guess,
                    "code": code,
                    "service_name": "",
                    "activity_code": "",
                    "activity_name": "",
                    "tasks_total": 0,
                    "total_lines": 0,
                    "reason": "activity_mapping_miss",
                    "detail": "code not found in activity.xlsx",
                }
            )
            continue

        try:
            tasks = get_tasks(s, url, key, label)
            tcnt, lines = calc_project_tasks_lines(
                s, url, key, tasks, ncloc_cache, new_cache
            )
        except Exception as e:
            unaccounted_projects += 1
            logger.exception("[%s] unaccounted project=%s code=%s reason=metrics_failed err=%s", label, key, code, e)
            unacc.append(
                {
                    "instance": label,
                    "project_key": key,
                    "prefix": prefix,
                    "svc_guess": meta.get("service_name", svc_guess),
                    "code": code,
                    "service_name": meta.get("service_name", ""),
                    "activity_code": meta.get("activity_code", ""),
                    "activity_name": meta.get("activity_name", ""),
                    "tasks_total": 0,
                    "total_lines": 0,
                    "reason": "metrics_failed",
                    "detail": str(e),
                }
            )
            continue

        accounted_projects += 1
        logger.info(
            '[%s] accounted project=%s service="%s" code=%s activity_code="%s" tasks=%d lines=%d',
            label,
            key,
            meta.get("service_name", ""),
            code,
            meta.get("activity_code", ""),
            tcnt,
            lines,
        )

        acc.setdefault(
            code,
            {
                "service_name": meta.get("service_name", ""),
                "code": code,
                "activity_code": meta.get("activity_code", ""),
                "activity_name": meta.get("activity_name", ""),
                "tasks_total": 0,
                "total_lines": 0,
            },
        )

        acc[code]["tasks_total"] += tcnt
        acc[code]["total_lines"] += lines

    logger.info(
        "[%s] done: accounted_projects=%d unaccounted_projects=%d services_accumulated=%d",
        label,
        accounted_projects,
        unaccounted_projects,
        len(acc),
    )


def write_xlsx(rows, unacc):
    wb = Workbook()

    ws = wb.active
    ws.title = "Отчет SonarQube"

    headers = [
        "Имя сервиса",
        "Код",
        "Код активности",
        "Наименование активности",
        "Кол-во тасок",
        "Строки",
        "% потребления",
    ]
    ws.append(headers)

    for c in ws[1]:
        c.font = Font(bold=True)

    total = sum(r["total_lines"] for r in rows) or 1

    for r in rows:
        ws.append(
            [
                r["service_name"],
                r["code"],
                r["activity_code"],
                r["activity_name"],
                int(r["tasks_total"]),
                int(r["total_lines"]),
                r["total_lines"] / total,
            ]
        )

    pct_col = headers.index("% потребления") + 1
    for rr in range(2, ws.max_row + 1):
        ws.cell(row=rr, column=pct_col).number_format = "0.0000%"

    ws2 = wb.create_sheet("Unaccounted")
    headers2 = [
        "instance",
        "project_key",
        "prefix",
        "svc_guess",
        "code",
        "service_name",
        "activity_code",
        "activity_name",
        "tasks_total",
        "total_lines",
        "reason",
        "detail",
    ]
    ws2.append(headers2)

    for c in ws2[1]:
        c.font = Font(bold=True)

    for r in unacc:
        ws2.append(
            [
                r.get("instance", ""),
                r.get("project_key", ""),
                r.get("prefix", ""),
                r.get("svc_guess", ""),
                r.get("code", ""),
                r.get("service_name", ""),
                r.get("activity_code", ""),
                r.get("activity_name", ""),
                int(r.get("tasks_total", 0) or 0),
                int(r.get("total_lines", 0) or 0),
                r.get("reason", ""),
                r.get("detail", ""),
            ]
        )

    wb.save(OUT_FILE)
    logger.info("file saved: %s", OUT_FILE)
    logger.info(
        "totals: accounted_lines=%d unaccounted_lines=%d accounted_services=%d unaccounted_rows=%d",
        sum(r["total_lines"] for r in rows),
        sum(int(r.get("total_lines", 0) or 0) for r in unacc),
        len(rows),
        len(unacc),
    )


def main():
    if not SONAR_URL or not SONAR_TOKEN:
        raise RuntimeError("SONAR_URL / SONAR_TOKEN не заданы")

    activity = load_activity(ACTIVITY_FILE)

    services = {}
    unaccounted = []

    process_sonar("SONAR1", SONAR_URL, SONAR_TOKEN, activity, services, unaccounted)
    process_sonar("SONAR2", SONAR2_URL, SONAR2_TOKEN, activity, services, unaccounted)

    rows = sorted(services.values(), key=lambda x: x["total_lines"], reverse=True)
    unaccounted_sorted = sorted(
        unaccounted, key=lambda x: int(x.get("total_lines", 0) or 0), reverse=True
    )

    write_xlsx(rows, unaccounted_sorted)


if __name__ == "__main__":
    main()