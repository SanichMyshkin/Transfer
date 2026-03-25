# main.py
import os
import logging
import re
import urllib3

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from humanfriendly import format_size

from nexus_sizes import (
    get_repository_data,
    get_repository_sizes,
    get_raw_top_folder_sizes,
)
from confluence_names import confluence_table_as_dicts, repo_to_service_map


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()

ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")
OUT_FILE = os.getenv("OUT_FILE", "nexus_report.xlsx")

SKIP_EMPTY_SERVICE = True

BAN_SERVICE_CODES = [
    15473,
]

KIMB_REPO = "kimb-dependencies"


def clean_spaces(s):
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def split_service_and_code(raw_service):
    s = clean_spaces(raw_service)
    if not s or s in {"-", "—"}:
        return "", ""

    parts = s.split("-")
    if len(parts) >= 2 and parts[-1].isdigit():
        return "-".join(parts[:-1]), parts[-1]

    m = re.search(r"(\d+)$", s)
    if m:
        code = m.group(1)
        name = s[: -len(code)].rstrip("-").strip()
        if name in {"-", "—"}:
            name = ""
        return name, code

    return s, ""


def normalize_number(x):
    s = str(x or "").strip()
    if not s:
        return ""
    m = re.search(r"(\d+)", s)
    return m.group(1) if m else ""


def to_int_bytes(x):
    if x is None:
        return 0
    return int(x)


def build_ban_set(ban_list):
    return {str(x).strip() for x in ban_list if str(x).strip()}


BAN_SET = build_ban_set(BAN_SERVICE_CODES)


def read_activity_map(path):
    if not path or not os.path.exists(path):
        raise RuntimeError(f"ACTIVITY_FILE не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    activity = {}
    for row in ws.iter_rows(values_only=True):
        code = normalize_number(row[0] if len(row) > 0 else "")
        if not code:
            continue

        if code in activity:
            continue

        service_name = clean_spaces(str(row[1] or "")) if len(row) > 1 else ""
        activity_code = clean_spaces(str(row[2] or "")) if len(row) > 2 else ""
        activity_name = clean_spaces(str(row[3] or "")) if len(row) > 3 else ""

        activity[code] = {
            "service_name": service_name,
            "activity_code": activity_code,
            "activity_name": activity_name,
        }

    wb.close()
    logging.info("activity loaded: %d", len(activity))
    return activity


def write_excel(path, rows, unaccounted_rows):
    wb = Workbook()

    ws = wb.active
    ws.title = "Отчет Nexus"

    header = [
        "Имя сервиса",
        "Код",
        "Код активности",
        "Наименование активности",
        "Объем",
        "% потребления",
    ]
    ws.append(header)

    bold = Font(bold=True)
    for i in range(1, len(header) + 1):
        ws.cell(row=1, column=i).font = bold

    for r in rows:
        ws.append(
            [
                r["service_name"],
                r["code"],
                r["activity_code"],
                r["activity_name"],
                r["size_human"],
                r["percent"],
            ]
        )

    ws2 = wb.create_sheet("Unaccounted")
    header2 = [
        "scope",
        "repo",
        "folder",
        "raw_service",
        "base_name",
        "code",
        "service_name",
        "activity_code",
        "activity_name",
        "bytes",
        "size_human",
        "reason",
        "detail",
    ]
    ws2.append(header2)
    for i in range(1, len(header2) + 1):
        ws2.cell(row=1, column=i).font = bold

    for u in unaccounted_rows:
        b = int(u.get("bytes") or 0)
        ws2.append(
            [
                u.get("scope", ""),
                u.get("repo", ""),
                u.get("folder", ""),
                u.get("raw_service", ""),
                u.get("base_name", ""),
                u.get("code", ""),
                u.get("service_name", ""),
                u.get("activity_code", ""),
                u.get("activity_name", ""),
                b,
                format_size(b, binary=True) if b else "",
                u.get("reason", ""),
                u.get("detail", ""),
            ]
        )

    wb.save(path)


def _add_to_totals(totals, code, base_name, size_bytes):
    if code not in totals:
        totals[code] = {"size_bytes": 0, "base_name": base_name}
    totals[code]["size_bytes"] += to_int_bytes(size_bytes)
    if base_name and len(base_name) > len(totals[code]["base_name"] or ""):
        totals[code]["base_name"] = base_name


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )

    conf_url = os.getenv("CONF_URL", "").strip()
    conf_page_id = os.getenv("CONF_PAGE_ID", "").strip()
    conf_user = os.getenv("CONF_USER", "").strip()
    conf_pass = os.getenv("CONF_PASS", "").strip()

    if not conf_url or not conf_page_id or not conf_user or not conf_pass:
        raise RuntimeError("Нужны CONF_URL, CONF_PAGE_ID, CONF_USER, CONF_PASS")

    def add_unaccounted(
        scope,
        repo,
        folder,
        raw_service,
        base_name,
        code,
        bytes_,
        reason,
        detail,
        service_name="",
        activity_code="",
        activity_name="",
    ):
        unaccounted.append(
            {
                "scope": scope,
                "repo": repo or "",
                "folder": folder or "",
                "raw_service": raw_service or "",
                "base_name": base_name or "",
                "code": code or "",
                "service_name": service_name or "",
                "activity_code": activity_code or "",
                "activity_name": activity_name or "",
                "bytes": int(bytes_ or 0),
                "reason": reason,
                "detail": detail,
            }
        )

    logging.info("Читаю таблицу из Confluence")
    conf_rows = confluence_table_as_dicts(conf_url, conf_page_id, conf_user, conf_pass)
    repo_service = repo_to_service_map(conf_rows)

    logging.info("Читаю activity")
    activity_map = read_activity_map(ACTIVITY_FILE)

    logging.info("Читаю репозитории из БД")
    repo_data = get_repository_data()

    logging.info("Считаю размеры репозиториев из БД")
    repo_sizes = get_repository_sizes()

    totals = {}
    unaccounted = []

    hosted_total = 0
    skipped_no_service = 0
    skipped_no_code = 0
    skipped_ban_service_code = 0

    for r in repo_data:
        if (r.get("repository_type") or "").strip().lower() != "hosted":
            continue

        hosted_total += 1
        repo_name = r["repository_name"]

        if repo_name == KIMB_REPO:
            logging.info(
                "map repo=%s -> SPECIAL (split by top-level folders; treat as pseudo-repos)",
                repo_name,
            )

            folder_sizes = get_raw_top_folder_sizes(repo_name)
            logging.info(
                "special-case repo=%s: folders=%d", repo_name, len(folder_sizes)
            )

            base_to_codes = {}
            for folder in folder_sizes.keys():
                base_name, code = split_service_and_code(folder)
                base_key = clean_spaces(base_name).lower()
                if base_key and code:
                    base_to_codes.setdefault(base_key, set()).add(code)

            for folder, size_bytes in folder_sizes.items():
                base_name, code = split_service_and_code(folder)

                if not code:
                    continue

                logging.info(
                    "map special: repo=%s -> folder=%s -> base=%s code=%s bytes=%d action=ADD",
                    repo_name,
                    folder,
                    base_name,
                    code,
                    int(size_bytes or 0),
                )

                if SKIP_EMPTY_SERVICE and (not base_name):
                    skipped_no_service += 1
                    logging.info(
                        "map special: folder=%s action=SKIP reason=no_service", folder
                    )
                    add_unaccounted(
                        scope="folder",
                        repo=repo_name,
                        folder=folder,
                        raw_service="",
                        base_name=base_name,
                        code=code,
                        bytes_=size_bytes,
                        reason="no_service",
                        detail="SKIP_EMPTY_SERVICE=True and base_name is empty",
                    )
                    continue

                if code in BAN_SET:
                    skipped_ban_service_code += 1
                    logging.info(
                        "map special: folder=%s action=SKIP reason=ban_service_code code=%s",
                        folder,
                        code,
                    )
                    meta = activity_map.get(code, {})
                    add_unaccounted(
                        scope="folder",
                        repo=repo_name,
                        folder=folder,
                        raw_service="",
                        base_name=base_name,
                        code=code,
                        bytes_=size_bytes,
                        reason="ban_service_code",
                        detail="code in BAN_SERVICE_CODES",
                        service_name=meta.get("service_name", ""),
                        activity_code=meta.get("activity_code", ""),
                        activity_name=meta.get("activity_name", ""),
                    )
                    continue

                _add_to_totals(totals, code, base_name, size_bytes)

            for folder, size_bytes in folder_sizes.items():
                base_name, code = split_service_and_code(folder)
                if code:
                    continue

                base_key = clean_spaces(base_name).lower()
                codes = list(base_to_codes.get(base_key, set()))

                if len(codes) == 1:
                    alias_code = codes[0]
                    logging.info(
                        "map special: repo=%s -> folder=%s -> base=%s code=%s bytes=%d action=ALIAS_ADD",
                        repo_name,
                        folder,
                        base_name,
                        alias_code,
                        int(size_bytes or 0),
                    )
                    _add_to_totals(totals, alias_code, base_name, size_bytes)
                    continue

                if len(codes) > 1:
                    skipped_no_code += 1
                    logging.info(
                        "map special: repo=%s -> folder=%s -> base=%s bytes=%d action=SKIP reason=ambiguous_alias codes=%s",
                        repo_name,
                        folder,
                        base_name,
                        int(size_bytes or 0),
                        ",".join(sorted(codes)),
                    )
                    add_unaccounted(
                        scope="folder",
                        repo=repo_name,
                        folder=folder,
                        raw_service="",
                        base_name=base_name,
                        code="",
                        bytes_=size_bytes,
                        reason="ambiguous_alias",
                        detail=f"multiple possible codes: {','.join(sorted(codes))}",
                    )
                    continue

                skipped_no_code += 1
                logging.info(
                    "map special: repo=%s -> folder=%s -> base=%s bytes=%d action=SKIP reason=no_code",
                    repo_name,
                    folder,
                    base_name,
                    int(size_bytes or 0),
                )
                add_unaccounted(
                    scope="folder",
                    repo=repo_name,
                    folder=folder,
                    raw_service="",
                    base_name=base_name,
                    code="",
                    bytes_=size_bytes,
                    reason="no_code",
                    detail="cannot extract code from folder name and no alias match",
                )

            continue

        raw_service = repo_service.get(repo_name, "")
        base_name, code = split_service_and_code(raw_service)
        size_bytes = to_int_bytes(repo_sizes.get(repo_name))

        if SKIP_EMPTY_SERVICE and (not base_name):
            skipped_no_service += 1
            logging.info(
                "map repo=%s -> raw_service=%s -> base=%s code=%s bytes=%d action=SKIP reason=no_service",
                repo_name,
                raw_service,
                base_name,
                code,
                size_bytes,
            )
            add_unaccounted(
                scope="repo",
                repo=repo_name,
                folder="",
                raw_service=raw_service,
                base_name=base_name,
                code=code,
                bytes_=size_bytes,
                reason="no_service",
                detail="SKIP_EMPTY_SERVICE=True and base_name is empty",
            )
            continue

        if not code:
            skipped_no_code += 1
            logging.info(
                "map repo=%s -> raw_service=%s -> base=%s code=%s bytes=%d action=SKIP reason=no_code",
                repo_name,
                raw_service,
                base_name,
                code,
                size_bytes,
            )
            add_unaccounted(
                scope="repo",
                repo=repo_name,
                folder="",
                raw_service=raw_service,
                base_name=base_name,
                code="",
                bytes_=size_bytes,
                reason="no_code",
                detail="cannot extract code from raw_service (Confluence mapping)",
            )
            continue

        if code in BAN_SET:
            skipped_ban_service_code += 1
            logging.info(
                "map repo=%s -> raw_service=%s -> base=%s code=%s bytes=%d action=SKIP reason=ban_service_code",
                repo_name,
                raw_service,
                base_name,
                code,
                size_bytes,
            )
            meta = activity_map.get(code, {})
            add_unaccounted(
                scope="repo",
                repo=repo_name,
                folder="",
                raw_service=raw_service,
                base_name=base_name,
                code=code,
                bytes_=size_bytes,
                reason="ban_service_code",
                detail="code in BAN_SERVICE_CODES",
                service_name=meta.get("service_name", ""),
                activity_code=meta.get("activity_code", ""),
                activity_name=meta.get("activity_name", ""),
            )
            continue

        logging.info(
            "map repo=%s -> raw_service=%s -> base=%s code=%s bytes=%d action=ADD",
            repo_name,
            raw_service,
            base_name,
            code,
            size_bytes,
        )
        _add_to_totals(totals, code, base_name, size_bytes)

    candidates = []
    skipped_activity_miss = 0

    for code, v in totals.items():
        size_bytes = v["size_bytes"]
        base_name = v["base_name"]

        meta = activity_map.get(code, {})
        service_name = meta.get("service_name", "") or base_name
        activity_code = meta.get("activity_code", "")
        activity_name = meta.get("activity_name", "")

        if code not in activity_map:
            skipped_activity_miss += 1
            add_unaccounted(
                scope="service",
                repo="",
                folder="",
                raw_service="",
                base_name=base_name,
                code=code,
                bytes_=size_bytes,
                reason="activity_mapping_miss",
                detail="service_id отсутствует в activity.xlsx",
                service_name=service_name,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        candidates.append(
            {
                "service_name": service_name,
                "code": code,
                "activity_code": activity_code,
                "activity_name": activity_name,
                "size_bytes": size_bytes,
            }
        )

    eligible_total = sum(x["size_bytes"] for x in candidates)

    rows = []
    for x in candidates:
        size_bytes = x["size_bytes"]
        percent = 0.0
        if eligible_total > 0:
            percent = (size_bytes / eligible_total) * 100.0

        rows.append(
            {
                "service_name": x["service_name"],
                "code": x["code"],
                "activity_code": x["activity_code"],
                "activity_name": x["activity_name"],
                "size_bytes": size_bytes,
                "size_human": format_size(size_bytes, binary=True),
                "percent": round(percent, 4),
            }
        )

    rows.sort(key=lambda x: x["size_bytes"], reverse=True)

    logging.info("hosted repos: %d", hosted_total)
    logging.info("skipped without service: %d", skipped_no_service)
    logging.info("skipped without code: %d", skipped_no_code)
    logging.info("skipped by service code ban: %d", skipped_ban_service_code)
    logging.info("skipped by activity miss: %d", skipped_activity_miss)
    logging.info("services in report: %d", len(rows))
    logging.info("eligible_total: %s", format_size(eligible_total, binary=True))
    logging.info("unaccounted rows: %d", len(unaccounted))
    logging.info("write excel: %s", OUT_FILE)

    write_excel(OUT_FILE, rows, unaccounted)

    logging.info("done")


if __name__ == "__main__":
    main()