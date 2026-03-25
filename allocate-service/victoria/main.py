import os
import sys
import time
import logging
from datetime import datetime, timezone
import re
from collections import defaultdict

import requests
import urllib3
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import pandas as pd

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

load_dotenv()

OUTPUT_FILE = os.getenv("OUT_FILE", "victoria_report.xlsx")
HTTP_TIMEOUT_SEC = 30
SLEEP_SEC = 0.01

BAN_TEAMS = []
BAN_SERVICE_IDS = [15473]

TEAM_SERVICE_ID_OVERRIDES = {
    "service": "1234"
}

EXCLUDE_NO_SERVICE_ID_AT_QUERY = False
EXTRAPOLATION_DAYS = 90

ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

TEAM_TAIL_ID_RE = re.compile(r"^(.*)-(\d+)$")

WINDOW_HOURS = 24
WINDOW_DAYS = WINDOW_HOURS / 24.0

REPORT_COLS = [
    "Имя сервиса",
    "Код",
    "Код активности",
    "Наименование активности",
    "samples_24h",
    "эксрополяция",
    "% от общего числа",
]

UNACC_COLS = [
    "stage",
    "team",
    "service_id",
    "service_name",
    "activity_code",
    "activity_name",
    "metric",
    "samples_24h",
    "reason",
    "detail",
]


def build_ban_set(ban_list):
    if not isinstance(ban_list, (list, tuple, set)):
        raise SystemExit("BAN_SERVICE_IDS должен быть list / tuple / set")
    return {str(x).strip() for x in ban_list if str(x).strip()}


ban_service_set = build_ban_set(BAN_SERVICE_IDS)


def clean_spaces(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def is_all_zeros(s: str) -> bool:
    s = (s or "").strip()
    return bool(s) and set(s) == {"0"}


def normalize_sid(sid: str) -> str:
    sid = (sid or "").strip()
    if not sid:
        return ""
    if is_all_zeros(sid):
        return ""
    return sid


def split_team_tail_id(team: str):
    team = (team or "").strip()
    m = TEAM_TAIL_ID_RE.match(team)
    if not m:
        return team, ""
    base = m.group(1).strip()
    tail_id = m.group(2)
    return (base if base else team), tail_id


def sid_rank(sid: str) -> int:
    sid = normalize_sid(sid)
    if not sid:
        return 0
    if sid.isdigit():
        return 3
    return 0


def pick_better_sid(a: str, b: str) -> str:
    a = normalize_sid(a)
    b = normalize_sid(b)
    return b if sid_rank(b) > sid_rank(a) else a


def http_query(vm_url: str, query: str, at_ts: float | None = None):
    url = vm_url.rstrip("/") + "/api/v1/query"
    log.info("QUERY: %s", query)
    params = {"query": query}
    if at_ts is not None:
        params["time"] = at_ts
    r = requests.get(url, params=params, verify=False, timeout=HTTP_TIMEOUT_SEC)
    r.raise_for_status()
    data = r.json()
    if data.get("status") != "success":
        raise RuntimeError(data)
    return data["data"]["result"]


def label(metric: dict, key: str) -> str:
    v = metric.get(key)
    return "" if v is None else str(v).strip()


def is_banned_team(team_base: str) -> bool:
    t = "" if team_base is None else str(team_base).strip()
    if t == "":
        return False
    for x in BAN_TEAMS:
        if x is None:
            continue
        if t == str(x).strip():
            return True
    return False


def read_activity_map(path: str) -> pd.DataFrame:
    if not path or not os.path.exists(path):
        log.warning("ACTIVITY_FILE не найден: %s", path)
        return pd.DataFrame(columns=["code", "service_name", "activity_code", "activity_name"])

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    rows = []
    for row in ws.iter_rows(values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else None)
        if not code:
            continue

        rows.append(
            {
                "code": code,
                "service_name": clean_spaces(row[1] if len(row) > 1 else ""),
                "activity_code": clean_spaces(row[2] if len(row) > 2 else ""),
                "activity_name": clean_spaces(row[3] if len(row) > 3 else ""),
            }
        )

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(columns=["code", "service_name", "activity_code", "activity_name"])

    out = out.drop_duplicates(subset=["code"], keep="first").copy()
    log.info("ACTIVITY loaded: %d", len(out))
    return out


def discover_series(vm_url: str):
    out = []
    seen = set()

    queries = [
        'count by (team, service_id, __name__) ({team=~".+", service_id=~".+"})',
        'count by (team, service_id, __name__) ({team!~".+", service_id=~".+"})',
        'count by (team, service_id, __name__) ({team=~".+", service_id!~".+"})',
        'count by (team, service_id, __name__) ({team!~".+", service_id!~".+"})',
    ]

    for q in queries:
        rows = http_query(vm_url, q)
        for r in rows or []:
            m = r.get("metric", {}) or {}
            team_raw = label(m, "team")
            service_id_raw = normalize_sid(label(m, "service_id"))
            metric = label(m, "__name__")
            if not metric:
                continue

            team_base, sid_from_team = split_team_tail_id(team_raw)
            sid_from_team = normalize_sid(sid_from_team)
            sid_seed = pick_better_sid(service_id_raw, sid_from_team)

            key = (team_raw, service_id_raw, metric)
            if key in seen:
                continue
            seen.add(key)

            out.append(
                {
                    "team_raw": team_raw,
                    "team_base": (team_base or "").strip(),
                    "service_id_raw": service_id_raw,
                    "sid_from_team": sid_from_team,
                    "sid_seed": sid_seed,
                    "metric": metric,
                }
            )

        time.sleep(SLEEP_SEC)

    return out


def build_team_to_sid_maps(series_rows):
    team_sids = defaultdict(set)
    for r in series_rows:
        team_base = (r.get("team_base") or "").strip()
        sid = normalize_sid(r.get("sid_seed"))
        if team_base and sid:
            team_sids[team_base].add(sid)

    team_to_sid = {}
    ambiguous_teams = set()

    for team_base, sids in team_sids.items():
        if len(sids) == 1:
            team_to_sid[team_base] = next(iter(sids))
        elif len(sids) > 1:
            ambiguous_teams.add(team_base)

    return team_to_sid, ambiguous_teams


def build_matchers_raw(team_raw: str, service_id_raw: str, metric_name: str) -> str:
    parts = []
    if (team_raw or "").strip():
        parts.append(f'team="{team_raw}"')
    else:
        parts.append('team!~".+"')

    service_id_raw = normalize_sid(service_id_raw)
    if service_id_raw:
        parts.append(f'service_id="{service_id_raw}"')
    else:
        parts.append('service_id!~".+"')

    parts.append(f'__name__="{metric_name}"')
    return ", ".join(parts)


def samples_24h_for_series(
    vm_url: str, metric_name: str, team_raw: str, service_id_raw: str, end_dt: datetime
) -> int:
    m = build_matchers_raw(team_raw, service_id_raw, metric_name)
    q = f"sum(count_over_time({{{m}}}[{WINDOW_HOURS}h]))"
    res = http_query(vm_url, q, at_ts=end_dt.timestamp())
    if not res:
        return 0
    v = res[0].get("value")
    if not isinstance(v, list) or len(v) < 2:
        return 0
    try:
        return int(float(v[1]))
    except Exception:
        return 0


def aggregate_to_group(metric_rows):
    acc = {}
    for r in metric_rows:
        team_base = (r.get("team_base") or "").strip()
        sid = normalize_sid(r.get("service_id_final"))
        samples = int(r.get("samples_24h", 0) or 0)

        key = (team_base, sid)
        if key not in acc:
            acc[key] = {"team_base": team_base, "service_id": sid, "samples_24h": 0}
        acc[key]["samples_24h"] += samples

    out = []
    for (_, _), v in acc.items():
        s24 = int(v["samples_24h"])
        extrap = (
            int(round((s24 / WINDOW_DAYS) * EXTRAPOLATION_DAYS)) if WINDOW_DAYS else 0
        )
        out.append(
            {
                "team": v["team_base"],
                "service_id": v["service_id"],
                "samples_24h": s24,
                "extrapolation": extrap,
            }
        )
    return out


def enrich_group_rows(group_rows, activity_df: pd.DataFrame):
    df = pd.DataFrame(group_rows)
    if df.empty:
        return pd.DataFrame(
            columns=[
                "team",
                "service_id",
                "code",
                "service_name",
                "activity_code",
                "activity_name",
                "samples_24h",
                "эксрополяция",
                "activity_found",
            ]
        )

    df["service_id"] = df["service_id"].astype(str).fillna("").map(normalize_sid)
    df["code"] = df["service_id"].astype(str)

    out = df.merge(activity_df, left_on="code", right_on="code", how="left")
    out["activity_found"] = (
        out["service_name"].fillna("").astype(str).map(lambda x: x.strip() != "")
    )

    out["service_name"] = out["service_name"].fillna("").astype(str)
    out.loc[out["service_name"] == "", "service_name"] = out["team"]

    out["activity_code"] = out["activity_code"].fillna("").astype(str).map(clean_spaces)
    out["activity_name"] = out["activity_name"].fillna("").astype(str).map(clean_spaces)

    out = out.rename(columns={"extrapolation": "эксрополяция"})
    out["samples_24h"] = out["samples_24h"].fillna(0).astype(int)
    out["эксрополяция"] = out["эксрополяция"].fillna(0).astype(int)

    return out[
        [
            "team",
            "service_id",
            "code",
            "service_name",
            "activity_code",
            "activity_name",
            "samples_24h",
            "эксрополяция",
            "activity_found",
        ]
    ]


def _first_non_empty(vals):
    for v in vals:
        s = "" if v is None else str(v).strip()
        if s:
            return s
    return ""


def dedupe_and_add_percent(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        out = df.copy()
        out["% от общего числа"] = []
        return out.reindex(columns=REPORT_COLS)

    df = df.copy()
    df["Код"] = df["Код"].fillna("").astype(str).map(normalize_sid)

    has_code = df[df["Код"] != ""].copy()
    no_code = df[df["Код"] == ""].copy()

    if not has_code.empty:
        has_code = has_code.groupby("Код", as_index=False).agg(
            {
                "Имя сервиса": _first_non_empty,
                "Код активности": _first_non_empty,
                "Наименование активности": _first_non_empty,
                "samples_24h": "sum",
                "эксрополяция": "sum",
            }
        )

    if not no_code.empty:
        no_code = no_code.groupby("Имя сервиса", as_index=False).agg(
            {
                "Код": _first_non_empty,
                "Код активности": _first_non_empty,
                "Наименование активности": _first_non_empty,
                "samples_24h": "sum",
                "эксрополяция": "sum",
            }
        )
        no_code = no_code[
            [
                "Имя сервиса",
                "Код",
                "Код активности",
                "Наименование активности",
                "samples_24h",
                "эксрополяция",
            ]
        ]

    out = pd.concat([has_code, no_code], ignore_index=True)

    total = float(out["samples_24h"].sum()) if "samples_24h" in out.columns else 0.0
    out["% от общего числа"] = (out["samples_24h"] / total * 100.0) if total else 0.0
    out["% от общего числа"] = out["% от общего числа"].round(5)

    return out.reindex(columns=REPORT_COLS)


def write_report(df_report: pd.DataFrame, df_unacc: pd.DataFrame):
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "samples_24h"
    df_report = df_report.reindex(columns=REPORT_COLS)
    ws.append(list(df_report.columns))
    for c in ws[1]:
        c.font = bold
    for row in df_report.itertuples(index=False):
        ws.append(list(row))
    if "% от общего числа" in df_report.columns:
        col_idx = df_report.columns.get_loc("% от общего числа") + 1
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                c.number_format = "0.00000"

    ws2 = wb.create_sheet("Unaccounted")
    if df_unacc is None or df_unacc.empty:
        ws2.append(["No data"])
        ws2["A1"].font = bold
    else:
        df_unacc = df_unacc.reindex(columns=UNACC_COLS)
        ws2.append(list(df_unacc.columns))
        for c in ws2[1]:
            c.font = bold
        for row in df_unacc.itertuples(index=False):
            ws2.append(list(row))

    wb.save(OUTPUT_FILE)


def main():
    vm_url = os.getenv("VM_URL", "").strip()
    if not vm_url:
        log.error("VM_URL не задан")
        sys.exit(1)

    log.info("VM_URL=%s", vm_url)
    log.info("WINDOW_HOURS=%s", WINDOW_HOURS)
    log.info("EXCLUDE_NO_SERVICE_ID_AT_QUERY=%s", EXCLUDE_NO_SERVICE_ID_AT_QUERY)
    log.info("EXTRAPOLATION_DAYS=%s", EXTRAPOLATION_DAYS)
    log.info("BAN_SERVICE_IDS=%s", sorted(ban_service_set) if ban_service_set else "[]")
    log.info("ACTIVITY_FILE=%s", ACTIVITY_FILE)

    activity_df = read_activity_map(ACTIVITY_FILE)

    overrides = {}
    for k, v in (TEAM_SERVICE_ID_OVERRIDES or {}).items():
        kk = clean_spaces(k)
        vv = normalize_sid(v)
        if kk and vv:
            overrides[kk] = vv

    unacc_map = {}

    def add_unacc_once(
        stage, team, service_id, metric, reason, detail, samples_24h=None
    ):
        key = (team or "", service_id or "", metric or "")
        if key in unacc_map:
            return

        sid = normalize_sid(service_id)
        service_name = ""
        activity_code = ""
        activity_name = ""

        if sid and not activity_df.empty:
            match = activity_df[activity_df["code"].astype(str) == sid]
            if not match.empty:
                first = match.iloc[0]
                service_name = clean_spaces(first.get("service_name", ""))
                activity_code = clean_spaces(first.get("activity_code", ""))
                activity_name = clean_spaces(first.get("activity_name", ""))

        unacc_map[key] = {
            "stage": stage,
            "team": team or "",
            "service_id": sid,
            "service_name": service_name,
            "activity_code": activity_code,
            "activity_name": activity_name,
            "metric": metric or "",
            "samples_24h": "" if samples_24h is None else int(samples_24h),
            "reason": reason,
            "detail": detail,
        }

    log.info("Discover series ...")
    series_rows = discover_series(vm_url)
    log.info("Series found: %d", len(series_rows))

    team_to_sid_map, ambiguous_teams = build_team_to_sid_maps(series_rows)
    log.info("Team->SID inferred map size: %d", len(team_to_sid_map))
    log.info("Ambiguous teams: %d", len(ambiguous_teams))
    log.info("Overrides: %d", len(overrides))

    end_dt = datetime.now(timezone.utc)

    metrics_audit = []
    accounted_metric_rows = []

    for idx, r in enumerate(series_rows, 1):
        team_raw = (r.get("team_raw") or "").strip()
        team_base = (r.get("team_base") or "").strip()
        service_id_raw = normalize_sid(r.get("service_id_raw"))
        sid_from_team = normalize_sid(r.get("sid_from_team"))
        sid_seed = normalize_sid(r.get("sid_seed"))
        metric = (r.get("metric") or "").strip()

        if not metric:
            continue

        service_id_final = sid_seed
        stage = ""
        reason = ""
        detail = ""
        status = "accounted"

        ov = overrides.get(team_base)
        if ov:
            service_id_final = ov
        else:
            if not service_id_final:
                if (
                    (not service_id_raw)
                    and (not sid_from_team)
                    and (team_base in ambiguous_teams)
                ):
                    status = "unaccounted"
                    stage = "infer"
                    reason = "ambiguous_service_id"
                    detail = "multiple service_id detected for team_base"
                else:
                    inferred = team_to_sid_map.get(team_base, "")
                    if (not service_id_raw) and (not sid_from_team) and inferred:
                        service_id_final = inferred

        if is_banned_team(team_base):
            status = "unaccounted"
            stage = "discover"
            reason = "banned_team"
            detail = "team in BAN_TEAMS"

        if (
            status == "accounted"
            and service_id_final
            and service_id_final in ban_service_set
        ):
            status = "unaccounted"
            stage = "discover"
            reason = "banned_service_id"
            detail = "service_id in BAN_SERVICE_IDS"

        if (
            status == "accounted"
            and EXCLUDE_NO_SERVICE_ID_AT_QUERY
            and not service_id_final
        ):
            status = "unaccounted"
            stage = "discover"
            reason = "excluded_no_service_id"
            detail = "EXCLUDE_NO_SERVICE_ID_AT_QUERY=True and service_id empty"

        s24 = 0
        try:
            s24 = samples_24h_for_series(vm_url, metric, team_raw, service_id_raw, end_dt)
        except Exception as e:
            add_unacc_once(
                "samples",
                team_base,
                service_id_final,
                metric,
                "samples_failed",
                str(e),
                samples_24h=None,
            )
            s24 = 0

        if status == "unaccounted":
            add_unacc_once(
                stage,
                team_base,
                service_id_final,
                metric,
                reason,
                detail,
                samples_24h=s24,
            )
        else:
            accounted_metric_rows.append(
                {
                    "team_base": team_base,
                    "service_id_final": service_id_final,
                    "samples_24h": int(s24),
                }
            )

        metrics_audit.append(
            {
                "team_base": team_base,
                "team_raw": team_raw,
                "service_id_raw": service_id_raw,
                "service_id_final": service_id_final,
                "metric": metric,
                "samples_24h": int(s24),
                "status": status,
                "stage": stage,
                "reason": reason,
                "detail": detail,
            }
        )

        if idx % 200 == 0:
            log.info("Processed: %d/%d", idx, len(series_rows))

        time.sleep(SLEEP_SEC)

    group_rows = aggregate_to_group(accounted_metric_rows)
    enriched = enrich_group_rows(group_rows, activity_df)

    if enriched.empty:
        df_report = pd.DataFrame(columns=REPORT_COLS)
        accounted = enriched.copy()
    else:
        accounted = enriched.copy()

        def mark_unacc_for_service(team, sid, stage, reason, detail):
            for rr in metrics_audit:
                if rr.get("status") != "accounted":
                    continue
                if (rr.get("team_base") or "") == (team or "") and normalize_sid(
                    rr.get("service_id_final")
                ) == normalize_sid(sid):
                    add_unacc_once(
                        stage,
                        rr.get("team_base", ""),
                        rr.get("service_id_final", ""),
                        rr.get("metric", ""),
                        reason,
                        detail,
                        samples_24h=int(rr.get("samples_24h", 0) or 0),
                    )
                    rr["status"] = "unaccounted"
                    rr["stage"] = stage
                    rr["reason"] = reason
                    rr["detail"] = detail

        m_activity_missing = ~accounted["activity_found"]
        for rr in accounted[m_activity_missing].to_dict("records"):
            mark_unacc_for_service(
                rr.get("team", ""),
                rr.get("service_id", ""),
                "enrich",
                "activity_mapping_miss",
                "service_id not found in activity.xlsx",
            )

        accounted = accounted[accounted["activity_found"]].copy()

        df_for_report = accounted.rename(
            columns={
                "service_name": "Имя сервиса",
                "service_id": "Код",
                "activity_code": "Код активности",
                "activity_name": "Наименование активности",
            }
        )[
            [
                "Имя сервиса",
                "Код",
                "Код активности",
                "Наименование активности",
                "samples_24h",
                "эксрополяция",
            ]
        ]

        df_report = dedupe_and_add_percent(df_for_report)
        df_report = df_report.sort_values(["samples_24h"], ascending=False).reset_index(
            drop=True
        )

    df_unacc = pd.DataFrame(list(unacc_map.values()))
    if not df_unacc.empty:
        df_unacc = df_unacc.reindex(columns=UNACC_COLS).fillna("")
        df_unacc = df_unacc.sort_values(
            ["stage", "reason", "team", "service_id", "metric"]
        ).reset_index(drop=True)

    log.info("Saving report: %s", OUTPUT_FILE)
    write_report(df_report, df_unacc)
    log.info("✔ Done")


if __name__ == "__main__":
    main()