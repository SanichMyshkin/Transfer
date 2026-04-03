import os
import sys
import time
import logging
import sqlite3
from datetime import datetime, timezone, timedelta
import re
from collections import defaultdict

import requests
import urllib3
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import pandas as pd

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

load_dotenv()

OUTPUT_FILE = os.getenv("OUT_FILE", "victoria_report.xlsx")
HTTP_TIMEOUT_SEC = int(os.getenv("HTTP_TIMEOUT_SEC", "30"))
SLEEP_SEC = float(os.getenv("SLEEP_SEC", "0.1"))
VERBOSE_LOG = (os.getenv("VERBOSE_LOG", "true").strip().lower() == "true")

BAN_TEAMS = []
BAN_SERVICE_IDS = [15473]

TEAM_SERVICE_ID_OVERRIDES = {"mlops": "15518"}

ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

TEAM_TAIL_ID_RE = re.compile(r"^(.*)-(\d+)$")

WINDOW_HOURS = int(os.getenv("WINDOW_HOURS", "24"))
MAX_METRICS = int(os.getenv("MAX_METRICS", "0"))

SQLITE_DB_FILE = os.getenv("SQLITE_DB_FILE", "victoria_daily.sqlite")

SNAPSHOT_LOOKBACK_DAYS = int(os.getenv("SNAPSHOT_LOOKBACK_DAYS", "0"))

UNACC_COLS = [
    "stage",
    "team",
    "service_id",
    "service_name",
    "activity_code",
    "activity_name",
    "metric",
    "samples_value",
    "reason",
    "detail",
]


def build_ban_set(ban_list):
    if not isinstance(ban_list, (list, tuple, set)):
        raise SystemExit("BAN_SERVICE_IDS должен быть list / tuple / set")
    return {str(x).strip() for x in ban_list if str(x).strip()}


ban_service_set = build_ban_set(BAN_SERVICE_IDS)


def clean_spaces(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def is_all_zeros(s: str) -> bool:
    s = (s or "").strip()
    return bool(s) and set(s) == {"0"}


def normalize_sid(sid: str) -> str:
    sid = (sid or "").strip()
    if not sid:
        return ""
    if is_all_zeros(sid):
        return ""
    return sid


def split_team_tail_id(team: str):
    team = (team or "").strip()
    m = TEAM_TAIL_ID_RE.match(team)
    if not m:
        return team, ""
    base = m.group(1).strip()
    tail_id = m.group(2)
    return (base if base else team), tail_id


def sid_rank(sid: str) -> int:
    sid = normalize_sid(sid)
    if not sid:
        return 0
    if sid.isdigit():
        return 3
    return 0


def pick_better_sid(a: str, b: str) -> str:
    a = normalize_sid(a)
    b = normalize_sid(b)
    return b if sid_rank(b) > sid_rank(a) else a


def http_query(vm_url: str, query: str, at_ts: float | None = None):
    url = vm_url.rstrip("/") + "/api/v1/query"
    params = {"query": query}
    if at_ts is not None:
        params["time"] = at_ts

    log.info("VM query start | url=%s | time=%s | query=%s", url, at_ts, query)

    started = time.time()
    r = requests.get(url, params=params, verify=False, timeout=HTTP_TIMEOUT_SEC)
    elapsed = time.time() - started

    log.info(
        "VM query response | status_code=%s | elapsed=%.3fs | query=%s",
        r.status_code,
        elapsed,
        query,
    )

    r.raise_for_status()
    data = r.json()

    if data.get("status") != "success":
        log.error("VM query failed payload=%s", data)
        raise RuntimeError(data)

    result = data["data"]["result"]
    log.info("VM query result rows=%d | query=%s", len(result), query)
    return result


def http_label_values(vm_url: str, label_name: str):
    url = vm_url.rstrip("/") + f"/api/v1/label/{label_name}/values"
    log.info("VM label values start | label=%s | url=%s", label_name, url)

    started = time.time()
    r = requests.get(url, verify=False, timeout=HTTP_TIMEOUT_SEC)
    elapsed = time.time() - started

    log.info(
        "VM label values response | label=%s | status_code=%s | elapsed=%.3fs",
        label_name,
        r.status_code,
        elapsed,
    )

    r.raise_for_status()
    data = r.json()
    if data.get("status") != "success":
        log.error("VM label values failed | label=%s | payload=%s", label_name, data)
        raise RuntimeError(data)

    values = data.get("data") or []
    log.info("VM label values loaded | label=%s | count=%d", label_name, len(values))
    return values


def label(metric: dict, key: str) -> str:
    v = metric.get(key)
    return "" if v is None else str(v).strip()


def is_banned_team(team_base: str) -> bool:
    t = "" if team_base is None else str(team_base).strip()
    if t == "":
        return False
    for x in BAN_TEAMS:
        if x is None:
            continue
        if t == str(x).strip():
            return True
    return False


def is_valid_metric_name(name: str) -> bool:
    name = (name or "").strip()
    if not name:
        return False
    if name.startswith("vm_"):
        return False
    return True


def read_activity_map(path: str) -> pd.DataFrame:
    if not path or not os.path.exists(path):
        log.warning("ACTIVITY_FILE не найден: %s", path)
        return pd.DataFrame(
            columns=["code", "service_name", "activity_code", "activity_name"]
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    rows = []
    for row in ws.iter_rows(values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else None)
        if not code:
            continue

        rows.append(
            {
                "code": code,
                "service_name": clean_spaces(row[1] if len(row) > 1 else ""),
                "activity_code": clean_spaces(row[2] if len(row) > 2 else ""),
                "activity_name": clean_spaces(row[3] if len(row) > 3 else ""),
            }
        )

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(
            columns=["code", "service_name", "activity_code", "activity_name"]
        )

    out = out.drop_duplicates(subset=["code"], keep="first").copy()
    log.info("ACTIVITY loaded: %d", len(out))
    return out


def discover_series(vm_url: str):
    started_at = time.time()

    metric_names = http_label_values(vm_url, "__name__")
    metric_names = [x for x in metric_names if is_valid_metric_name(x)]
    metric_names.sort()

    if MAX_METRICS > 0:
        metric_names = metric_names[:MAX_METRICS]

    total_metrics = len(metric_names)
    log.info(
        "Discovery start | metric_names=%d | MAX_METRICS=%s",
        total_metrics,
        MAX_METRICS if MAX_METRICS > 0 else "ALL",
    )

    out = []
    at_ts = datetime.now(timezone.utc).timestamp()

    for idx, metric_name in enumerate(metric_names, 1):
        q = (
            f"sum by (team, service_id, __name__) "
            f"(count_over_time({metric_name}[{WINDOW_HOURS}h]))"
        )

        log.info(
            "Metric start | idx=%d/%d | metric=%s",
            idx,
            total_metrics,
            metric_name,
        )

        try:
            rows = http_query(vm_url, q, at_ts=at_ts)
        except Exception as e:
            log.warning(
                "Metric failed | idx=%d/%d | metric=%s | err=%s",
                idx,
                total_metrics,
                metric_name,
                e,
            )
            time.sleep(SLEEP_SEC)
            continue

        log.info(
            "Metric rows fetched | idx=%d/%d | metric=%s | rows=%d",
            idx,
            total_metrics,
            metric_name,
            len(rows),
        )

        rows_added = 0

        for row_idx, r in enumerate(rows or [], 1):
            m = r.get("metric", {}) or {}

            team_raw = label(m, "team")
            service_id_raw = normalize_sid(label(m, "service_id"))
            metric = label(m, "__name__") or metric_name

            v = r.get("value")
            if not isinstance(v, list) or len(v) < 2:
                samples_value = 0
            else:
                try:
                    samples_value = int(float(v[1]))
                except Exception:
                    samples_value = 0

            team_base, sid_from_team = split_team_tail_id(team_raw)
            sid_from_team = normalize_sid(sid_from_team)
            sid_seed = pick_better_sid(service_id_raw, sid_from_team)

            out.append(
                {
                    "team_raw": team_raw,
                    "team_base": (team_base or "").strip(),
                    "service_id_raw": service_id_raw,
                    "sid_from_team": sid_from_team,
                    "sid_seed": sid_seed,
                    "metric": metric,
                    "samples_value": samples_value,
                }
            )
            rows_added += 1

            if VERBOSE_LOG:
                log.info(
                    "Metric row parsed | metric=%s | row=%d/%d | team_raw=%s | team_base=%s | service_id_raw=%s | sid_from_team=%s | sid_seed=%s | samples=%s",
                    metric,
                    row_idx,
                    len(rows),
                    team_raw,
                    team_base,
                    service_id_raw,
                    sid_from_team,
                    sid_seed,
                    samples_value,
                )

        elapsed = time.time() - started_at
        rate = idx / elapsed if elapsed > 0 else 0.0

        log.info(
            "Metric done | idx=%d/%d | metric=%s | rows_added=%d | total_rows=%d | rate=%.2f m/s",
            idx,
            total_metrics,
            metric_name,
            rows_added,
            len(out),
            rate,
        )

        time.sleep(SLEEP_SEC)

    log.info("Discovery finished | total_rows=%d", len(out))
    return out


def build_team_to_sid_maps(series_rows):
    team_sids = defaultdict(set)
    for r in series_rows:
        team_base = (r.get("team_base") or "").strip()
        sid = normalize_sid(r.get("sid_seed"))
        if team_base and sid:
            team_sids[team_base].add(sid)

    team_to_sid = {}
    ambiguous_teams = set()

    for team_base, sids in team_sids.items():
        if len(sids) == 1:
            team_to_sid[team_base] = next(iter(sids))
        elif len(sids) > 1:
            ambiguous_teams.add(team_base)

    return team_to_sid, ambiguous_teams


def aggregate_to_group(metric_rows):
    acc = {}
    for r in metric_rows:
        team_base = (r.get("team_base") or "").strip()
        sid = normalize_sid(r.get("service_id_final"))
        samples = int(r.get("samples_value", 0) or 0)

        key = (team_base, sid)
        if key not in acc:
            acc[key] = {"team_base": team_base, "service_id": sid, "samples_value": 0}
        acc[key]["samples_value"] += samples

    out = []
    for _, v in acc.items():
        out.append(
            {
                "team": v["team_base"],
                "service_id": v["service_id"],
                "samples_value": int(v["samples_value"]),
            }
        )
    return out


def enrich_group_rows(group_rows, activity_df: pd.DataFrame):
    df = pd.DataFrame(group_rows)
    if df.empty:
        return pd.DataFrame(
            columns=[
                "team",
                "service_id",
                "code",
                "service_name",
                "activity_code",
                "activity_name",
                "samples_value",
                "activity_found",
            ]
        )

    df["service_id"] = df["service_id"].astype(str).fillna("").map(normalize_sid)
    df["code"] = df["service_id"].astype(str)

    out = df.merge(activity_df, left_on="code", right_on="code", how="left")
    out["activity_found"] = (
        out["service_name"].fillna("").astype(str).map(lambda x: x.strip() != "")
    )

    out["service_name"] = out["service_name"].fillna("").astype(str)
    out.loc[out["service_name"] == "", "service_name"] = out["team"]

    out["activity_code"] = out["activity_code"].fillna("").astype(str).map(clean_spaces)
    out["activity_name"] = out["activity_name"].fillna("").astype(str).map(clean_spaces)
    out["samples_value"] = out["samples_value"].fillna(0).astype(int)

    return out[
        [
            "team",
            "service_id",
            "code",
            "service_name",
            "activity_code",
            "activity_name",
            "samples_value",
            "activity_found",
        ]
    ]


def _first_non_empty(vals):
    for v in vals:
        s = "" if v is None else str(v).strip()
        if s:
            return s
    return ""


def dedupe_daily_report(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=[
                "Имя сервиса",
                "Код",
                "Код активности",
                "Наименование активности",
                "samples_value",
            ]
        )

    df = df.copy()
    df["Код"] = df["Код"].fillna("").astype(str).map(normalize_sid)
    df["Имя сервиса"] = df["Имя сервиса"].fillna("").astype(str).map(clean_spaces)
    df["Код активности"] = df["Код активности"].fillna("").astype(str).map(clean_spaces)
    df["Наименование активности"] = (
        df["Наименование активности"].fillna("").astype(str).map(clean_spaces)
    )
    df["samples_value"] = df["samples_value"].fillna(0).astype(int)

    has_code = df[df["Код"] != ""].copy()
    no_code = df[df["Код"] == ""].copy()

    out_parts = []

    if not has_code.empty:
        has_code = has_code.groupby(
            ["Код", "Код активности"],
            as_index=False,
        ).agg(
            {
                "Имя сервиса": _first_non_empty,
                "Наименование активности": _first_non_empty,
                "samples_value": "sum",
            }
        )
        out_parts.append(
            has_code[
                [
                    "Имя сервиса",
                    "Код",
                    "Код активности",
                    "Наименование активности",
                    "samples_value",
                ]
            ]
        )

    if not no_code.empty:
        no_code = no_code.groupby(
            ["Имя сервиса", "Код активности"],
            as_index=False,
        ).agg(
            {
                "Код": _first_non_empty,
                "Наименование активности": _first_non_empty,
                "samples_value": "sum",
            }
        )
        out_parts.append(
            no_code[
                [
                    "Имя сервиса",
                    "Код",
                    "Код активности",
                    "Наименование активности",
                    "samples_value",
                ]
            ]
        )

    if not out_parts:
        return pd.DataFrame(
            columns=[
                "Имя сервиса",
                "Код",
                "Код активности",
                "Наименование активности",
                "samples_value",
            ]
        )

    out = pd.concat(out_parts, ignore_index=True)
    out["Имя сервиса"] = out["Имя сервиса"].fillna("").astype(str).map(clean_spaces)
    out["Код"] = out["Код"].fillna("").astype(str).map(normalize_sid)
    out["Код активности"] = (
        out["Код активности"].fillna("").astype(str).map(clean_spaces)
    )
    out["Наименование активности"] = (
        out["Наименование активности"].fillna("").astype(str).map(clean_spaces)
    )
    out["samples_value"] = out["samples_value"].fillna(0).astype(int)
    return out


def build_daily_df_report(accounted: pd.DataFrame) -> pd.DataFrame:
    if accounted.empty:
        return pd.DataFrame(
            columns=[
                "Имя сервиса",
                "Код",
                "Код активности",
                "Наименование активности",
                "samples_value",
            ]
        )

    df_for_report = accounted.rename(
        columns={
            "service_name": "Имя сервиса",
            "service_id": "Код",
            "activity_code": "Код активности",
            "activity_name": "Наименование активности",
        }
    )[
        [
            "Имя сервиса",
            "Код",
            "Код активности",
            "Наименование активности",
            "samples_value",
        ]
    ]

    return dedupe_daily_report(df_for_report)


class SqliteDb:
    def __init__(self, path: str):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row

    def init_schema(self):
        cur = self.conn.cursor()

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS snapshot_runs (
                snapshot_date TEXT PRIMARY KEY,
                rows_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_activity_samples (
                snapshot_date TEXT NOT NULL,
                service_key TEXT NOT NULL,
                service_id TEXT NOT NULL DEFAULT '',
                service_name TEXT NOT NULL DEFAULT '',
                activity_code TEXT NOT NULL DEFAULT '',
                activity_name TEXT NOT NULL DEFAULT '',
                samples_value INTEGER NOT NULL DEFAULT 0,
                window_hours INTEGER NOT NULL DEFAULT 24,
                created_at TEXT NOT NULL,
                PRIMARY KEY (snapshot_date, service_key, activity_code)
            )
            """
        )

        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_daily_activity_samples_snapshot_date
            ON daily_activity_samples (snapshot_date)
            """
        )

        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_daily_activity_samples_service_id
            ON daily_activity_samples (service_id)
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_unaccounted (
                snapshot_date TEXT NOT NULL,
                stage TEXT NOT NULL DEFAULT '',
                team TEXT NOT NULL DEFAULT '',
                service_id TEXT NOT NULL DEFAULT '',
                service_name TEXT NOT NULL DEFAULT '',
                activity_code TEXT NOT NULL DEFAULT '',
                activity_name TEXT NOT NULL DEFAULT '',
                metric TEXT NOT NULL DEFAULT '',
                samples_value TEXT NOT NULL DEFAULT '',
                reason TEXT NOT NULL DEFAULT '',
                detail TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL
            )
            """
        )

        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_daily_unaccounted_snapshot_date
            ON daily_unaccounted (snapshot_date)
            """
        )

        self.conn.commit()

    def has_snapshot(self, snapshot_date: str) -> bool:
        cur = self.conn.cursor()
        cur.execute(
            "SELECT 1 FROM snapshot_runs WHERE snapshot_date = ? LIMIT 1",
            (snapshot_date,),
        )
        return cur.fetchone() is not None

    def save_daily_snapshot(self, snapshot_date: str, df_daily: pd.DataFrame):
        now_str = datetime.now(timezone.utc).isoformat()
        cur = self.conn.cursor()

        try:
            cur.execute("BEGIN")

            cur.execute(
                "SELECT 1 FROM snapshot_runs WHERE snapshot_date = ? LIMIT 1",
                (snapshot_date,),
            )
            if cur.fetchone() is not None:
                self.conn.rollback()
                log.info("Snapshot for %s already exists in DB", snapshot_date)
                return

            rows = []
            for row in df_daily.to_dict("records"):
                service_name = clean_spaces(row.get("Имя сервиса", ""))
                service_id = normalize_sid(row.get("Код", ""))
                activity_code = clean_spaces(row.get("Код активности", ""))
                activity_name = clean_spaces(row.get("Наименование активности", ""))
                samples_value = int(row.get("samples_value", 0) or 0)
                service_key = service_id if service_id else service_name

                rows.append(
                    (
                        snapshot_date,
                        service_key,
                        service_id,
                        service_name,
                        activity_code,
                        activity_name,
                        samples_value,
                        WINDOW_HOURS,
                        now_str,
                    )
                )

            if rows:
                cur.executemany(
                    """
                    INSERT OR IGNORE INTO daily_activity_samples (
                        snapshot_date,
                        service_key,
                        service_id,
                        service_name,
                        activity_code,
                        activity_name,
                        samples_value,
                        window_hours,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    rows,
                )

            cur.execute(
                """
                INSERT INTO snapshot_runs (
                    snapshot_date,
                    rows_count,
                    created_at
                )
                VALUES (?, ?, ?)
                """,
                (snapshot_date, len(rows), now_str),
            )

            self.conn.commit()
            log.info("Snapshot saved into DB: %s rows=%d", snapshot_date, len(rows))
        except Exception:
            self.conn.rollback()
            raise

    def replace_daily_unaccounted(self, snapshot_date: str, df_unacc: pd.DataFrame):
        now_str = datetime.now(timezone.utc).isoformat()
        cur = self.conn.cursor()

        try:
            cur.execute("BEGIN")

            cur.execute(
                "DELETE FROM daily_unaccounted WHERE snapshot_date = ?",
                (snapshot_date,),
            )

            rows = []
            if df_unacc is not None and not df_unacc.empty:
                df_unacc = df_unacc.reindex(columns=UNACC_COLS).fillna("")

                for row in df_unacc.to_dict("records"):
                    samples_value = row.get("samples_value", "")
                    if samples_value == "":
                        samples_value_str = ""
                    else:
                        samples_value_str = str(samples_value)

                    rows.append(
                        (
                            snapshot_date,
                            str(row.get("stage", "") or "").strip(),
                            str(row.get("team", "") or "").strip(),
                            normalize_sid(row.get("service_id", "")),
                            clean_spaces(row.get("service_name", "")),
                            clean_spaces(row.get("activity_code", "")),
                            clean_spaces(row.get("activity_name", "")),
                            str(row.get("metric", "") or "").strip(),
                            samples_value_str,
                            str(row.get("reason", "") or "").strip(),
                            str(row.get("detail", "") or "").strip(),
                            now_str,
                        )
                    )

            if rows:
                cur.executemany(
                    """
                    INSERT INTO daily_unaccounted (
                        snapshot_date,
                        stage,
                        team,
                        service_id,
                        service_name,
                        activity_code,
                        activity_name,
                        metric,
                        samples_value,
                        reason,
                        detail,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    rows,
                )

            self.conn.commit()
            log.info("Unaccounted replaced for %s: rows=%d", snapshot_date, len(rows))
        except Exception:
            self.conn.rollback()
            raise

    def load_daily_unaccounted(self, snapshot_date: str) -> pd.DataFrame:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                stage,
                team,
                service_id,
                service_name,
                activity_code,
                activity_name,
                metric,
                samples_value,
                reason,
                detail
            FROM daily_unaccounted
            WHERE snapshot_date = ?
            ORDER BY stage, reason, team, service_id, metric
            """,
            (snapshot_date,),
        )

        rows = [dict(x) for x in cur.fetchall()]
        if not rows:
            return pd.DataFrame(columns=UNACC_COLS)
        return pd.DataFrame(rows, columns=UNACC_COLS)

    def _period_start_date(self, lookback_days: int) -> str | None:
        if lookback_days <= 0:
            return None
        start_dt = datetime.now(timezone.utc).date() - timedelta(days=lookback_days - 1)
        return start_dt.isoformat()

    def load_period_rows(self, lookback_days: int) -> pd.DataFrame:
        cur = self.conn.cursor()
        start_date = self._period_start_date(lookback_days)

        if start_date is None:
            cur.execute(
                """
                SELECT
                    snapshot_date,
                    service_id,
                    service_name,
                    activity_code,
                    activity_name,
                    samples_value,
                    window_hours
                FROM daily_activity_samples
                """
            )
        else:
            cur.execute(
                """
                SELECT
                    snapshot_date,
                    service_id,
                    service_name,
                    activity_code,
                    activity_name,
                    samples_value,
                    window_hours
                FROM daily_activity_samples
                WHERE snapshot_date >= ?
                """,
                (start_date,),
            )

        rows = [dict(x) for x in cur.fetchall()]
        if not rows:
            return pd.DataFrame(
                columns=[
                    "snapshot_date",
                    "service_id",
                    "service_name",
                    "activity_code",
                    "activity_name",
                    "samples_value",
                    "window_hours",
                ]
            )
        return pd.DataFrame(rows)

    def close(self):
        if getattr(self, "conn", None) is not None:
            self.conn.close()
            self.conn = None


def calc_period_distance_days(df_period: pd.DataFrame) -> int:
    if df_period is None or df_period.empty or "snapshot_date" not in df_period.columns:
        return 1

    dates = pd.to_datetime(df_period["snapshot_date"], errors="coerce").dropna()
    if dates.empty:
        return 1

    min_dt = dates.min().date()
    max_dt = dates.max().date()
    delta_days = (max_dt - min_dt).days + 1

    if delta_days <= 0:
        return 1
    return int(delta_days)


def build_period_samples_col_name(df_period: pd.DataFrame) -> str:
    days_distance = calc_period_distance_days(df_period)
    return f"samples_за_{days_distance}d"


def build_period_report(df_period: pd.DataFrame) -> pd.DataFrame:
    samples_col = build_period_samples_col_name(df_period)

    report_cols = [
        "Имя сервиса",
        "Код",
        "Код активности",
        "Наименование активности",
        samples_col,
        "% от общего числа",
    ]

    if df_period.empty:
        return pd.DataFrame(columns=report_cols)

    df = df_period.copy()
    df["Код"] = df["service_id"].fillna("").astype(str).map(normalize_sid)
    df["Имя сервиса"] = df["service_name"].fillna("").astype(str).map(clean_spaces)
    df["Код активности"] = df["activity_code"].fillna("").astype(str).map(clean_spaces)
    df["Наименование активности"] = (
        df["activity_name"].fillna("").astype(str).map(clean_spaces)
    )
    df["samples_value"] = df["samples_value"].fillna(0).astype(int)

    has_code = df[df["Код"] != ""].copy()
    no_code = df[df["Код"] == ""].copy()

    out_parts = []

    if not has_code.empty:
        has_code = has_code.groupby(
            ["Код", "Код активности"],
            as_index=False,
        ).agg(
            {
                "Имя сервиса": _first_non_empty,
                "Наименование активности": _first_non_empty,
                "samples_value": "sum",
            }
        )
        out_parts.append(
            has_code.rename(columns={"samples_value": samples_col})[
                [
                    "Имя сервиса",
                    "Код",
                    "Код активности",
                    "Наименование активности",
                    samples_col,
                ]
            ]
        )

    if not no_code.empty:
        no_code = no_code.groupby(
            ["Имя сервиса", "Код активности"],
            as_index=False,
        ).agg(
            {
                "Код": _first_non_empty,
                "Наименование активности": _first_non_empty,
                "samples_value": "sum",
            }
        )
        out_parts.append(
            no_code.rename(columns={"samples_value": samples_col})[
                [
                    "Имя сервиса",
                    "Код",
                    "Код активности",
                    "Наименование активности",
                    samples_col,
                ]
            ]
        )

    if not out_parts:
        return pd.DataFrame(columns=report_cols)

    out = pd.concat(out_parts, ignore_index=True)
    out[samples_col] = out[samples_col].fillna(0).astype(int)

    total = float(out[samples_col].sum())
    out["% от общего числа"] = (out[samples_col] / total * 100.0) if total else 0.0
    out["% от общего числа"] = out["% от общего числа"].round(5)

    out = out[report_cols]
    out = out.sort_values([samples_col], ascending=False).reset_index(drop=True)
    return out


def write_report(df_report: pd.DataFrame, df_unacc: pd.DataFrame):
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "samples_за_период"
    ws.append(list(df_report.columns))
    for c in ws[1]:
        c.font = bold

    for row in df_report.itertuples(index=False):
        ws.append(list(row))

    if "% от общего числа" in df_report.columns:
        col_idx = df_report.columns.get_loc("% от общего числа") + 1
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                c.number_format = "0.00000"

    ws2 = wb.create_sheet("Unaccounted")
    if df_unacc is None or df_unacc.empty:
        ws2.append(["No data"])
        ws2["A1"].font = bold
    else:
        df_unacc = df_unacc.reindex(columns=UNACC_COLS)
        ws2.append(list(df_unacc.columns))
        for c in ws2[1]:
            c.font = bold
        for row in df_unacc.itertuples(index=False):
            ws2.append(list(row))

    wb.save(OUTPUT_FILE)


def main():
    vm_url = os.getenv("VM_URL", "").strip()
    if not vm_url:
        log.error("VM_URL не задан")
        sys.exit(1)

    db = SqliteDb(SQLITE_DB_FILE)
    try:
        db.init_schema()

        log.info("VM_URL=%s", vm_url)
        log.info("WINDOW_HOURS=%s", WINDOW_HOURS)
        log.info("MAX_METRICS=%s", MAX_METRICS if MAX_METRICS > 0 else "ALL")
        log.info("VERBOSE_LOG=%s", VERBOSE_LOG)
        log.info(
            "BAN_SERVICE_IDS=%s", sorted(ban_service_set) if ban_service_set else "[]"
        )
        log.info("ACTIVITY_FILE=%s", ACTIVITY_FILE)
        log.info("SQLITE_DB_FILE=%s", SQLITE_DB_FILE)
        log.info("SNAPSHOT_LOOKBACK_DAYS=%s", SNAPSHOT_LOOKBACK_DAYS)

        activity_df = read_activity_map(ACTIVITY_FILE)

        overrides = {}
        for k, v in (TEAM_SERVICE_ID_OVERRIDES or {}).items():
            kk = clean_spaces(k)
            vv = normalize_sid(v)
            if kk and vv:
                overrides[kk] = vv

        today_snapshot = datetime.now(timezone.utc).date().isoformat()

        unacc_map = {}
        collected_unacc_today = False

        def add_unacc_once(
            stage, team, service_id, metric, reason, detail, samples_value=None
        ):
            key = (team or "", service_id or "", metric or "")
            if key in unacc_map:
                if VERBOSE_LOG:
                    log.info(
                        "Unaccounted skip duplicate | team=%s | service_id=%s | metric=%s",
                        team,
                        service_id,
                        metric,
                    )
                return

            sid = normalize_sid(service_id)
            service_name = ""
            activity_code = ""
            activity_name = ""

            if sid and not activity_df.empty:
                match = activity_df[activity_df["code"].astype(str) == sid]
                if not match.empty:
                    first = match.iloc[0]
                    service_name = clean_spaces(first.get("service_name", ""))
                    activity_code = clean_spaces(first.get("activity_code", ""))
                    activity_name = clean_spaces(first.get("activity_name", ""))

            unacc_map[key] = {
                "stage": stage,
                "team": team or "",
                "service_id": sid,
                "service_name": service_name,
                "activity_code": activity_code,
                "activity_name": activity_name,
                "metric": metric or "",
                "samples_value": "" if samples_value is None else int(samples_value),
                "reason": reason,
                "detail": detail,
            }

            if VERBOSE_LOG:
                log.info(
                    "Unaccounted add | stage=%s | team=%s | service_id=%s | metric=%s | reason=%s | detail=%s | samples=%s",
                    stage,
                    team,
                    sid,
                    metric,
                    reason,
                    detail,
                    samples_value,
                )

        if db.has_snapshot(today_snapshot):
            log.info("Snapshot for today already exists: %s", today_snapshot)
            log.info("Heavy Victoria collection skipped")
        else:
            log.info("No snapshot for today, starting heavy Victoria collection")

            log.info("Discover series ...")
            series_rows = discover_series(vm_url)
            log.info("Series found: %d", len(series_rows))

            team_to_sid_map, ambiguous_teams = build_team_to_sid_maps(series_rows)
            log.info("Team->SID inferred map size: %d", len(team_to_sid_map))
            log.info("Ambiguous teams: %d", len(ambiguous_teams))
            log.info("Overrides: %d", len(overrides))

            metrics_audit = []
            accounted_metric_rows = []

            for idx, r in enumerate(series_rows, 1):
                team_raw = (r.get("team_raw") or "").strip()
                team_base = (r.get("team_base") or "").strip()
                service_id_raw = normalize_sid(r.get("service_id_raw"))
                sid_from_team = normalize_sid(r.get("sid_from_team"))
                sid_seed = normalize_sid(r.get("sid_seed"))
                metric = (r.get("metric") or "").strip()

                if not metric:
                    continue

                service_id_final = sid_seed
                stage = ""
                reason = ""
                detail = ""
                status = "accounted"

                ov = overrides.get(team_base)
                if ov:
                    service_id_final = ov
                else:
                    if not service_id_final:
                        if (
                            (not service_id_raw)
                            and (not sid_from_team)
                            and (team_base in ambiguous_teams)
                        ):
                            status = "unaccounted"
                            stage = "infer"
                            reason = "ambiguous_service_id"
                            detail = "multiple service_id detected for team_base"
                        else:
                            inferred = team_to_sid_map.get(team_base, "")
                            if (
                                (not service_id_raw)
                                and (not sid_from_team)
                                and inferred
                            ):
                                service_id_final = inferred

                if is_banned_team(team_base):
                    status = "unaccounted"
                    stage = "discover"
                    reason = "banned_team"
                    detail = "team in BAN_TEAMS"

                if (
                    status == "accounted"
                    and service_id_final
                    and service_id_final in ban_service_set
                ):
                    status = "unaccounted"
                    stage = "discover"
                    reason = "banned_service_id"
                    detail = "service_id in BAN_SERVICE_IDS"

                if status == "accounted" and not service_id_final:
                    status = "unaccounted"
                    stage = "discover"
                    reason = "excluded_no_service_id"
                    detail = "service_id empty"

                samples_value = int(r.get("samples_value", 0) or 0)

                if VERBOSE_LOG:
                    log.info(
                        "Metric decision | metric=%s | team_raw=%s | team_base=%s | service_id_raw=%s | sid_from_team=%s | sid_seed=%s | service_id_final=%s | samples=%s | status=%s | stage=%s | reason=%s",
                        metric,
                        team_raw,
                        team_base,
                        service_id_raw,
                        sid_from_team,
                        sid_seed,
                        service_id_final,
                        samples_value,
                        status,
                        stage,
                        reason,
                    )

                if status == "unaccounted":
                    if VERBOSE_LOG:
                        log.info(
                            "Metric routed | destination=unaccounted | metric=%s | team=%s | service_id=%s | reason=%s | detail=%s | samples=%s",
                            metric,
                            team_base,
                            service_id_final,
                            reason,
                            detail,
                            samples_value,
                        )

                    add_unacc_once(
                        stage,
                        team_base,
                        service_id_final,
                        metric,
                        reason,
                        detail,
                        samples_value=samples_value,
                    )
                else:
                    if VERBOSE_LOG:
                        log.info(
                            "Metric routed | destination=accounted_metric_rows | metric=%s | team=%s | service_id=%s | samples=%s",
                            metric,
                            team_base,
                            service_id_final,
                            samples_value,
                        )

                    accounted_metric_rows.append(
                        {
                            "team_base": team_base,
                            "service_id_final": service_id_final,
                            "samples_value": int(samples_value),
                        }
                    )

                metrics_audit.append(
                    {
                        "team_base": team_base,
                        "team_raw": team_raw,
                        "service_id_raw": service_id_raw,
                        "service_id_final": service_id_final,
                        "metric": metric,
                        "samples_value": int(samples_value),
                        "status": status,
                        "stage": stage,
                        "reason": reason,
                        "detail": detail,
                    }
                )

                if idx % 200 == 0:
                    log.info("Processed: %d/%d", idx, len(series_rows))

                time.sleep(SLEEP_SEC)

            log.info(
                "Metric routing summary | total_series=%d | accounted_metric_rows=%d | unaccounted_rows=%d",
                len(series_rows),
                len(accounted_metric_rows),
                len(unacc_map),
            )

            group_rows = aggregate_to_group(accounted_metric_rows)
            log.info("Aggregate done | grouped_rows=%d", len(group_rows))

            enriched = enrich_group_rows(group_rows, activity_df)
            log.info("Enrich done | enriched_rows=%d", len(enriched))

            if not enriched.empty:
                accounted = enriched.copy()

                def mark_unacc_for_service(team, sid, stage, reason, detail):
                    for rr in metrics_audit:
                        if rr.get("status") != "accounted":
                            continue
                        if (rr.get("team_base") or "") == (
                            team or ""
                        ) and normalize_sid(
                            rr.get("service_id_final")
                        ) == normalize_sid(sid):
                            if VERBOSE_LOG:
                                log.info(
                                    "Enrich re-route | metric=%s | team=%s | service_id=%s | reason=%s",
                                    rr.get("metric", ""),
                                    rr.get("team_base", ""),
                                    rr.get("service_id_final", ""),
                                    reason,
                                )

                            add_unacc_once(
                                stage,
                                rr.get("team_base", ""),
                                rr.get("service_id_final", ""),
                                rr.get("metric", ""),
                                reason,
                                detail,
                                samples_value=int(rr.get("samples_value", 0) or 0),
                            )
                            rr["status"] = "unaccounted"
                            rr["stage"] = stage
                            rr["reason"] = reason
                            rr["detail"] = detail

                m_activity_missing = ~accounted["activity_found"]
                for rr in accounted[m_activity_missing].to_dict("records"):
                    mark_unacc_for_service(
                        rr.get("team", ""),
                        rr.get("service_id", ""),
                        "enrich",
                        "activity_mapping_miss",
                        "service_id not found in activity.xlsx",
                    )

                accounted = accounted[accounted["activity_found"]].copy()
                log.info("Accounted after activity filter | rows=%d", len(accounted))
                df_daily = build_daily_df_report(accounted)
            else:
                df_daily = pd.DataFrame(
                    columns=[
                        "Имя сервиса",
                        "Код",
                        "Код активности",
                        "Наименование активности",
                        "samples_value",
                    ]
                )

            log.info(
                "DB save daily snapshot | snapshot_date=%s | rows=%d",
                today_snapshot,
                len(df_daily),
            )
            db.save_daily_snapshot(today_snapshot, df_daily)

            df_unacc_new = pd.DataFrame(list(unacc_map.values()))
            if not df_unacc_new.empty:
                df_unacc_new = df_unacc_new.reindex(columns=UNACC_COLS).fillna("")
                df_unacc_new = df_unacc_new.sort_values(
                    ["stage", "reason", "team", "service_id", "metric"]
                ).reset_index(drop=True)
            else:
                df_unacc_new = pd.DataFrame(columns=UNACC_COLS)

            log.info(
                "DB replace unaccounted | snapshot_date=%s | rows=%d",
                today_snapshot,
                len(df_unacc_new),
            )
            db.replace_daily_unaccounted(today_snapshot, df_unacc_new)
            collected_unacc_today = True

        df_period = db.load_period_rows(SNAPSHOT_LOOKBACK_DAYS)

        log.info("Period rows loaded: %d", len(df_period))
        log.info("Period distance days: %d", calc_period_distance_days(df_period))

        df_report = build_period_report(df_period)

        df_unacc = db.load_daily_unaccounted(today_snapshot)
        if not df_unacc.empty:
            df_unacc = df_unacc.reindex(columns=UNACC_COLS).fillna("")
            df_unacc = df_unacc.sort_values(
                ["stage", "reason", "team", "service_id", "metric"]
            ).reset_index(drop=True)

        log.info("Unaccounted loaded from DB: %d", len(df_unacc))
        log.info("Collected unaccounted today in current run: %s", collected_unacc_today)

        log.info("Saving report: %s", OUTPUT_FILE)
        write_report(df_report, df_unacc)
        log.info("✔ Done")
    finally:
        db.close()


if __name__ == "__main__":
    main()