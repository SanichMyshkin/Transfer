#!/usr/bin/env python3
import os
import sys
import logging
import re

import clickhouse_connect
import humanize
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

load_dotenv()

HOST = os.getenv("CH_HOST")
PORT = int(os.getenv("CH_PORT", 8123))
USER = os.getenv("CH_USER", "default")
PASSWORD = os.getenv("CH_PASSWORD", "")
DATABASE = os.getenv("CH_DATABASE")

SECURE = False
VERIFY = False

ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")
OUTPUT_FILE = os.getenv("CH_OUT", "openTelemetry_report.xlsx")

LOG = logging.getLogger("ch_table_sizes")

SQL = """
SELECT
    t.name AS table,
    coalesce(sum(p.bytes_on_disk), 0) AS bytes_on_disk
FROM system.tables AS t
LEFT JOIN system.parts AS p
    ON p.database = t.database
   AND p.table = t.name
   AND p.active
WHERE t.database = {db:String}
  AND t.engine NOT IN ('View', 'MaterializedView', 'LiveView')
  AND t.engine != 'Distributed'
GROUP BY table
ORDER BY bytes_on_disk DESC
"""

TABLE_PREFIX_OVERRIDES = {
    "test_table1": ("test", 11111),
    "test_table": ("test", 11111),
}

BAN_SERVICE_IDS = {
    15473,
}

NAME_RE = re.compile(r"^otel_([a-z0-9]+)_(\d+)_traces(?:_trace_id_ts)?$", re.IGNORECASE)


def setup_logging():
    level = os.getenv("LOG_LEVEL", "INFO").upper()
    logging.basicConfig(
        level=getattr(logging, level, logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )


def die(msg: str, code: int = 2):
    LOG.error(msg)
    raise SystemExit(code)


def clean_spaces(s: str) -> str:
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_number(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    return s


def map_table(table_name: str):
    for base, (svc_name, svc_id) in TABLE_PREFIX_OVERRIDES.items():
        if table_name == base or table_name == f"{base}_trace_id_ts":
            return svc_name, int(svc_id), "override"

    m = NAME_RE.match(table_name)
    if not m:
        return None, None, "unmatched"

    return m.group(1).upper(), int(m.group(2)), "parsed"


def read_activity_map(path):
    if not path or not os.path.exists(path):
        raise RuntimeError(f"ACTIVITY_FILE не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    activity = {}
    for row in ws.iter_rows(values_only=True):
        code = normalize_number(row[0] if len(row) > 0 else "")
        if not code:
            continue

        if code in activity:
            continue

        service_name = clean_spaces(str(row[1] or "")) if len(row) > 1 else ""
        activity_code = clean_spaces(str(row[2] or "")) if len(row) > 2 else ""
        activity_name = clean_spaces(str(row[3] or "")) if len(row) > 3 else ""

        activity[code] = {
            "service_name": service_name,
            "activity_code": activity_code,
            "activity_name": activity_name,
        }

    wb.close()
    LOG.info("activity loaded: %d", len(activity))
    return activity


def main() -> int:
    setup_logging()

    if not HOST or not DATABASE:
        die("CH_HOST and CH_DATABASE must be set", 1)

    LOG.info("Connecting to %s:%s secure=%s", HOST, PORT, SECURE)

    activity = read_activity_map(ACTIVITY_FILE)

    try:
        client = clickhouse_connect.get_client(
            host=HOST,
            port=PORT,
            username=USER,
            password=PASSWORD,
            secure=SECURE,
            verify=VERIFY,
        )
    except Exception:
        LOG.exception("Connection failed")
        return 2

    try:
        LOG.info("Executing query for database=%s", DATABASE)
        res = client.query(SQL, parameters={"db": DATABASE})
        table_rows = res.result_rows
        LOG.info("Query returned %d tables (including empty)", len(table_rows))
    except Exception:
        LOG.exception("Query failed")
        return 3

    agg = {}
    unaccounted_rows = []

    unmatched = 0
    banned = 0
    activity_miss = 0

    for table, bytes_on_disk in table_rows:
        b = int(bytes_on_disk or 0)

        parsed_service_name, service_id, src = map_table(table)

        if src == "unmatched" or service_id is None:
            unmatched += 1
            LOG.warning("Unmatched table: %s (bytes=%d)", table, b)
            unaccounted_rows.append(
                {
                    "table": table,
                    "service_id": "",
                    "service_name": "",
                    "activity_code": "",
                    "activity_name": "",
                    "bytes_on_disk": b,
                    "size_h": humanize.naturalsize(b, binary=True),
                    "reason": "unmatched",
                    "detail": "table name does not match NAME_RE and no TABLE_PREFIX_OVERRIDES hit",
                    "source": src,
                }
            )
            continue

        if service_id in BAN_SERVICE_IDS:
            banned += 1
            LOG.info("Banned service_id: table=%s service_id=%s bytes=%d", table, service_id, b)
            meta = activity.get(str(service_id), {})
            unaccounted_rows.append(
                {
                    "table": table,
                    "service_id": service_id,
                    "service_name": meta.get("service_name", parsed_service_name or ""),
                    "activity_code": meta.get("activity_code", ""),
                    "activity_name": meta.get("activity_name", ""),
                    "bytes_on_disk": b,
                    "size_h": humanize.naturalsize(b, binary=True),
                    "reason": "banned_service_id",
                    "detail": "service_id in BAN_SERVICE_IDS",
                    "source": src,
                }
            )
            continue

        if service_id not in agg:
            agg[service_id] = {
                "parsed_service_name": parsed_service_name or "",
                "bytes": 0,
            }

        agg[service_id]["bytes"] += b

        if src == "override" and not agg[service_id]["parsed_service_name"] and parsed_service_name:
            agg[service_id]["parsed_service_name"] = parsed_service_name

    wb = Workbook()

    ws = wb.active
    ws.title = "Отчет OpenTelemetry"

    headers = [
        "Наименование сервиса",
        "Код сервиса",
        "Код активности",
        "Наименование активности",
        "Объем (bytes)",
        "Объем",
        "% от учтенного итога",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for i in range(1, len(headers) + 1):
        ws.cell(row=1, column=i).font = bold

    final_rows = []
    for service_id, payload in agg.items():
        meta = activity.get(str(service_id))
        service_bytes = int(payload["bytes"] or 0)

        if not meta:
            activity_miss += 1
            LOG.warning("No activity mapping for service_id=%s bytes=%d", service_id, service_bytes)
            unaccounted_rows.append(
                {
                    "table": "",
                    "service_id": service_id,
                    "service_name": payload.get("parsed_service_name", ""),
                    "activity_code": "",
                    "activity_name": "",
                    "bytes_on_disk": service_bytes,
                    "size_h": humanize.naturalsize(service_bytes, binary=True),
                    "reason": "activity_mapping_miss",
                    "detail": "service_id отсутствует в activity.xlsx",
                    "source": "aggregated",
                }
            )
            continue

        final_rows.append(
            {
                "service_id": service_id,
                "service_name": meta.get("service_name", "") or payload.get("parsed_service_name", ""),
                "activity_code": meta.get("activity_code", ""),
                "activity_name": meta.get("activity_name", ""),
                "bytes": service_bytes,
            }
        )

    final_rows.sort(key=lambda x: x["bytes"], reverse=True)

    total_accounted = sum(x["bytes"] for x in final_rows)

    for row in final_rows:
        pct = (row["bytes"] / total_accounted * 100.0) if total_accounted > 0 else 0.0
        ws.append(
            [
                row["service_name"],
                row["service_id"],
                row["activity_code"],
                row["activity_name"],
                row["bytes"],
                humanize.naturalsize(row["bytes"], binary=True),
                round(pct, 4),
            ]
        )

    ws_un = wb.create_sheet("Unaccounted")
    un_headers = [
        "table",
        "service_id",
        "service_name",
        "activity_code",
        "activity_name",
        "bytes_on_disk",
        "size_human",
        "reason",
        "detail",
        "source",
    ]
    ws_un.append(un_headers)

    for i in range(1, len(un_headers) + 1):
        ws_un.cell(row=1, column=i).font = bold

    for row in unaccounted_rows:
        ws_un.append(
            [
                row.get("table", ""),
                row.get("service_id", ""),
                row.get("service_name", ""),
                row.get("activity_code", ""),
                row.get("activity_name", ""),
                int(row.get("bytes_on_disk") or 0),
                row.get("size_h", ""),
                row.get("reason", ""),
                row.get("detail", ""),
                row.get("source", ""),
            ]
        )

    wb.save(OUTPUT_FILE)

    LOG.info(
        "Report saved: %s | services=%d | tables_total=%d | unmatched_tables=%d | banned_tables=%d | activity_miss=%d | unaccounted_rows=%d",
        OUTPUT_FILE,
        len(final_rows),
        len(table_rows),
        unmatched,
        banned,
        activity_miss,
        len(unaccounted_rows),
    )

    return 0


if __name__ == "__main__":
    sys.exit(main())