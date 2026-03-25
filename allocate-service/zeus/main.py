import os
import logging
import urllib3
import re

import gitlab
import yaml
import pandas as pd
from dotenv import load_dotenv

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger("zeus_report")

load_dotenv()

GITLAB_URL = os.getenv("GITLAB_URL", "").rstrip("/")
TOKEN = os.getenv("TOKEN", "")
GROUP_ID = os.getenv("GROUP_ID", "").strip()

ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

OUTPUT_XLSX = "zeus_report.xlsx"
GIT_REF = "main"

BAN_SERVICES = [
    "TEST",
    "UNAITP",
]

BAN_SERVICE_IDS = [15473]

EXCLUDE_ZERO_CODES = True

CPU_WEIGHT = 0.40
RAM_WEIGHT = 0.60


def clean_spaces(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def load_activity_map(path: str):
    log.info("Читаем activity (A=КОД, B=Наименование сервиса, C=Код активности, D=Наименование активности)...")

    if not path or not os.path.exists(path):
        raise RuntimeError(f"ACTIVITY_FILE не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    mp = {}

    for row in ws.iter_rows(values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else None)
        if not code:
            continue

        if code in mp:
            continue

        mp[code] = {
            "service_name": clean_spaces(row[1] if len(row) > 1 else ""),
            "activity_code": clean_spaces(row[2] if len(row) > 2 else ""),
            "activity_name": clean_spaces(row[3] if len(row) > 3 else ""),
        }

    wb.close()
    log.info("ACTIVITY: загружено сервисов по КОД: %d", len(mp))
    return mp


def gl_connect():
    log.info("Подключаемся к GitLab...")
    gl = gitlab.Gitlab(GITLAB_URL, private_token=TOKEN, ssl_verify=False, timeout=60)
    gl.auth()
    log.info("GitLab: ok")
    return gl


def get_group_projects(gl):
    log.info("Получаем проекты группы GROUP_ID=%s (включая сабгруппы)...", GROUP_ID)
    group = gl.groups.get(GROUP_ID)
    projs = group.projects.list(all=True, include_subgroups=True)
    log.info("Проектов найдено: %d", len(projs))
    return projs


def repo_tree(proj, path=None):
    if path:
        return proj.repository_tree(path=path, all=True)
    return proj.repository_tree(all=True)


def get_file_text(proj, file_path: str, ref: str) -> str:
    f = proj.files.get(file_path=file_path, ref=ref)
    return f.decode().decode("utf-8")


def find_deployment_files(proj):
    try:
        root = repo_tree(proj)
    except Exception:
        return []

    zeus_dir = next(
        (i for i in root if i["type"] == "tree" and i["name"].startswith("zeus-")),
        None,
    )
    if not zeus_dir:
        return []

    try:
        zeus_items = repo_tree(proj, zeus_dir["path"])
        subfolders = [i for i in zeus_items if i["type"] == "tree"]
    except Exception:
        return []

    files = []
    for sub in subfolders:
        try:
            sub_items = repo_tree(proj, sub["path"])
        except Exception:
            continue

        for f in sub_items:
            if f["type"] != "blob":
                continue
            n = (f.get("name") or "").lower()
            if n.endswith("-deployment.yaml") or n.endswith("-deployment.yml"):
                files.append(f["path"])

    return files


def parse_cpu_to_cores(v):
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    if s.endswith("m"):
        try:
            return float(s[:-1].strip()) / 1000.0
        except Exception:
            return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_mem_to_bytes(v):
    if v is None:
        return 0
    s = str(v).strip()
    if not s:
        return 0

    units = {
        "Ki": 1024,
        "Mi": 1024**2,
        "Gi": 1024**3,
        "Ti": 1024**4,
        "K": 1000,
        "M": 1000**2,
        "G": 1000**3,
        "T": 1000**4,
    }
    for u, mul in units.items():
        if s.endswith(u):
            try:
                return int(float(s[: -len(u)]) * mul)
            except Exception:
                return 0
    try:
        return int(float(s))
    except Exception:
        return 0


def bytes_to_mib(b: int) -> float:
    return float(b) / (1024.0**2)


def parse_deployment_limits(text: str):
    try:
        docs = list(yaml.safe_load_all(text))
    except Exception:
        return 0.0, 0

    cpu_sum = 0.0
    mem_sum = 0

    for doc in docs:
        if not isinstance(doc, dict):
            continue
        if (doc.get("kind") or "") != "Deployment":
            continue

        containers = (
            doc.get("spec", {})
            .get("template", {})
            .get("spec", {})
            .get("containers", [])
        )

        for c in containers:
            res = c.get("resources", {})
            lim = res.get("limits", {})

            cpu_sum += parse_cpu_to_cores(lim.get("cpu"))
            mem_sum += parse_mem_to_bytes(lim.get("memory"))

    return cpu_sum, mem_sum


def split_service_and_code(project_name: str):
    m = re.match(r"^(.*?)-(\d+)$", project_name or "")
    if not m:
        return project_name, ""
    return m.group(1), m.group(2)


def is_zero_code(code: str):
    return code and set(code) == {"0"}


def build_ban_set(vals):
    return {str(x).strip() for x in vals if str(x).strip()}


def collect_rows(gl, projects, activity_map):
    ban_service_ids = build_ban_set(BAN_SERVICE_IDS)
    totals = {}
    unaccounted = []

    for p in projects:
        proj = gl.projects.get(p.id)

        log.info("PROJECT: %s (%s)", p.path_with_namespace, p.id)

        files = find_deployment_files(proj)

        cpu_total = 0.0
        mem_total = 0

        for path in files:
            try:
                raw = get_file_text(proj, path, GIT_REF)
            except Exception as e:
                log.warning("Не удалось прочитать файл %s в проекте %s: %s", path, p.path_with_namespace, e)
                continue

            cpu, mem = parse_deployment_limits(raw)
            cpu_total += cpu
            mem_total += mem

        service_from_git, code = split_service_and_code(p.name)

        def add_unacc(reason, detail, service_name="", activity_code="", activity_name=""):
            unaccounted.append(
                {
                    "project_id": p.id,
                    "project": p.name,
                    "path_with_namespace": p.path_with_namespace,
                    "service_guess": service_from_git,
                    "code": code,
                    "service_name": service_name,
                    "activity_code": activity_code,
                    "activity_name": activity_name,
                    "cpu_cores": cpu_total,
                    "mem_mib": bytes_to_mib(mem_total),
                    "deployment_files": len(files),
                    "reason": reason,
                    "detail": detail,
                }
            )

        if not code:
            log.info('UNACCOUNTED project=%s reason=no_code', p.path_with_namespace)
            add_unacc("no_code", "project name does not contain code")
            continue

        if EXCLUDE_ZERO_CODES and is_zero_code(code):
            log.info('UNACCOUNTED project=%s code=%s reason=zero_code', p.path_with_namespace, code)
            add_unacc("zero_code", "code consists of zeros")
            continue

        meta = activity_map.get(code, {})
        activity_service = clean_spaces(meta.get("service_name", ""))
        activity_code = clean_spaces(meta.get("activity_code", ""))
        activity_name = clean_spaces(meta.get("activity_name", ""))

        if code in ban_service_ids:
            log.info('UNACCOUNTED project=%s code=%s reason=banned_service_id', p.path_with_namespace, code)
            add_unacc(
                "banned_service_id",
                "code in BAN_SERVICE_IDS",
                service_name=activity_service,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        service = activity_service or service_from_git

        if service in BAN_SERVICES:
            log.info('UNACCOUNTED project=%s code=%s reason=banned_service service=%s', p.path_with_namespace, code, service)
            add_unacc(
                "banned_service",
                service,
                service_name=activity_service,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        if code not in activity_map:
            log.info('UNACCOUNTED project=%s code=%s reason=activity_mapping_miss', p.path_with_namespace, code)
            add_unacc(
                "activity_mapping_miss",
                "code not found in activity.xlsx",
                service_name=activity_service,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        if cpu_total <= 0 and mem_total <= 0:
            log.info('SKIP project=%s code=%s reason=empty_resources', p.path_with_namespace, code)
            continue

        key = code

        if key not in totals:
            totals[key] = {
                "service": activity_service or service_from_git,
                "code": code,
                "activity_code": activity_code,
                "activity_name": activity_name,
                "cpu": 0.0,
                "mem": 0,
            }

        totals[key]["cpu"] += cpu_total
        totals[key]["mem"] += mem_total

        log.info(
            'ACCOUNTED project=%s code=%s service="%s" activity_code="%s" cpu=%.6f mem_mib=%.2f files=%d',
            p.path_with_namespace,
            code,
            totals[key]["service"],
            totals[key]["activity_code"],
            cpu_total,
            bytes_to_mib(mem_total),
            len(files),
        )

    total_cpu = sum(v["cpu"] for v in totals.values())
    total_mem = sum(v["mem"] for v in totals.values())

    rows = []

    for code, v in totals.items():
        cpu = v["cpu"]
        mem = v["mem"]

        cpu_pct = (cpu / total_cpu * 100.0) if total_cpu else 0.0
        mem_pct = (mem / total_mem * 100.0) if total_mem else 0.0

        pct = cpu_pct * CPU_WEIGHT + mem_pct * RAM_WEIGHT

        rows.append(
            {
                "service": v["service"],
                "code": code,
                "activity_code": v["activity_code"],
                "activity_name": v["activity_name"],
                "cpu_cores": cpu,
                "mem_mib": bytes_to_mib(mem),
                "pct": pct,
            }
        )

    rows.sort(key=lambda r: r["pct"], reverse=True)
    unaccounted.sort(
        key=lambda r: ((r.get("cpu_cores", 0.0) * CPU_WEIGHT) + (r.get("mem_mib", 0.0) * RAM_WEIGHT)),
        reverse=True,
    )

    return rows, unaccounted


def write_excel(rows, unaccounted_rows, out_file):
    wb = Workbook()

    ws = wb.active
    ws.title = "Report"

    headers = [
        "Имя сервиса",
        "Код",
        "Код активности",
        "Наименование активности",
        "CPU (cores)",
        "RAM (MiB)",
        "% потребления",
    ]

    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True)

    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, r["service"])
        ws.cell(i, 2, r["code"])
        ws.cell(i, 3, r["activity_code"])
        ws.cell(i, 4, r["activity_name"])
        ws.cell(i, 5, round(r["cpu_cores"], 6))
        ws.cell(i, 6, round(r["mem_mib"], 2))
        ws.cell(i, 7, round(r["pct"], 2))

    ws2 = wb.create_sheet("Unaccounted")

    headers2 = [
        "project_id",
        "project",
        "path_with_namespace",
        "service_guess",
        "code",
        "service_name",
        "activity_code",
        "activity_name",
        "CPU (cores)",
        "RAM (MiB)",
        "deployment_files",
        "reason",
        "detail",
    ]

    for i, h in enumerate(headers2, 1):
        c = ws2.cell(1, i, h)
        c.font = Font(bold=True)

    for i, r in enumerate(unaccounted_rows, 2):
        ws2.cell(i, 1, r["project_id"])
        ws2.cell(i, 2, r["project"])
        ws2.cell(i, 3, r["path_with_namespace"])
        ws2.cell(i, 4, r["service_guess"])
        ws2.cell(i, 5, r["code"])
        ws2.cell(i, 6, r["service_name"])
        ws2.cell(i, 7, r["activity_code"])
        ws2.cell(i, 8, r["activity_name"])
        ws2.cell(i, 9, round(r["cpu_cores"], 6))
        ws2.cell(i, 10, round(r["mem_mib"], 2))
        ws2.cell(i, 11, r["deployment_files"])
        ws2.cell(i, 12, r["reason"])
        ws2.cell(i, 13, r["detail"])

    wb.save(out_file)
    log.info("Excel отчет создан: %s", out_file)


def main():
    activity = load_activity_map(ACTIVITY_FILE)

    gl = gl_connect()
    projects = get_group_projects(gl)

    rows, unaccounted = collect_rows(gl, projects, activity)

    write_excel(rows, unaccounted, OUTPUT_XLSX)


if __name__ == "__main__":
    main()