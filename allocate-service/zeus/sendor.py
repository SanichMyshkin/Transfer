import os
import logging
from collections import defaultdict
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv
import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

DB_HOST = os.getenv("DB_HOST")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

SINCE_DAYS = int(os.getenv("SINCE", "90"))
OUT_XLSX = os.getenv("OUT_XLSX", "sender_report.xlsx")
ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

EXCLUDE_SERVICE_IDS = {
    # "15473",
}

ALLOW_ZERO_SERVICE_ID = False
SKIP_ACTIVITY_MAPPING_MISS = False


def clean_spaces(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_code(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    return "".join(ch for ch in s if ch.isdigit())


def is_all_zeros(s: str) -> bool:
    return bool(s) and set(s) == {"0"}


def get_counts_by_service_since_days_sql(days: int):
    since_dt = datetime.now(timezone.utc) - timedelta(days=days)
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
    )
    conn.set_session(readonly=True)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    with raw as (
                        select
                            'express'::text as source_type,
                            nullif(
                                substring(metadata_source_id from 'id:([0-9]+)'),
                                ''
                            )::bigint as service_id
                        from sender.express_events_history
                        where created >= %s

                        union all

                        select
                            'mail'::text as source_type,
                            nullif(
                                substring(metadata_source_id from 'id:([0-9]+)'),
                                ''
                            )::bigint as service_id
                        from sender.mail_events_history
                        where created >= %s
                    )
                    select
                        service_id,
                        count(*) filter (where source_type = 'mail') as mail_cnt,
                        count(*) filter (where source_type = 'express') as express_cnt,
                        count(*) as total_cnt
                    from raw
                    group by 1
                    order by 4 desc
                    """,
                    (since_dt, since_dt),
                )
                rows = cur.fetchall()
                log.info("DB rows fetched: %d", len(rows))
                for service_id, mail_cnt, express_cnt, total_cnt in rows[:20]:
                    log.info(
                        "DB sample row: service_id=%r mail=%s express=%s total=%s",
                        service_id,
                        mail_cnt,
                        express_cnt,
                        total_cnt,
                    )
                return rows
    finally:
        conn.close()


def read_activity_map(path: str):
    if not path or not os.path.exists(path):
        log.warning("ACTIVITY_FILE не найден: %s", path)
        return {}

    wb = load_workbook(path, read_only=True, data_only=True)
    log.info("ACTIVITY workbook sheets: %s", wb.sheetnames)
    ws = wb.worksheets[0]

    out = {}
    rows = 0
    ok = 0

    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        rows += 1

        if i == 1:
            log.info("ACTIVITY header: %s", r)
            continue

        code = normalize_code(r[0] if len(r) > 0 else "")
        if not code:
            continue

        if code in out:
            continue

        out[code] = {
            "service_name": clean_spaces(r[1] if len(r) > 1 else ""),
            "activity_code": clean_spaces(r[2] if len(r) > 2 else ""),
            "activity_name": clean_spaces(r[3] if len(r) > 3 else ""),
        }
        ok += 1

    wb.close()

    log.info("ACTIVITY_FILE exists=%s path=%s", os.path.exists(path), path)
    log.info("ACTIVITY: rows=%d mapped_codes=%d ok_rows=%d", rows, len(out), ok)
    log.info("ACTIVITY_MAP has 15473=%s", "15473" in out)
    log.info("ACTIVITY_MAP[15473]=%s", out.get("15473"))
    log.info("ACTIVITY sample keys: %s", list(out.keys())[:20])

    return out


def aggregate_and_enrich(db_rows, exclude_service_ids, activity_map):
    include_counts = defaultdict(lambda: {"mail": 0, "express": 0, "total": 0})
    unaccounted_rows = []

    for service_id, mail_cnt, express_cnt, total_cnt in db_rows:
        mail_c = int(mail_cnt or 0)
        express_c = int(express_cnt or 0)
        total_c = int(total_cnt or 0)

        if total_c <= 0:
            continue

        if service_id is None:
            unaccounted_rows.append(
                ["", "", "", "", mail_c, express_c, total_c, "missing_service_id"]
            )
            continue

        sid = str(service_id).strip()

        if not sid:
            unaccounted_rows.append(
                ["", "", "", "", mail_c, express_c, total_c, "missing_service_id"]
            )
            continue

        if is_all_zeros(sid) and not ALLOW_ZERO_SERVICE_ID:
            unaccounted_rows.append(
                [sid, "", "", "", mail_c, express_c, total_c, "zero_service_id"]
            )
            continue

        if sid in exclude_service_ids:
            unaccounted_rows.append(
                [sid, "", "", "", mail_c, express_c, total_c, "ban_service_id"]
            )
            continue

        include_counts[sid]["mail"] += mail_c
        include_counts[sid]["express"] += express_c
        include_counts[sid]["total"] += total_c

    rows_main = []
    total_included = 0
    skipped_activity_miss = 0

    for sid, counts in sorted(
        include_counts.items(),
        key=lambda x: x[1]["total"],
        reverse=True,
    ):
        meta = activity_map.get(sid, {})
        if not meta:
            log.info("activity miss sid=%r", sid)

        service_name = clean_spaces(meta.get("service_name", ""))
        activity_code = clean_spaces(meta.get("activity_code", ""))
        activity_name = clean_spaces(meta.get("activity_name", ""))

        if SKIP_ACTIVITY_MAPPING_MISS and not service_name:
            skipped_activity_miss += 1
            unaccounted_rows.append(
                [
                    sid,
                    "",
                    "",
                    "",
                    int(counts["mail"]),
                    int(counts["express"]),
                    int(counts["total"]),
                    "activity_mapping_miss",
                ]
            )
            continue

        total_included += int(counts["total"])
        rows_main.append(
            {
                "service_name": service_name or sid,
                "service_id": sid,
                "activity_code": activity_code,
                "activity_name": activity_name,
                "mail_messages_count": int(counts["mail"]),
                "express_messages_count": int(counts["express"]),
                "total_messages_count": int(counts["total"]),
                "percent_of_all_total": 0.0,
            }
        )

    for row in rows_main:
        cnt = int(row["total_messages_count"] or 0)
        pct = 0.0 if total_included <= 0 else (cnt * 100.0 / total_included)
        row["percent_of_all_total"] = pct

    log.info(
        "aggregate_and_enrich: included_before=%d main_rows=%d skipped_activity_miss=%d unaccounted=%d total_included=%d",
        len(include_counts),
        len(rows_main),
        skipped_activity_miss,
        len(unaccounted_rows),
        total_included,
    )

    return rows_main, unaccounted_rows, total_included


def write_excel(path, rows_main, unaccounted_rows):
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Отчет Sender"

    headers = [
        "Наименование сервиса",
        "Код сервиса",
        "Код активности",
        "Наименование активности",
        "Кол-во писем (Outlook)",
        "Кол-во сообщений (Express)",
        "Сумма сообщений",
        "% потребления",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold

    for r in rows_main:
        ws.append(
            [
                r["service_name"],
                r["service_id"],
                r["activity_code"],
                r["activity_name"],
                int(r["mail_messages_count"]),
                int(r["express_messages_count"]),
                int(r["total_messages_count"]),
                float(r["percent_of_all_total"]),
            ]
        )

    pct_col = headers.index("% потребления") + 1
    for rr in range(2, ws.max_row + 1):
        ws.cell(row=rr, column=pct_col).number_format = "0.00000"

    ws2 = wb.create_sheet("unaccounted")
    headers2 = [
        "service_id",
        "service_name",
        "activity_code",
        "activity_name",
        "mail_messages_count",
        "express_messages_count",
        "total_messages_count",
        "reason",
    ]
    ws2.append(headers2)
    for c in ws2[1]:
        c.font = bold

    ua_sorted = sorted(
        unaccounted_rows,
        key=lambda r: (str(r[7]), -int(r[6] or 0), str(r[0])),
    )
    for r in ua_sorted:
        ws2.append(r)

    wb.save(path)
    log.info("XLSX saved %s", path)


def main():
    log.info("START")
    log.info("SINCE_DAYS=%s", SINCE_DAYS)
    log.info("EXCLUDE_SERVICE_IDS=%s", sorted(EXCLUDE_SERVICE_IDS))
    log.info("ALLOW_ZERO_SERVICE_ID=%s", ALLOW_ZERO_SERVICE_ID)
    log.info("SKIP_ACTIVITY_MAPPING_MISS=%s", SKIP_ACTIVITY_MAPPING_MISS)
    log.info("ACTIVITY_FILE=%s", ACTIVITY_FILE)

    activity_map = read_activity_map(ACTIVITY_FILE)

    db_rows = get_counts_by_service_since_days_sql(SINCE_DAYS)

    rows_main, unaccounted_rows, total_included = aggregate_and_enrich(
        db_rows=db_rows,
        exclude_service_ids=EXCLUDE_SERVICE_IDS,
        activity_map=activity_map,
    )

    log.info(
        "RESULT: main_rows=%d unaccounted_rows=%d total_included=%d",
        len(rows_main),
        len(unaccounted_rows),
        total_included,
    )

    write_excel(OUT_XLSX, rows_main, unaccounted_rows)
    log.info("DONE")


if __name__ == "__main__":
    main()