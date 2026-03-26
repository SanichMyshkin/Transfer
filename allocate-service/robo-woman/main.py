import os
import time
import logging
from collections import defaultdict

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from zabbix_utils import ZabbixAPI

load_dotenv()

ZABBIX_URL = os.getenv("ZABBIX_URL", "").rstrip("/")
ZABBIX_TOKEN = os.getenv("ZABBIX_TOKEN", "")
DB_FILE = os.getenv("DB_FILE")
ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "robo-dama_report.xlsx")

TRIGGER_TAGS = [
    "robo_dama",
    "robo_dama_a",
]

EVENT_DAYS = 30
BAN_SERVICE_IDS = [15473]
SKIP_EMPTY_SERVICE_ID = True

ZBX_CHUNK = 200

MAX_TRIGGERIDS_IN_CELL = 200

logger = logging.getLogger("zabbix_events_report")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
handler.setFormatter(
    logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S")
)
logger.handlers.clear()
logger.addHandler(handler)


def die(msg: str):
    logger.error(msg)
    raise SystemExit(2)


def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i : i + n]


def clean_spaces(s):
    if s is None:
        return ""
    s = str(s).strip()
    return " ".join(s.replace(",", " ").split())


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def clean_dns(s):
    return (s or "").strip().lower()


def clean_ip_only_32(s):
    s = (s or "").strip()
    if s.endswith("/32"):
        s = s[:-3].strip()
    return s


def pick_primary_interface(interfaces):
    if not interfaces:
        return {}
    for it in interfaces:
        if str(it.get("main")) == "1":
            return it
    return interfaces[0]


def fetch_active_hosts(api):
    raw = api.host.get(
        output=["hostid", "host", "status"],
        selectInterfaces=["ip", "dns", "main"],
    )
    rows = []
    for h in raw or []:
        if int(h.get("status", 0)) != 0:
            continue
        iface = pick_primary_interface(h.get("interfaces") or [])
        rows.append(
            {
                "hostid": str(h.get("hostid", "")).strip(),
                "host": (h.get("host") or "").strip(),
                "ip": clean_ip_only_32(iface.get("ip")),
                "dns": clean_dns(iface.get("dns")),
            }
        )
    return pd.DataFrame(rows)


def fetch_triggers_by_tags(api, hostids):
    triggerids_all = set()
    host_trigger_ids = defaultdict(set)
    host_tags_found = defaultdict(set)

    for tag in TRIGGER_TAGS:
        logger.info("Получаем триггеры по тегу: %s", tag)
        for part in chunks(hostids, ZBX_CHUNK):
            trigs = api.trigger.get(
                hostids=part,
                output=["triggerid", "status"],
                tags=[{"tag": tag}],
                selectHosts=["hostid"],
            )
            for t in trigs or []:
                if str(t.get("status", "0")) != "0":
                    continue
                tid = str(t.get("triggerid", "")).strip()
                if not tid:
                    continue
                triggerids_all.add(tid)
                for h in t.get("hosts") or []:
                    hid = str(h.get("hostid", "")).strip()
                    if hid:
                        host_trigger_ids[hid].add(tid)
                        host_tags_found[hid].add(tag)

    return sorted(triggerids_all), host_trigger_ids, host_tags_found


def fetch_events_count_by_host(api, triggerids, time_from, time_till):
    host_events = defaultdict(int)
    for part in chunks(triggerids, ZBX_CHUNK):
        evs = api.event.get(
            object=0,
            objectids=part,
            output=["eventid"],
            selectHosts=["hostid"],
            time_from=time_from,
            time_till=time_till,
        )
        for e in evs or []:
            for h in e.get("hosts") or []:
                hid = str(h.get("hostid", "")).strip()
                if hid:
                    host_events[hid] += 1
    return dict(host_events)


def load_db(path):
    df = pd.read_excel(path, usecols="A,B,N,V", dtype=str, engine="openpyxl")
    df.columns = ["service", "service_id", "dns", "ip"]
    df = df.fillna("")
    df["service"] = df["service"].map(clean_spaces)
    df["service_id"] = df["service_id"].map(normalize_code)
    df["ip"] = df["ip"].map(clean_ip_only_32)
    df["dns"] = df["dns"].map(clean_dns)
    return df


def load_activity_map(path):
    if not path or not os.path.isfile(path):
        die(f"ACTIVITY_FILE не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    mp = {}
    for row in ws.iter_rows(values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else None)
        if not code:
            continue
        if code in mp:
            continue
        mp[code] = {
            "service_name": clean_spaces(row[1] if len(row) > 1 else ""),
            "activity_code": clean_spaces(row[2] if len(row) > 2 else ""),
            "activity_name": clean_spaces(row[3] if len(row) > 3 else ""),
        }

    wb.close()
    logger.info("ACTIVITY loaded: %d", len(mp))
    return mp


def build_map(df, keys):
    tmp = df.copy()
    tmp["_k"] = list(map(tuple, tmp[keys].to_numpy()))
    last = tmp.drop_duplicates("_k", keep="last")
    return {
        k: (r["service"], r["service_id"])
        for k, r in zip(last["_k"], last.to_dict("records"))
    }


def main():
    logger.info("Старт отчета: tagged triggers + events + mapping")

    if not ZABBIX_URL:
        die("ENV ZABBIX_URL пустой")
    if not ZABBIX_TOKEN:
        die("ENV ZABBIX_TOKEN пустой")
    if not DB_FILE or not os.path.isfile(DB_FILE):
        die(f"DB_FILE не найден: {DB_FILE}")
    if not ACTIVITY_FILE or not os.path.isfile(ACTIVITY_FILE):
        die(f"ACTIVITY_FILE не найден: {ACTIVITY_FILE}")
    if not TRIGGER_TAGS:
        die("TRIGGER_TAGS пуст")

    now = int(time.time())
    time_from = now - EVENT_DAYS * 86400
    time_till = now
    ban_set = {str(x).strip() for x in BAN_SERVICE_IDS if str(x).strip()}

    logger.info("BAN_SERVICE_IDS=%s", sorted(ban_set) if ban_set else "[]")
    logger.info("EVENT_DAYS=%s", EVENT_DAYS)

    api = ZabbixAPI(url=ZABBIX_URL)
    api.login(token=ZABBIX_TOKEN)

    try:
        df_hosts = fetch_active_hosts(api)
        if df_hosts.empty:
            die("Нет активных хостов")

        hostids = df_hosts["hostid"].tolist()
        logger.info("Активных хостов: %d", len(hostids))

        triggerids_all, host_trigger_ids, host_tags_found = fetch_triggers_by_tags(
            api, hostids
        )
        logger.info("Уникальных триггеров по всем тегам: %d", len(triggerids_all))

        if triggerids_all:
            host_events = fetch_events_count_by_host(
                api, triggerids_all, time_from, time_till
            )
        else:
            host_events = {}
    finally:
        try:
            api.logout()
        except Exception:
            pass

    df_hosts["triggers"] = df_hosts["hostid"].map(
        lambda x: len(host_trigger_ids.get(str(x), set()))
    )
    df_hosts["events"] = df_hosts["hostid"].map(
        lambda x: int(host_events.get(str(x), 0))
    )

    df_db = load_db(DB_FILE)
    activity_map = load_activity_map(ACTIVITY_FILE)

    map_both = build_map(df_db, ["ip", "dns"])
    map_ip = build_map(df_db[df_db["ip"] != ""], ["ip"])
    map_dns = build_map(df_db[df_db["dns"] != ""], ["dns"])

    per_service_events = defaultdict(int)
    per_service_triggers = defaultdict(int)
    unaccounted = []

    df_tagged_hosts = df_hosts[df_hosts["triggers"] > 0].copy()

    for r in df_tagged_hosts.to_dict("records"):
        ip = r["ip"]
        dns = r["dns"]
        trg = int(r["triggers"])
        evc = int(r["events"])
        svc_key = None

        if ip and dns and (ip, dns) in map_both:
            svc_key = map_both[(ip, dns)]
        elif ip and (ip,) in map_ip:
            svc_key = map_ip[(ip,)]
        elif dns and (dns,) in map_dns:
            svc_key = map_dns[(dns,)]

        def add_unacc(reason, detail, service="", service_id="", service_name="", activity_code="", activity_name=""):
            unaccounted.append(
                {
                    "reason": reason,
                    "detail": detail,
                    "Хост": r["host"],
                    "DNS": dns,
                    "IP": ip,
                    "service_from_db": service,
                    "КОД": service_id,
                    "Имя сервиса": service_name,
                    "Код активности": activity_code,
                    "Наименование активности": activity_name,
                    "Кол-во триггеров": trg,
                    "События": evc,
                }
            )

        if not svc_key:
            add_unacc("no_mapping", "no match by (ip,dns) / ip / dns")
            continue

        service, service_id = svc_key
        service_id = normalize_code(service_id)

        meta = activity_map.get(service_id, {})
        service_name = meta.get("service_name", "")
        activity_code = meta.get("activity_code", "")
        activity_name = meta.get("activity_name", "")

        if SKIP_EMPTY_SERVICE_ID and not service_id:
            add_unacc(
                "empty_service_id",
                "SKIP_EMPTY_SERVICE_ID=True and service_id empty",
                service=service,
                service_id=service_id,
                service_name=service_name,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        if service_id in ban_set:
            add_unacc(
                "banned_service_id",
                "service_id in BAN_SERVICE_IDS",
                service=service,
                service_id=service_id,
                service_name=service_name,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        if service_id not in activity_map:
            add_unacc(
                "activity_mapping_miss",
                "service_id not found in activity.xlsx",
                service=service,
                service_id=service_id,
                service_name=service_name,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        key = (service_name or service, service_id, activity_code, activity_name)
        per_service_events[key] += evc
        per_service_triggers[key] += trg

    rows = []
    for key, trg_sum in per_service_triggers.items():
        service_name, service_id, activity_code, activity_name = key
        ev_sum = per_service_events[key]
        rows.append(
            {
                "Имя сервиса": service_name,
                "Код": service_id,
                "Код активности": activity_code,
                "Наименование активности": activity_name,
                "Кол-во триггеров": int(trg_sum),
                "События": int(ev_sum),
            }
        )

    total_events = sum(x["События"] for x in rows)
    for x in rows:
        x["% потребления (events)"] = (
            round(x["События"] / total_events, 4) if total_events else 0.0
        )

    if rows:
        df_report = pd.DataFrame(rows).sort_values(
            ["События", "Кол-во триггеров"],
            ascending=[False, False],
        )
    else:
        df_report = pd.DataFrame(
            columns=[
                "Имя сервиса",
                "Код",
                "Код активности",
                "Наименование активности",
                "Кол-во триггеров",
                "События",
                "% потребления (events)",
            ]
        )

    if unaccounted:
        df_unaccounted = pd.DataFrame(unaccounted).sort_values(
            ["reason", "События", "Кол-во триггеров"], ascending=[True, False, False]
        )
    else:
        df_unaccounted = pd.DataFrame(
            columns=[
                "reason",
                "detail",
                "Хост",
                "DNS",
                "IP",
                "service_from_db",
                "КОД",
                "Имя сервиса",
                "Код активности",
                "Наименование активности",
                "Кол-во триггеров",
                "События",
            ]
        )

    logger.info("Итого событий (в отчете): %d", total_events)
    logger.info("Сервисов в отчете: %d", len(df_report))
    logger.info("Unaccounted rows: %d", len(df_unaccounted))

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df_report.to_excel(writer, index=False, sheet_name="Отчет Zabbix")
        df_unaccounted.to_excel(writer, index=False, sheet_name="Unaccounted")


if __name__ == "__main__":
    main()