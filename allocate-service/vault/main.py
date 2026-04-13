import os
import logging
import urllib3
import re

import requests
import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook
from dotenv import load_dotenv

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger("vault_report")

load_dotenv()

VAULT_ADDR = os.getenv("VAULT_ADDR")
ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")
OUT_FILE = os.getenv("OUT_FILE", "vault_report.xlsx")

BAN_SERVICE_IDS = [15473]

SKIP_EMPTY_SECRETS = True
SKIP_ZERO_SERVICE_IDS = True


def die(msg: str, code: int = 2):
    log.error(msg)
    raise SystemExit(code)


def build_ban_set(ban_list):
    if not isinstance(ban_list, (list, tuple, set)):
        die("BAN_SERVICE_IDS должен быть list / tuple / set")
    return {str(x).strip() for x in ban_list if str(x).strip()}


ban_set = build_ban_set(BAN_SERVICE_IDS)


def clean_spaces(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def get_vault_metrics_prometheus() -> str:
    url = f"{VAULT_ADDR}/v1/sys/metrics?format=prometheus"
    log.info("Запрашиваю метрики Vault: %s", url)
    r = requests.get(url, verify=False, timeout=20)
    r.raise_for_status()
    return r.text


def parse_kv_metrics(metrics_text: str) -> pd.DataFrame:
    pattern = re.compile(
        r'vault[_\s]*secret[_\s]*kv[_\s]*count\s*\{[^}]*mount_point="([^"]+)"[^}]*\}\s+(\d+)',
        re.IGNORECASE,
    )

    rows = []
    for m in pattern.finditer(metrics_text):
        kv = (m.group(1) or "").rstrip("/")
        if "test" in kv.lower():
            continue
        rows.append({"kv": kv, "secrets": int(m.group(2))})

    return pd.DataFrame(rows)


def read_activity_map(path: str) -> pd.DataFrame:
    if not path:
        die("ACTIVITY_FILE не задан")
    if not os.path.exists(path):
        die(f"ACTIVITY_FILE не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    rows = []
    for row in ws.iter_rows(values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else None)
        if not code:
            continue

        rows.append(
            {
                "code": code,
                "service_name": clean_spaces(row[1] if len(row) > 1 else ""),
                "activity_code": clean_spaces(row[2] if len(row) > 2 else ""),
                "activity_name": clean_spaces(row[3] if len(row) > 3 else ""),
            }
        )

    wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(columns=["code", "service_name", "activity_code", "activity_name"])

    out = out.drop_duplicates(subset=["code"], keep="first").copy()
    log.info("ACTIVITY загружено: %d", len(out))
    return out


def main():
    if not VAULT_ADDR:
        die("Не задан VAULT_ADDR")

    log.info("==== START ====")
    log.info("VAULT_ADDR=%s", VAULT_ADDR)
    log.info("SKIP_EMPTY_SECRETS=%s", SKIP_EMPTY_SECRETS)
    log.info("SKIP_ZERO_SERVICE_IDS=%s", SKIP_ZERO_SERVICE_IDS)
    log.info("BAN_SERVICE_IDS=%s", sorted(ban_set) if ban_set else "[]")

    activity_df = read_activity_map(ACTIVITY_FILE)

    metrics = get_vault_metrics_prometheus()
    base_df = parse_kv_metrics(metrics)

    if base_df.empty:
        die("Нет данных KV в метриках (после пропуска test)")

    log.info("KV строк после парсинга метрик: %d", len(base_df))

    unacc_rows = []

    base_df["code"] = base_df["kv"].astype(str).str.extract(r"(\d+)$", expand=False)

    # --- no_code ---
    no_code = base_df[base_df["code"].isna()].copy()
    if not no_code.empty:
        no_code["reason"] = "no_code_in_kv"
        no_code["detail"] = "cannot extract code via r'(\\d+)$' from kv"
        no_code["service_name"] = ""
        no_code["activity_code"] = ""
        no_code["activity_name"] = ""
        unacc_rows.append(no_code)
        log.info("Unaccounted no_code_in_kv: %d", len(no_code))

    df = base_df[base_df["code"].notna()].copy()
    df["code"] = df["code"].astype(str)

    # --- zero_service_id ---
    if SKIP_ZERO_SERVICE_IDS:
        zero_code = df[df["code"].str.fullmatch(r"0+")].copy()
        if not zero_code.empty:
            zero_code = zero_code.merge(activity_df, on="code", how="left")
            zero_code["reason"] = "zero_service_id"
            zero_code["detail"] = "service_id consists only of zeros"
            unacc_rows.append(zero_code)
            log.info("Unaccounted zero_service_id: %d", len(zero_code))

        df = df[~df["code"].str.fullmatch(r"0+")].copy()

    # --- banned ---
    banned = df[df["code"].isin(ban_set)].copy()
    if not banned.empty:
        banned = banned.merge(activity_df, on="code", how="left")
        banned["reason"] = "banned_service_id"
        banned["detail"] = "code in BAN_SERVICE_IDS"
        unacc_rows.append(banned)
        log.info("Unaccounted banned_service_id: %d", len(banned))

    df = df[~df["code"].isin(ban_set)].copy()

    # --- empty secrets ---
    if SKIP_EMPTY_SECRETS:
        empty_sec = df[df["secrets"] <= 0].copy()
        if not empty_sec.empty:
            empty_sec = empty_sec.merge(activity_df, on="code", how="left")
            empty_sec["reason"] = "empty_secrets"
            empty_sec["detail"] = "SKIP_EMPTY_SECRETS=True and secrets<=0"
            unacc_rows.append(empty_sec)
            log.info("Unaccounted empty_secrets: %d", len(empty_sec))
        df = df[df["secrets"] > 0].copy()

    if df.empty:
        die("Нет данных KV после фильтров")

    out = df.merge(activity_df, on="code", how="left")

    activity_miss = out[out["service_name"].map(clean_spaces) == ""].copy()
    if not activity_miss.empty:
        activity_miss["reason"] = "activity_mapping_miss"
        activity_miss["detail"] = "code not found in activity.xlsx"
        unacc_rows.append(activity_miss)
        log.info("Unaccounted activity_mapping_miss: %d", len(activity_miss))

    out = out[out["service_name"].map(clean_spaces) != ""].copy()

    if out.empty:
        die("Нет данных после маппинга activity.xlsx")

    total = int(out["secrets"].sum()) or 1
    out["percent"] = out["secrets"] / total

    out = out.rename(
        columns={
            "service_name": "Имя сервиса",
            "code": "Код",
            "activity_code": "Код активности",
            "activity_name": "Наименование активности",
            "secrets": "Кол-во секретов",
            "percent": "% потребления",
        }
    )

    out = out.sort_values("Кол-во секретов", ascending=False).reset_index(drop=True)

    unaccounted = (
        pd.concat(unacc_rows, ignore_index=True) if unacc_rows else pd.DataFrame()
    )

    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Отчет Vault")
        ws = writer.book["Отчет Vault"]
        for c in ws[1]:
            c.font = Font(bold=True)

        if not unaccounted.empty:
            unaccounted.to_excel(writer, index=False, sheet_name="Unaccounted")
            ws_u = writer.book["Unaccounted"]
            for c in ws_u[1]:
                c.font = Font(bold=True)

    log.info("==== DONE ====")


if __name__ == "__main__":
    main()