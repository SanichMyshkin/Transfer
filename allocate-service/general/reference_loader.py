from pathlib import Path
from openpyxl import load_workbook


def normalize_percent(value):
    if value is None:
        return None

    s = str(value).strip().replace("%", "").replace(",", ".")
    if not s:
        return None

    try:
        return float(s)
    except ValueError:
        return None


def load_reference_rows(
    file_path,
    sheet_name=None,
    service_name_col=2,
    service_code_col=3,
    owner_col=4,
    percent_col=6,
    header_row=1,
):
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(path, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{sheet_name}' not found in {file_path}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    result = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        service_name = row[service_name_col - 1] if len(row) >= service_name_col else None
        service_code = row[service_code_col - 1] if len(row) >= service_code_col else None
        owner = row[owner_col - 1] if len(row) >= owner_col else None
        percent = row[percent_col - 1] if len(row) >= percent_col else None

        if service_name is None and service_code is None:
            continue

        service_name = str(service_name).strip() if service_name is not None else ""
        service_code = str(service_code).strip() if service_code is not None else ""
        owner = str(owner).strip() if owner is not None else ""

        if not service_name and not service_code:
            continue

        service_display = service_name
        if service_code:
            service_display = f"{service_name} ({service_code})"

        result.append(
            {
                "service_name": service_name,
                "service_code": service_code,
                "service_display": service_display,
                "owner": owner,
                "percent": normalize_percent(percent),
            }
        )

    return result