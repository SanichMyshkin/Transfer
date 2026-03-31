import os
import re
import logging

import requests
import urllib3
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)

log = logging.getLogger(__name__)

CONF_URL = os.getenv("CONF_URL")
CONF_PAGE_ID = os.getenv("CONF_PAGE_ID")
CONF_USER = os.getenv("CONF_USER")
CONF_PASS = os.getenv("CONF_PASS")

REFERENCES_DIR = os.getenv("REFERENCES_DIR", "references")
OUT_FILE = os.getenv("OUT_FILE", "employees.xlsx")


def set_bold_row(ws, row, cols):
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).font = Font(bold=True)


def clean_spaces(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_code(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def try_parse_percent(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        num = float(value)
        if num > 1:
            return num / 100
        return num

    s = str(value).strip().replace(",", ".")

    if "/" in s:
        return None

    if s.endswith("%"):
        s = s[:-1]

    if not s:
        return None

    if not re.fullmatch(r"\d+(\.\d+)?", s):
        return None

    num = float(s)

    if num > 1:
        return num / 100

    return num


def fetch_confluence_table():
    url = f"{CONF_URL.rstrip('/')}/rest/api/content/{CONF_PAGE_ID}"
    params = {"expand": "body.storage"}

    log.info("Loading Confluence table")

    r = requests.get(
        url,
        params=params,
        auth=(CONF_USER, CONF_PASS),
        timeout=30,
        verify=False,
    )
    r.raise_for_status()

    html = r.json()["body"]["storage"]["value"]

    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    rows = table.find_all("tr")

    headers = [c.get_text(strip=True) for c in rows[0].find_all(["th", "td"])]
    service_headers = headers[3:]

    data_rows = []

    for tr in rows[1:]:
        cols = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]

        if not cols:
            continue

        employee = cols[0]
        load_value = cols[1] if len(cols) > 1 else ""
        service_values = cols[3:]

        if len(service_values) < len(service_headers):
            service_values += [""] * (len(service_values) - len(service_headers))

        data_rows.append(
            {
                "employee": employee,
                "load": load_value,
                "services": dict(zip(service_headers, service_values)),
            }
        )

    return headers, service_headers, data_rows


def write_employees_sheet(ws, headers, service_headers, data_rows):
    ws.append(headers)
    set_bold_row(ws, 1, len(headers))

    platform_col_map = {}

    service_start = 4
    service_end = service_start + len(service_headers) - 1

    for idx, service in enumerate(service_headers, start=service_start):
        platform_col_map[service] = idx

    for row_idx, item in enumerate(data_rows, start=2):
        row = [item["employee"], item["load"], None]

        for service in service_headers:
            raw = item["services"].get(service, "")
            parsed = try_parse_percent(raw)
            row.append(parsed if parsed is not None else raw)

        ws.append(row)

        start_letter = get_column_letter(service_start)
        end_letter = get_column_letter(service_end)

        total_cell = ws.cell(row=row_idx, column=3)
        total_cell.value = f"=SUM({start_letter}{row_idx}:{end_letter}{row_idx})"
        total_cell.number_format = "0%"

    sum_row = ws.max_row + 1
    ws.cell(row=sum_row, column=1, value="SUM")

    for col in range(service_start, service_end + 1):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=sum_row, column=col)
        cell.value = f"=SUM({col_letter}2:{col_letter}{sum_row - 1})"
        cell.number_format = "0%"

    set_bold_row(ws, sum_row, len(headers))

    return platform_col_map, sum_row


def get_cell(row, idx_1_based):
    if not idx_1_based or idx_1_based < 1:
        return None
    pos = idx_1_based - 1
    if pos >= len(row):
        return None
    return row[pos]


def load_reference_rows(
    file_path,
    service_id_col,
    service_name_col,
    activity_code_col,
    activity_name_col,
    percent_col,
    header_row=1,
):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    result = []
    rows_seen = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < header_row + 1:
            continue

        rows_seen += 1

        service_id = normalize_code(get_cell(row, service_id_col))
        service_name = clean_spaces(get_cell(row, service_name_col))
        activity_code = clean_spaces(get_cell(row, activity_code_col))
        activity_name = clean_spaces(get_cell(row, activity_name_col))
        percent = try_parse_percent(get_cell(row, percent_col))

        if not service_id and not service_name:
            continue

        result.append(
            {
                "service_id": service_id,
                "service_name": service_name,
                "activity_code": activity_code,
                "activity_name": activity_name,
                "percent": 0.0 if percent is None else percent,
            }
        )

    wb.close()

    log.info(
        "Loaded rows from %s: scanned=%d loaded=%d",
        file_path,
        rows_seen,
        len(result),
    )

    return result


def build_source_rows(rows, source_name):
    result = []

    for r in rows:
        service_id = normalize_code(r.get("service_id", ""))
        service_name = clean_spaces(r.get("service_name", ""))
        activity_code = clean_spaces(r.get("activity_code", ""))
        activity_name = clean_spaces(r.get("activity_name", ""))
        percent = r.get("percent")

        if not service_id and not service_name:
            continue

        result.append(
            {
                "service_id": service_id,
                "service_name": service_name,
                "activity_code": activity_code,
                "activity_name": activity_name,
                f"{source_name}_percent": 0.0 if percent is None else percent,
            }
        )

    return result


def merge_source_rows(rows_list):
    merged = {}

    for rows in rows_list:
        for item in rows:
            key = (
                item["service_id"],
                item["service_name"],
                item["activity_code"],
                item["activity_name"],
            )

            if key not in merged:
                merged[key] = {
                    "service_id": item["service_id"],
                    "service_name": item["service_name"],
                    "activity_code": item["activity_code"],
                    "activity_name": item["activity_name"],
                }

            for k, v in item.items():
                if k in ("service_id", "service_name", "activity_code", "activity_name"):
                    continue
                merged[key][k] = 0.0 if v is None else v

    return list(merged.values())


def write_sources_sheet(ws, merged_rows, source_configs, platform_map, sum_row):
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)

    ws.cell(row=1, column=1, value="service_id").font = Font(bold=True)
    ws.cell(row=1, column=2, value="service_name").font = Font(bold=True)
    ws.cell(row=1, column=3, value="activity_code").font = Font(bold=True)
    ws.cell(row=1, column=4, value="activity_name").font = Font(bold=True)

    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=4).alignment = Alignment(horizontal="center", vertical="center")

    current_col = 5

    for cfg in source_configs:
        source = cfg["name"]
        subtitle = cfg["subtitle"]

        ws.merge_cells(
            start_row=1,
            start_column=current_col,
            end_row=1,
            end_column=current_col + 1,
        )
        ws.merge_cells(
            start_row=2,
            start_column=current_col,
            end_row=2,
            end_column=current_col + 1,
        )

        head = ws.cell(row=1, column=current_col, value=source)
        head.font = Font(bold=True)
        head.alignment = Alignment(horizontal="center", vertical="center")

        desc = ws.cell(row=2, column=current_col, value=subtitle)
        desc.font = Font(bold=True)
        desc.alignment = Alignment(horizontal="center", vertical="center")

        current_col += 2

    for row_idx, item in enumerate(merged_rows, start=3):
        ws.cell(row=row_idx, column=1, value=item["service_id"])
        ws.cell(row=row_idx, column=2, value=item["service_name"])
        ws.cell(row=row_idx, column=3, value=item["activity_code"])
        ws.cell(row=row_idx, column=4, value=item["activity_name"])

        current_col = 5

        for cfg in source_configs:
            source = cfg["name"]

            percent_value = item.get(f"{source}_percent")
            if percent_value is None:
                percent_value = 0.0

            percent_cell = ws.cell(row=row_idx, column=current_col, value=percent_value)
            percent_cell.number_format = "0.00000"

            weight_cell = ws.cell(row=row_idx, column=current_col + 1)

            platform_col = platform_map.get(source)
            if platform_col:
                platform_letter = get_column_letter(platform_col)
                percent_letter = get_column_letter(current_col)
                weight_cell.value = (
                    f"=Employees!{platform_letter}{sum_row}"
                    f"*{percent_letter}{row_idx}"
                )
            else:
                weight_cell.value = 0.0

            weight_cell.number_format = "0.00000"

            current_col += 2


def main():
    headers, service_headers, employees = fetch_confluence_table()

    source_configs = [
        {
            "name": "Nexus",
            "subtitle": "Объем репозиториев",
            "file_name": "nexus.xlsx",
            "service_id_col": 3,
            "service_name_col": 2,
            "activity_code_col": 4,
            "activity_name_col": 5,
            "percent_col": 6,
            "header_row": 1,
        },
        {
            "name": "Gitlab",
            "subtitle": "Объем проектов",
            "file_name": "gitlab.xlsx",
            "service_id_col": 3,
            "service_name_col": 2,
            "activity_code_col": 4,
            "activity_name_col": 5,
            "percent_col": 11,
            "header_row": 1,
        },
        {
            "name": "Sender",
            "subtitle": "Сообщения",
            "file_name": "sender.xlsx",
            "service_id_col": 1,
            "service_name_col": 2,
            "activity_code_col": 3,
            "activity_name_col": 4,
            "percent_col": 8,
            "header_row": 1,
        },
    ]

    all_rows_for_merge = []

    for cfg in source_configs:
        log.info("Loading %s data", cfg["name"])

        raw_rows = load_reference_rows(
            file_path=os.path.join(REFERENCES_DIR, cfg["file_name"]),
            service_id_col=cfg["service_id_col"],
            service_name_col=cfg["service_name_col"],
            activity_code_col=cfg["activity_code_col"],
            activity_name_col=cfg["activity_name_col"],
            percent_col=cfg["percent_col"],
            header_row=cfg["header_row"],
        )

        cfg["rows"] = build_source_rows(raw_rows, cfg["name"])
        all_rows_for_merge.append(cfg["rows"])

    merged_rows = merge_source_rows(all_rows_for_merge)

    for item in merged_rows:
        for cfg in source_configs:
            key = f"{cfg['name']}_percent"
            if key not in item or item[key] is None:
                item[key] = 0.0

    merged_rows.sort(
        key=lambda x: (
            x["service_name"] or "",
            x["service_id"] or "",
            x["activity_code"] or "",
            x["activity_name"] or "",
        )
    )

    wb = Workbook()

    ws_emp = wb.active
    ws_emp.title = "Employees"

    platform_map, sum_row = write_employees_sheet(
        ws_emp,
        headers,
        service_headers,
        employees,
    )

    ws_sources = wb.create_sheet("Sources")

    write_sources_sheet(
        ws_sources,
        merged_rows,
        source_configs,
        platform_map,
        sum_row,
    )

    wb.save(OUT_FILE)
    log.info("Report saved to %s", OUT_FILE)


if __name__ == "__main__":
    main()