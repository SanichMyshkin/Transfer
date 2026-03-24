import os
import sys
import logging
import urllib3
import pandas as pd
from dotenv import load_dotenv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from jenkins_client import JenkinsGroovyClient

logger = logging.getLogger("jenkins_report")
logger.setLevel(logging.INFO)
formatter = logging.Formatter(
    "%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S"
)
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(formatter)
logger.handlers.clear()
logger.addHandler(console_handler)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()

JENKINS_URL = os.getenv("JENKINS_URL")
USER = os.getenv("USER")
TOKEN = os.getenv("TOKEN")

EXCLUDE_PROJECTS_WITHOUT_TEAM_NUMBER = True

ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")
OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "jenkins_report.xlsx")

BAN_SERVICE_IDS = [15473]

SCRIPT_BUILDS = r"""
import jenkins.model.Jenkins
import groovy.json.JsonOutput
def jobs = Jenkins.instance.getAllItems()
def jobList = []
jobs.each { j ->
    def info = [
        name: j.fullName,
        isFolder: j.class.simpleName.contains("Folder")
    ]
    try {
        if (!info.isFolder && j.metaClass.respondsTo(j, "getBuilds")) {
            def builds = j.getBuilds()
            info.buildCount = (builds != null) ? builds.size() : 0
        } else {
            info.buildCount = 0
        }
    } catch (Exception e) {
        info.buildCount = 0
        info.error = e.message
    }
    jobList << info
}
JsonOutput.toJson([jobs: jobList, total: jobs.size()])
"""

client = JenkinsGroovyClient(JENKINS_URL, USER, TOKEN, is_https=False)


def die(msg: str, code: int = 2):
    logger.error(msg)
    raise SystemExit(code)


def clean_spaces(s: str) -> str:
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def build_ban_set(ban_list):
    if not isinstance(ban_list, (list, tuple, set)):
        die("BAN_SERVICE_IDS должен быть list / tuple / set")
    return {str(x).strip() for x in ban_list if str(x).strip()}


ban_set = build_ban_set(BAN_SERVICE_IDS)


def validate_env_and_files():
    logger.info("Проверяем ENV и входные файлы...")

    if not JENKINS_URL:
        die("ENV JENKINS_URL пустой")
    if not USER:
        die("ENV USER пустой")
    if not TOKEN:
        die("ENV TOKEN пустой")

    if not ACTIVITY_FILE or not os.path.isfile(ACTIVITY_FILE):
        die(f"ACTIVITY_FILE не найден: {ACTIVITY_FILE}")

    out_dir = os.path.dirname(OUTPUT_XLSX)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    logger.info(f"Бан-лист (КОД): {sorted(ban_set) if ban_set else 'пусто'}")
    logger.info(
        f"EXCLUDE_PROJECTS_WITHOUT_TEAM_NUMBER={EXCLUDE_PROJECTS_WITHOUT_TEAM_NUMBER}"
    )
    logger.info("ENV/файлы ок.")


def get_builds_inventory():
    logger.info("Получаем инвентарь Jenkins (jobs + buildCount)...")
    data = client.run_script(SCRIPT_BUILDS)
    total = data.get("total")
    jobs_len = len(data.get("jobs", []) or [])
    logger.info(f"Jenkins вернул items(total): {total}, jobs в payload: {jobs_len}")
    return data


def split_project_and_team(project_raw: str):
    if not project_raw:
        return "", ""
    parts = project_raw.split("-")
    if len(parts) >= 2 and parts[-1].isdigit():
        return "-".join(parts[:-1]), parts[-1]
    return project_raw, ""


def normalize_number(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if not s:
        return None
    if "." in s and s.replace(".", "", 1).isdigit():
        try:
            f = float(s)
            if f.is_integer():
                s = str(int(f))
        except Exception:
            pass
    return s


def load_activity_mapping(path: str):
    logger.info("Читаем ACTIVITY (A=КОД, B=Наименование сервиса, C=Код активности, D=Наименование активности)...")

    df = pd.read_excel(path, dtype=str, engine="openpyxl").fillna("")

    map_by_number = {}

    for _, row in df.iterrows():
        service_id = normalize_number(row.iloc[0] if len(row) > 0 else "")
        service_name = clean_spaces(row.iloc[1] if len(row) > 1 else "")
        activity_code = clean_spaces(row.iloc[2] if len(row) > 2 else "")
        activity_name = clean_spaces(row.iloc[3] if len(row) > 3 else "")

        if not service_id:
            continue

        if service_id in map_by_number:
            continue

        map_by_number[service_id] = {
            "service_name": service_name,
            "activity_code": activity_code,
            "activity_name": activity_name,
        }

    logger.info(f"ACTIVITY: загружено сервисов по КОД: {len(map_by_number)}")
    return map_by_number


def aggregate_builds_by_service(
    data, activity_map, exclude_without_team=True
):
    logger.info("Агрегируем билды по сервису...")

    acc = defaultdict(int)
    unaccounted = []

    def add_unaccounted(
        root_project,
        team_number,
        full_name,
        builds,
        reason,
        detail,
        service_name="",
        activity_code="",
        activity_name="",
    ):
        unaccounted.append(
            {
                "root_project": root_project,
                "team_number": team_number,
                "job_full_name": full_name,
                "buildCount": builds,
                "reason": reason,
                "detail": detail,
                "service_name": service_name,
                "activity_code": activity_code,
                "activity_name": activity_name,
            }
        )

    for j in data.get("jobs", []) or []:
        if j.get("isFolder"):
            continue

        full_name = (j.get("name") or "").strip()
        if not full_name:
            continue

        root = full_name.split("/", 1)[0].strip()
        if not root:
            continue

        proj, team_number = split_project_and_team(root)
        builds = int(j.get("buildCount") or 0)

        if exclude_without_team and not team_number:
            add_unaccounted(
                root_project=proj or root,
                team_number="",
                full_name=full_name,
                builds=builds,
                reason="no_team_number",
                detail="root project name does not end with -<digits> (or exclude_without_team=True)",
            )
            continue

        if team_number and team_number in ban_set:
            meta = activity_map.get(team_number, {})
            add_unaccounted(
                root_project=proj or root,
                team_number=team_number,
                full_name=full_name,
                builds=builds,
                reason="banned_service_id",
                detail="team_number in BAN_SERVICE_IDS",
                service_name=meta.get("service_name", ""),
                activity_code=meta.get("activity_code", ""),
                activity_name=meta.get("activity_name", ""),
            )
            continue

        if not team_number:
            add_unaccounted(
                root_project=proj or root,
                team_number="",
                full_name=full_name,
                builds=builds,
                reason="no_team_number_included",
                detail="team_number is empty but exclude_without_team=False; not attributable to service",
            )
            continue

        acc[team_number] += builds

    candidates = []
    filtered_activity_miss = 0

    for team_number, builds in acc.items():
        meta = activity_map.get(team_number)
        service_name = (meta or {}).get("service_name", "")
        activity_code = (meta or {}).get("activity_code", "")
        activity_name = (meta or {}).get("activity_name", "")

        if not meta or not clean_spaces(service_name):
            filtered_activity_miss += 1
            add_unaccounted(
                root_project="",
                team_number=team_number,
                full_name="",
                builds=builds,
                reason="activity_mapping_miss",
                detail="no match in activity.xlsx for this team_number",
                service_name=service_name,
                activity_code=activity_code,
                activity_name=activity_name,
            )
            continue

        candidates.append(
            {
                "service_name": service_name,
                "team_number": team_number,
                "activity_code": activity_code,
                "activity_name": activity_name,
                "builds": builds,
            }
        )

    eligible_total_builds = sum(x["builds"] for x in candidates)

    rows = []
    for x in sorted(candidates, key=lambda d: d["builds"], reverse=True):
        builds = x["builds"]
        pct = (
            (builds / eligible_total_builds * 100.0)
            if eligible_total_builds > 0
            else 0.0
        )
        rows.append(
            [
                x["service_name"],
                x["team_number"],
                x["activity_code"],
                x["activity_name"],
                builds,
                round(pct, 2),
            ]
        )

    logger.info(
        f"Фильтры: activity_miss={filtered_activity_miss}. "
        f"Итого учтенных сервисов={len(candidates)}"
    )

    return rows, unaccounted


def export_excel(rows, unaccounted_rows, filename):
    logger.info("Формируем Excel...")

    wb = Workbook()

    ws = wb.active
    ws.title = "Отчет Jenkins"

    headers = [
        "Имя сервиса",
        "Код",
        "Код активности",
        "Наименование активности",
        "Кол-во билдов",
        "% от общего кол-ва билдов",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    for r in rows:
        ws.append(r)

    ws2 = wb.create_sheet("Unaccounted")
    headers2 = [
        "root_project",
        "team_number",
        "job_full_name",
        "buildCount",
        "reason",
        "detail",
        "service_name",
        "activity_code",
        "activity_name",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = bold

    for r in unaccounted_rows:
        ws2.append(
            [
                r.get("root_project", ""),
                r.get("team_number", ""),
                r.get("job_full_name", ""),
                int(r.get("buildCount") or 0),
                r.get("reason", ""),
                r.get("detail", ""),
                r.get("service_name", ""),
                r.get("activity_code", ""),
                r.get("activity_name", ""),
            ]
        )

    wb.save(filename)
    logger.info(f"Excel сохранён: {filename}")


def main():
    try:
        validate_env_and_files()

        activity_map = load_activity_mapping(ACTIVITY_FILE)

        data = get_builds_inventory()

        rows, unaccounted = aggregate_builds_by_service(
            data,
            activity_map=activity_map,
            exclude_without_team=EXCLUDE_PROJECTS_WITHOUT_TEAM_NUMBER,
        )

        export_excel(rows, unaccounted, OUTPUT_XLSX)

        logger.info(
            f"Инвентаризация билдов завершена успешно. accounted={len(rows)} unaccounted={len(unaccounted)}"
        )
    except Exception as e:
        logger.exception(f"Ошибка при инвентаризации: {e}")


if __name__ == "__main__":
    main()