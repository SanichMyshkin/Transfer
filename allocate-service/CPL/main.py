import os
import re
import logging
from collections import defaultdict
from typing import Any, Optional, Tuple
from urllib.parse import urlparse

import humanize
from dotenv import load_dotenv
from opensearchpy import OpenSearch
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

load_dotenv()

SD_FILE = os.getenv("SD_FILE")
BK_FILE = os.getenv("BK_FILE")

OPENSEARCH_URL = os.getenv("OPENSEARCH_URL")
OPENSEARCH_PORT = int(os.getenv("OPENSEARCH_PORT", "9200"))

USER = os.getenv("USER")
PASS = os.getenv("PASS")

BAN_SERVICE_IDS = {15473}

BAN_BUSINESS_TYPES: set[str] = {
    # "Розница",
}

SKIP_UNKNOWN_SERVICE_IDS: bool = True

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("cpl")

TEAM_SERVICE_RE = re.compile(r"^index_([a-z0-9\-]+)-(\d+)_", re.IGNORECASE)
DIGITS_RE = re.compile(r"(\d+)")
TAIL_SERVICE_RE = re.compile(r"-(\d+)(?:_\d+)?$", re.IGNORECASE)


def clean_spaces(s: str) -> str:
    return " ".join((s or "").replace(",", " ").split()).strip()


def humanize_bytes(n: int) -> str:
    return humanize.naturalsize(int(n or 0), binary=True)


def parse_host_and_ssl(raw: str, default_port: int) -> Tuple[str, int, bool]:
    if not raw:
        raise RuntimeError("OPENSEARCH_URL не задан")

    if raw.startswith("http://") or raw.startswith("https://"):
        u = urlparse(raw)
        return u.hostname, u.port or default_port, u.scheme == "https"

    return raw, default_port, True


def normalize_index_name(index_name: str) -> Optional[Tuple[str, int, str]]:
    name = (index_name or "").strip()
    m = TEAM_SERVICE_RE.match(name)
    if not m:
        return None

    team = m.group(1).lower()
    head_service_id = int(m.group(2))

    if team == "ib":
        t = TAIL_SERVICE_RE.search(name)
        if t:
            return team, int(t.group(1)), "ib-tail"
        return team, head_service_id, "ib-head"

    return team, head_service_id, "default"


def read_sd_map(path: str) -> dict[int, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    sd: dict[int, dict[str, str]] = {}
    for row in ws.iter_rows(values_only=True):
        m = DIGITS_RE.search(str(row[1] or ""))
        if not m:
            continue

        service_id = int(m.group(1))
        if service_id in sd:
            continue

        sd[service_id] = {
            "service_name": clean_spaces(row[3]),
            "owner": clean_spaces(row[7]),
        }

    wb.close()
    log.info(f"SD loaded: {len(sd)} services")
    return sd


def read_bk_map(path: str) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    out: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 45:
            continue

        fio = clean_spaces(f"{row[1]} {row[0]} {row[2]}")
        business_type = clean_spaces(row[44])

        if not fio or not business_type:
            continue

        out.setdefault(fio, business_type)

    wb.close()
    log.info(f"BK loaded: {len(out)} owners")
    return out


def fetch_and_aggregate(client: OpenSearch) -> list[dict[str, Any]]:
    raw = client.cat.indices(format="json", bytes="b")

    acc = defaultdict(int)
    parsed_cnt = 0
    skipped_cnt = 0
    banned_cnt = 0

    for idx in raw:
        index_name = idx.get("index") or ""
        parsed = normalize_index_name(index_name)
        if not parsed:
            skipped_cnt += 1
            continue

        team, service_id, rule = parsed

        if service_id in BAN_SERVICE_IDS:
            banned_cnt += 1
            continue

        size_b = int(idx.get("store.size", 0))
        acc[service_id] += size_b
        parsed_cnt += 1

        log.info(
            f"INDEX_MAP | rule={rule} | index={index_name} "
            f"-> team={team} service_id={service_id} size={size_b}B"
        )

    rows = [
        {
            "service_id": service_id,
            "total_bytes": total,
        }
        for service_id, total in acc.items()
    ]

    log.info(
        f"OpenSearch parsed={parsed_cnt} skipped={skipped_cnt} "
        f"banned={banned_cnt} aggregated_services={len(rows)}"
    )
    return rows


def enrich(
    rows: list[dict[str, Any]],
    sd: dict[int, dict[str, str]],
    bk: dict[str, str],
) -> list[dict[str, Any]]:
    for r in rows:
        meta = sd.get(r["service_id"], {})
        r["service_name"] = meta.get("service_name", "")
        r["owner"] = meta.get("owner", "")

        owner_norm = clean_spaces(r["owner"])
        r["business_type"] = clean_spaces(bk.get(owner_norm, ""))

    return rows


def apply_unknown_service_filter(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not SKIP_UNKNOWN_SERVICE_IDS:
        return rows

    kept: list[dict[str, Any]] = []
    skipped = 0

    for r in rows:
        service_id = int(r.get("service_id") or 0)
        if service_id == 0 or not clean_spaces(r.get("service_name", "")):
            skipped += 1
            log.info(
                f"SKIP_UNKNOWN | service_id={service_id} "
                f"size={int(r.get('total_bytes') or 0)}B"
            )
            continue
        kept.append(r)

    log.info(f"Unknown services skipped: {skipped} (kept {len(kept)})")
    return kept


def apply_business_type_ban(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not BAN_BUSINESS_TYPES:
        return rows

    kept: list[dict[str, Any]] = []
    banned = 0

    for r in rows:
        bt = clean_spaces(r.get("business_type", ""))
        if bt in BAN_BUSINESS_TYPES:
            banned += 1
            log.info(
                f"BAN_BT | business_type={bt!r} | "
                f"service_id={r.get('service_id')} "
                f"service_name={r.get('service_name')!r}"
            )
            continue
        kept.append(r)

    log.info(f"BusinessType banned rows: {banned} (kept {len(kept)})")
    return kept


def finalize(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    total_all = sum(int(r["total_bytes"]) for r in rows)

    for r in rows:
        r["size_human"] = humanize_bytes(r["total_bytes"])
        r["pct"] = (r["total_bytes"] / total_all) if total_all else 0

    rows.sort(key=lambda x: x["total_bytes"], reverse=True)
    log.info(f"TOTAL (after filters): {humanize_bytes(total_all)}")
    return rows


def write_to_excel(path: str, rows: list[dict[str, Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "CPL report"

    header = [
        "Тип бизнеса",
        "Наименование сервиса",
        "КОД",
        "Владелец",
        "Объем",
        "% потребления",
    ]
    ws.append(header)

    bold = Font(bold=True)
    for i in range(1, len(header) + 1):
        ws.cell(row=1, column=i).font = bold

    for r in rows:
        ws.append(
            [
                r.get("business_type", ""),
                r.get("service_name", ""),
                r.get("service_id", ""),
                r.get("owner", ""),
                r.get("size_human", ""),
                r.get("pct", 0),
            ]
        )

    for i in range(2, len(rows) + 2):
        ws.cell(row=i, column=6).number_format = "0.0000%"

    wb.save(path)
    log.info(f"Report written: {path}")


def main():
    host, port, use_ssl = parse_host_and_ssl(OPENSEARCH_URL, OPENSEARCH_PORT)

    client = OpenSearch(
        hosts=[{"host": host, "port": port}],
        http_auth=(USER, PASS),
        use_ssl=use_ssl,
        verify_certs=False,
        ssl_show_warn=False,
    )

    rows = fetch_and_aggregate(client)
    sd = read_sd_map(SD_FILE)
    bk = read_bk_map(BK_FILE)

    rows = enrich(rows, sd, bk)
    rows = apply_unknown_service_filter(rows)
    rows = apply_business_type_ban(rows)
    rows = finalize(rows)

    write_to_excel("CPL_report.xlsx", rows)


if __name__ == "__main__":
    main()