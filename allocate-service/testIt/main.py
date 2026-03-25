import os
import logging
import re
from collections import defaultdict

import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s %(levelname)s | %(message)s"
)

load_dotenv()

PG_HOST = os.getenv("PG_HOST")
PG_PORT = os.getenv("PG_PORT")
PG_USER = os.getenv("PG_USER")
PG_PASSWORD = os.getenv("PG_PASSWORD")
PG_DB = os.getenv("PG_DB")

ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")
OUT_XLSX = os.getenv("OUT_XLSX", "testIt_report.xlsx")

BAN_SERVICE_IDS = {15473}

PFP_RE = re.compile(r"пфп\s*[-–—:]?\s*(\d+)", re.IGNORECASE)

QUERY_PROJECTS = """
SELECT
    p."Id",
    p."Name",
    p."Description",
    COUNT(DISTINCT at."Id") AS "AutotestsCount",
    (
        SELECT COUNT(wi."Id")
        FROM "WorkItems" wi
        WHERE wi."ProjectId" = p."Id"
          AND wi."EntityTypeName" = 'TestCases'
          AND (wi."IsDeleted" = FALSE OR wi."IsDeleted" IS NULL)
          AND (wi."IsActual" = TRUE OR wi."IsActual" IS NULL)
    ) AS "TestCasesCount"
FROM "Projects" p
LEFT JOIN "AutoTests" at ON at."ProjectId" = p."Id"
GROUP BY p."Id", p."Name", p."Description"
ORDER BY p."Id";
"""


def die(msg: str, code: int = 2):
    logging.error(msg)
    raise SystemExit(code)


def exec_query(conn, query: str):
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(query)
        return cur.fetchall()


def clean(s):
    if s is None:
        return ""
    return " ".join(str(s).replace(",", " ").split())


def normalize_code(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def extract_pfp(description: str):
    if not description:
        return None
    m = PFP_RE.search(description)
    return m.group(1) if m else None


def to_int(v):
    try:
        return int(v or 0)
    except Exception:
        return 0


def load_activity(path):
    if not path or not os.path.exists(path):
        raise RuntimeError(f"ACTIVITY_FILE не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    m = {}
    for r in ws.iter_rows(values_only=True):
        code = normalize_code(r[0] if len(r) > 0 else "")
        if not code:
            continue

        if code in m:
            continue

        m[code] = {
            "service_name": clean(r[1] if len(r) > 1 else ""),
            "activity_code": clean(r[2] if len(r) > 2 else ""),
            "activity_name": clean(r[3] if len(r) > 3 else ""),
        }

    wb.close()
    logging.info("activity loaded: %d", len(m))
    return m


def write_sheet(wb: Workbook, sheet_name: str, rows, *, make_active=False):
    ws = wb.active if make_active else wb.create_sheet(sheet_name)
    ws.title = sheet_name

    if not rows:
        ws.cell(row=1, column=1, value="No data")
        return ws

    headers = list(rows[0].keys())
    header_font = Font(bold=True)

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = header_font

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

    return ws


def main():
    for k in ("PG_HOST", "PG_PORT", "PG_USER", "PG_PASSWORD", "PG_DB"):
        if not globals().get(k):
            die(f"Env var {k} is not set")

    activity = load_activity(ACTIVITY_FILE)

    logging.info("Connecting to PROJECT DB...")
    with psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        dbname=PG_DB,
        user=PG_USER,
        password=PG_PASSWORD,
    ) as conn:
        projects = exec_query(conn, QUERY_PROJECTS)

    logging.info("Projects loaded: %d", len(projects))

    agg = defaultdict(
        lambda: {
            "ProjectsCount": 0,
            "TestCasesCount": 0,
            "AutotestsCount": 0,
        }
    )
    unaccounted = []

    banned_count = 0
    activity_miss_count = 0
    no_pfp_count = 0

    for p in projects:
        desc = p.get("Description")
        code = extract_pfp(desc)

        tc = to_int(p.get("TestCasesCount"))
        at = to_int(p.get("AutotestsCount"))
        total_tests = tc + at

        if not code:
            no_pfp_count += 1
            unaccounted.append(
                {
                    "ProjectId": p.get("Id"),
                    "ProjectName": p.get("Name"),
                    "Code": "",
                    "ServiceName": "",
                    "ActivityCode": "",
                    "ActivityName": "",
                    "TestCasesCount": tc,
                    "AutotestsCount": at,
                    "TotalTests": total_tests,
                    "Reason": "no_pfp_in_description" if desc else "empty_description",
                    "Detail": desc or "",
                }
            )
            continue

        if code in {str(x) for x in BAN_SERVICE_IDS}:
            banned_count += 1
            meta = activity.get(code, {})
            unaccounted.append(
                {
                    "ProjectId": p.get("Id"),
                    "ProjectName": p.get("Name"),
                    "Code": code,
                    "ServiceName": meta.get("service_name", ""),
                    "ActivityCode": meta.get("activity_code", ""),
                    "ActivityName": meta.get("activity_name", ""),
                    "TestCasesCount": tc,
                    "AutotestsCount": at,
                    "TotalTests": total_tests,
                    "Reason": "banned_service_id",
                    "Detail": "code in BAN_SERVICE_IDS",
                }
            )
            continue

        meta = activity.get(code)
        if not meta:
            activity_miss_count += 1
            unaccounted.append(
                {
                    "ProjectId": p.get("Id"),
                    "ProjectName": p.get("Name"),
                    "Code": code,
                    "ServiceName": "",
                    "ActivityCode": "",
                    "ActivityName": "",
                    "TestCasesCount": tc,
                    "AutotestsCount": at,
                    "TotalTests": total_tests,
                    "Reason": "activity_mapping_miss",
                    "Detail": "code not found in activity.xlsx",
                }
            )
            continue

        agg[code]["ProjectsCount"] += 1
        agg[code]["TestCasesCount"] += tc
        agg[code]["AutotestsCount"] += at

        logging.info(
            'accounted project_id=%s project="%s" service="%s" code=%s tc=%d at=%d total=%d',
            p.get("Id"),
            p.get("Name"),
            meta.get("service_name", ""),
            code,
            tc,
            at,
            total_tests,
        )

    rows = []
    total_all_tests = 0

    for code, v in agg.items():
        total_tests = v["TestCasesCount"] + v["AutotestsCount"]
        total_all_tests += total_tests

        meta = activity.get(code, {})

        rows.append(
            {
                "Имя сервиса": meta.get("service_name", ""),
                "Код": code,
                "Код активности": meta.get("activity_code", ""),
                "Наименование активности": meta.get("activity_name", ""),
                "Кол-во проектов": v["ProjectsCount"],
                "TestCasesCount": v["TestCasesCount"],
                "AutotestsCount": v["AutotestsCount"],
                "TotalTests": total_tests,
                "% потребления": 0.0,
            }
        )

    if total_all_tests == 0:
        total_all_tests = 1

    for r in rows:
        r["% потребления"] = r["TotalTests"] / total_all_tests

    rows.sort(key=lambda r: int(r["Код"]) if str(r["Код"]).isdigit() else 10**18)
    unaccounted.sort(key=lambda x: (-to_int(x.get("TotalTests")), to_int(x.get("ProjectId"))))

    wb = Workbook()
    ws_main = write_sheet(wb, "Отчет TestIt", rows, make_active=True)
    write_sheet(wb, "Unaccounted", unaccounted)

    if rows:
        pct_col = list(rows[0].keys()).index("% потребления") + 1
        for rr in range(2, ws_main.max_row + 1):
            ws_main.cell(row=rr, column=pct_col).number_format = "0.0000%"

    wb.save(OUT_XLSX)

    logging.info("Excel saved: %s", OUT_XLSX)
    logging.info("Projects total: %d", len(projects))
    logging.info("Accounted services: %d", len(rows))
    logging.info("Unaccounted projects: %d", len(unaccounted))
    logging.info("No PFP in description: %d", no_pfp_count)
    logging.info("Banned service ids: %d", banned_count)
    logging.info("Activity mapping miss: %d", activity_miss_count)
    logging.info("Total tests (accounted only): %d", total_all_tests)


if __name__ == "__main__":
    main()