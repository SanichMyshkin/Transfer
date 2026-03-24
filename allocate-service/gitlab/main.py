import os
import re
import time
import logging
from pathlib import Path
from collections import defaultdict

import urllib3
import gitlab
import humanize
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()

GITLAB_URL = (os.getenv("GITLAB_URL") or "").rstrip("/")
GITLAB_TOKEN = os.getenv("GITLAB_TOKEN") or ""

OUT_XLSX = os.getenv("OUT_XLSX", "gitlab_report.xlsx")
ACTIVITY_FILE = os.getenv("ACTIVITY_FILE", "activity.xlsx")

SSL_VERIFY = False
SLEEP_SEC = 0.02
LOG_EVERY = 100
LIMIT = 0

BAN_SERVICE_IDS = {
    "15473",
    "0",
}

SKIP_UNKNOWN_SERVICE_IDS = True

SERVICE_ID_RE = re.compile(r"^service[_-]?id\s*:\s*(\d+)\s*$", re.IGNORECASE)

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("gitlab_sizes")


def die(msg: str, code: int = 2):
    log.error(msg)
    raise SystemExit(code)


def clean_spaces(s: str) -> str:
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def normalize_number(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    return s


def build_ban_set(ban_list):
    if not isinstance(ban_list, (list, tuple, set)):
        die("BAN_SERVICE_IDS должен быть list / tuple / set")
    return {str(x).strip() for x in ban_list if str(x).strip()}


def validate_files():
    if not ACTIVITY_FILE or not os.path.isfile(ACTIVITY_FILE):
        die(f"ACTIVITY_FILE не найден: {ACTIVITY_FILE}")


def load_activity_mapping(path: str):
    df = pd.read_excel(path, dtype=str, engine="openpyxl").fillna("")

    map_by_number = {}

    for _, row in df.iterrows():
        service_id = normalize_number(row.iloc[0] if len(row) > 0 else "")
        service_name = clean_spaces(row.iloc[1] if len(row) > 1 else "")
        activity_code = clean_spaces(row.iloc[2] if len(row) > 2 else "")
        activity_name = clean_spaces(row.iloc[3] if len(row) > 3 else "")

        if not service_id:
            continue

        if service_id in map_by_number:
            continue

        map_by_number[service_id] = {
            "service_name": service_name,
            "activity_code": activity_code,
            "activity_name": activity_name,
        }

    log.info("ACTIVITY: загружено сервисов по номеру=%d", len(map_by_number))
    return map_by_number


def connect():
    if not GITLAB_URL:
        die("GITLAB_URL не задан")
    if not GITLAB_TOKEN:
        die("GITLAB_TOKEN не задан")

    gl = gitlab.Gitlab(
        GITLAB_URL,
        private_token=GITLAB_TOKEN,
        ssl_verify=SSL_VERIFY,
        timeout=60,
        per_page=100,
    )
    gl.auth()
    log.info("Подключение к GitLab успешно")
    return gl


def extract_service_id_info(topics):
    ids = []
    for t in topics or []:
        s = (t or "").strip()
        m = SERVICE_ID_RE.match(s)
        if m:
            ids.append(m.group(1))

    if not ids:
        return "", "MISSING"

    uniq = sorted(set(ids))
    if len(uniq) == 1:
        return uniq[0], "OK"

    return "", "CONFLICT"


def pct(part: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return (part / total) * 100.0


def resolve_activity(service_id: str, map_by_number):
    if not service_id:
        return "", "", ""

    norm_number = normalize_number(service_id)
    if norm_number and norm_number in map_by_number:
        item = map_by_number[norm_number]
        return (
            item.get("service_name", "") or "",
            item.get("activity_code", "") or "",
            item.get("activity_name", "") or "",
        )

    return "", "", ""


def main():
    validate_files()

    ban_set = build_ban_set(BAN_SERVICE_IDS)

    log.info("BAN_SERVICE_IDS=%s", sorted(ban_set) if ban_set else "[]")
    log.info("SKIP_UNKNOWN_SERVICE_IDS=%s", SKIP_UNKNOWN_SERVICE_IDS)

    activity_map = load_activity_mapping(ACTIVITY_FILE)

    gl = connect()

    out_path = str(Path(OUT_XLSX).resolve())
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "ByServiceId"

    ws_sum.append(
        [
            "Имя сервиса",
            "Код",
            "Код активности",
            "Наименование активности",
            "кол-во проектов",
            "Объем репозиториев (bytes)",
            "Объем репозиториев",
            "Объем джобов(bytes)",
            "Объем джобов",
            "Суммарный объем(bytes)",
            "Суммарный объем",
            "% потребления",
        ]
    )
    for c in ws_sum[1]:
        c.font = Font(bold=True)

    ws_unmapped = wb.create_sheet("Unmapped")
    ws_unmapped.append(
        [
            "reason",
            "detail",
            "sid_status",
            "service_id",
            "service_name",
            "activity_code",
            "activity_name",
            "project_id",
            "project",
            "web_url",
            "repo_size_b",
            "repo_size_h",
            "job_artifacts_b",
            "job_artifacts_h",
            "total_b",
            "total_h",
        ]
    )
    for c in ws_unmapped[1]:
        c.font = Font(bold=True)

    agg = defaultdict(
        lambda: {
            "projects": 0,
            "repo": 0,
            "job": 0,
            "service_name": "",
            "activity_code": "",
            "activity_name": "",
        }
    )

    total_repo_all = 0
    total_job_all = 0

    errors = 0
    processed = 0

    skipped_ban_sid = 0
    skipped_activity_miss = 0
    unmapped_missing = 0
    unmapped_conflict = 0

    log.info("Начинаем обход проектов...")

    for i, p in enumerate(gl.projects.list(all=True, iterator=True), start=1):
        if LIMIT and processed >= LIMIT:
            log.info("LIMIT=%d достигнут, останавливаемся", LIMIT)
            break

        proj_id = getattr(p, "id", None)

        try:
            full = gl.projects.get(proj_id, statistics=True)
            name = getattr(full, "path_with_namespace", "") or getattr(full, "name", "") or str(proj_id)
            web_url = getattr(full, "web_url", "") or ""

            topics = list(getattr(full, "topics", []) or [])
            service_id, sid_status = extract_service_id_info(topics)

            stats = getattr(full, "statistics", {}) or {}
            repo_bytes = int(stats.get("repository_size", 0) or 0)
            job_bytes = int(stats.get("job_artifacts_size", 0) or 0)
            total_bytes = repo_bytes + job_bytes

            total_repo_all += repo_bytes
            total_job_all += job_bytes

            service_name = ""
            activity_code = ""
            activity_name = ""

            if sid_status == "OK":
                service_name, activity_code, activity_name = resolve_activity(service_id, activity_map)

                if service_id in ban_set:
                    skipped_ban_sid += 1
                    ws_unmapped.append(
                        [
                            "banned_service_id",
                            "service_id in BAN_SERVICE_IDS",
                            sid_status,
                            service_id,
                            service_name,
                            activity_code,
                            activity_name,
                            proj_id,
                            name,
                            web_url,
                            repo_bytes,
                            humanize.naturalsize(repo_bytes, binary=True),
                            job_bytes,
                            humanize.naturalsize(job_bytes, binary=True) if job_bytes else "",
                            total_bytes,
                            humanize.naturalsize(total_bytes, binary=True),
                        ]
                    )
                    processed += 1
                    continue

                if SKIP_UNKNOWN_SERVICE_IDS and not clean_spaces(service_name):
                    skipped_activity_miss += 1
                    ws_unmapped.append(
                        [
                            "activity_mapping_miss",
                            "no match in activity.xlsx by service_id",
                            sid_status,
                            service_id,
                            service_name,
                            activity_code,
                            activity_name,
                            proj_id,
                            name,
                            web_url,
                            repo_bytes,
                            humanize.naturalsize(repo_bytes, binary=True),
                            job_bytes,
                            humanize.naturalsize(job_bytes, binary=True) if job_bytes else "",
                            total_bytes,
                            humanize.naturalsize(total_bytes, binary=True),
                        ]
                    )
                    processed += 1
                    continue

                a = agg[service_id]
                a["projects"] += 1
                a["repo"] += repo_bytes
                a["job"] += job_bytes

                if not a["service_name"] and service_name:
                    a["service_name"] = service_name
                if not a["activity_code"] and activity_code:
                    a["activity_code"] = activity_code
                if not a["activity_name"] and activity_name:
                    a["activity_name"] = activity_name

            else:
                if sid_status == "MISSING":
                    unmapped_missing += 1
                    reason = "service_id_missing"
                    detail = "no service_id topic found"
                else:
                    unmapped_conflict += 1
                    reason = "service_id_conflict"
                    detail = "multiple different service_id topics"

                ws_unmapped.append(
                    [
                        reason,
                        detail,
                        sid_status,
                        service_id,
                        service_name,
                        activity_code,
                        activity_name,
                        proj_id,
                        name,
                        web_url,
                        repo_bytes,
                        humanize.naturalsize(repo_bytes, binary=True),
                        job_bytes,
                        humanize.naturalsize(job_bytes, binary=True) if job_bytes else "",
                        total_bytes,
                        humanize.naturalsize(total_bytes, binary=True),
                    ]
                )

            processed += 1

        except Exception as e:
            errors += 1
            log.warning("FAIL project_id=%s err=%s", proj_id, e)

        if LOG_EVERY and i % LOG_EVERY == 0:
            log.info("PROGRESS i=%d processed=%d errors=%d", i, processed, errors)

        if SLEEP_SEC:
            time.sleep(SLEEP_SEC)

    total_all = total_repo_all + total_job_all

    def _sort_key(x: str):
        return int(x) if str(x).isdigit() else str(x)

    for service_id in sorted(agg.keys(), key=_sort_key):
        repo_b = agg[service_id]["repo"]
        job_b = agg[service_id]["job"]
        total_b = repo_b + job_b

        ws_sum.append(
            [
                agg[service_id]["service_name"],
                service_id,
                agg[service_id]["activity_code"],
                agg[service_id]["activity_name"],
                agg[service_id]["projects"],
                repo_b,
                humanize.naturalsize(repo_b, binary=True),
                job_b,
                humanize.naturalsize(job_b, binary=True) if job_b else "",
                total_b,
                humanize.naturalsize(total_b, binary=True),
                round(pct(total_b, total_all), 4),
            ]
        )

    wb.save(out_path)

    log.info(
        "Saved: %s | projects=%d | services=%d | unmapped=%d | errors=%d | total=%s",
        out_path,
        processed,
        len(agg),
        ws_unmapped.max_row - 1,
        errors,
        humanize.naturalsize(total_all, binary=True),
    )
    log.info(
        "Unmapped breakdown: missing=%d conflict=%d ban_sid=%d activity_miss=%d",
        unmapped_missing,
        unmapped_conflict,
        skipped_ban_sid,
        skipped_activity_miss,
    )
    log.info("Готово")


if __name__ == "__main__":
    main()