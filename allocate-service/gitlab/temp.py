import os
import time
import logging
from pathlib import Path
from collections import defaultdict

import urllib3
import gitlab
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()

GITLAB_URL = (os.getenv("GITLAB_URL") or "").rstrip("/")
GITLAB_TOKEN = os.getenv("GITLAB_TOKEN") or ""

# Файл, который ты уже создаёшь/правишь: 1-я колонка project_id, 2-я колонка service_id
MAP_XLSX = os.getenv("MAP_XLSX", "gitlab_pr_with_service_id.xlsx")

# Итоговый агрегированный отчёт
OUT_AGG_XLSX = os.getenv("OUT_AGG_XLSX", "gitlab_by_service_id.xlsx")

SSL_VERIFY = False
SLEEP_SEC = float(os.getenv("SLEEP_SEC", "0.02"))
LOG_EVERY = int(os.getenv("LOG_EVERY", "100"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("gitlab_service_agg")


def die(msg: str, code: int = 2):
    log.error(msg)
    raise SystemExit(code)


def connect():
    if not GITLAB_URL:
        die("GITLAB_URL не задан")
    if not GITLAB_TOKEN:
        die("GITLAB_TOKEN не задан")

    gl = gitlab.Gitlab(
        GITLAB_URL,
        private_token=GITLAB_TOKEN,
        ssl_verify=SSL_VERIFY,
        timeout=60,
        per_page=100,
    )
    gl.auth()
    log.info("Подключение к GitLab успешно")
    return gl


def autosize_columns(ws, max_width=80):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        best = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        ws.column_dimensions[letter].width = min(max(10, best + 2), max_width)


def normalize_service_id(v) -> str:
    # service_id может быть "0", 0, "", None, "abc" — всё оставляем строкой
    if v is None:
        return ""
    return str(v).strip()


def read_project_to_service_map(xlsx_path: str) -> dict[int, str]:
    """
    Берём ТОЛЬКО 2 колонки:
    1) project_id
    2) service_id
    """
    path = Path(xlsx_path)
    if not path.exists():
        die(f"MAP_XLSX не найден: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    mapping: dict[int, str] = {}
    bad = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        proj_id = row[0] if len(row) > 0 else None
        service_id = row[1] if len(row) > 1 else None
        if proj_id is None:
            continue

        try:
            pid = int(str(proj_id).strip())
        except Exception:
            bad += 1
            continue

        mapping[pid] = normalize_service_id(service_id)

    log.info(f"Мапа: {xlsx_path} | mapped={len(mapping)} bad_rows={bad}")
    return mapping


def collect_raw_from_gitlab_in_memory(gl):
    """
    Сбор raw байт в память.
    Возвращает dict: project_id -> (repo_bytes, job_bytes, total_bytes)
    """
    raw = {}
    errors = 0
    start_ts = time.time()

    log.info("Начинаем обход всех проектов (raw байты)...")

    for i, p in enumerate(gl.projects.list(all=True, iterator=True), start=1):
        proj_id = getattr(p, "id", None)

        try:
            full = gl.projects.get(proj_id, statistics=True)
            stats = getattr(full, "statistics", {}) or {}

            repo_bytes = int(stats.get("repository_size", 0) or 0)
            job_bytes = int(stats.get("job_artifacts_size", 0) or 0)
            total_bytes = repo_bytes + job_bytes

            raw[int(proj_id)] = (repo_bytes, job_bytes, total_bytes)

        except Exception as e:
            errors += 1
            proj_name = getattr(p, "path_with_namespace", "") or getattr(p, "name", "") or str(proj_id)
            log.warning(f'FAIL project_id={proj_id} project="{proj_name}" err={e}')

        if LOG_EVERY and i % LOG_EVERY == 0:
            elapsed = time.time() - start_ts
            rate = i / elapsed if elapsed > 0 else 0.0
            log.info(f"PROGRESS i={i} rate={rate:.2f}/s errors={errors}")

        if SLEEP_SEC:
            time.sleep(SLEEP_SEC)

    log.info(f"Сбор raw завершён | projects={len(raw)} errors={errors}")
    return raw


def build_agg_report(raw_by_project: dict[int, tuple[int, int, int]], proj2sid: dict[int, str], out_xlsx_path: str):
    """
    Агрегация по service_id:
    service_id / кол-во проектов / repo_bytes / job_bytes / total_bytes / % от всех
    """
    # agg[sid] = [projects_count, repo_sum, job_sum, total_sum]
    agg = defaultdict(lambda: [0, 0, 0, 0])

    total_all = 0
    missed_mapping = 0

    for pid, (repo_b, job_b, total_b) in raw_by_project.items():
        sid = proj2sid.get(pid, None)
        if sid is None:
            missed_mapping += 1
            sid = ""  # нет строки в мапе — группируем в пустой ключ
        sid = normalize_service_id(sid)

        rec = agg[sid]
        rec[0] += 1
        rec[1] += int(repo_b or 0)
        rec[2] += int(job_b or 0)
        rec[3] += int(total_b or 0)

        total_all += int(total_b or 0)

    out_path = str(Path(out_xlsx_path).resolve())
    wb = Workbook()
    ws = wb.active
    ws.title = "ByService"

    headers = [
        "service_id",
        "projects_count",
        "repo_bytes",
        "job_bytes",
        "total_bytes",
        "percent_of_all",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    items = sorted(agg.items(), key=lambda kv: kv[1][3], reverse=True)

    for sid, (cnt, repo_sum, job_sum, total_sum) in items:
        pct = (total_sum / total_all) if total_all > 0 else 0.0
        ws.append([sid, cnt, repo_sum, job_sum, total_sum, pct])

    # формат процента
    for cell in ws["F"][1:]:
        cell.number_format = "0.00%"

    autosize_columns(ws)
    wb.save(out_path)

    log.info(
        f"Saved: {out_path} | groups={len(items)} total_all_bytes={total_all} missed_mapping={missed_mapping}"
    )


def main():
    gl = connect()

    # 1) raw байты в память
    raw_by_project = collect_raw_from_gitlab_in_memory(gl)

    # 2) мапа project_id -> service_id из твоего файла (2 колонки)
    proj2sid = read_project_to_service_map(MAP_XLSX)

    # 3) агрегация и отчёт
    build_agg_report(raw_by_project, proj2sid, OUT_AGG_XLSX)

    log.info("Готово")


if __name__ == "__main__":
    main()
