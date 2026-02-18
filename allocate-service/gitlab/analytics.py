import os
import time
import logging
from pathlib import Path

import urllib3
import gitlab
import pandas as pd
from dotenv import load_dotenv
import humanize
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()

GITLAB_URL = os.getenv("GITLAB_URL")
GITLAB_TOKEN = os.getenv("GITLAB_TOKEN")

BK_FILE = os.getenv("BK_FILE", "bk_all_users.xlsx")
OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "gitlab_projects_detailed.xlsx")

SLEEP_SEC = 0.02
SSL_VERIFY = False
LOG_EVERY = 50

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("gitlab_projects_detailed")


def die(msg: str, code: int = 2):
    log.error(msg)
    raise SystemExit(code)


def normalize_login(x: str) -> str:
    return (x or "").strip().lower()


def clean_spaces(s: str) -> str:
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = " ".join(s.split())
    return s


def connect():
    if not GITLAB_URL:
        die("GITLAB_URL не задан")
    if not GITLAB_TOKEN:
        die("GITLAB_TOKEN не задан")

    log.info("Подключение к GitLab api...")
    gl = gitlab.Gitlab(
        GITLAB_URL,
        private_token=GITLAB_TOKEN,
        ssl_verify=SSL_VERIFY,
        timeout=60,
        per_page=100,
    )
    gl.auth()
    log.info("Подключение успешно")
    return gl


def load_bk_login_business_type_map(path: str):
    if not path or not os.path.isfile(path):
        die(f"BK_FILE не найден: {path}")

    # индексы колонок (0-based) как у тебя
    LOGIN_COL_IDX = 55
    BUSINESS_TYPE_COL_IDX = 44

    df = pd.read_excel(path, dtype=str, engine="openpyxl", header=0).fillna("")

    max_idx = max(LOGIN_COL_IDX, BUSINESS_TYPE_COL_IDX)
    if df.shape[1] <= max_idx:
        die(f"BK_FILE: недостаточно колонок: cols={df.shape[1]}, нужен минимум индекс {max_idx}")

    sub = df.iloc[:, [LOGIN_COL_IDX, BUSINESS_TYPE_COL_IDX]].copy()
    sub.columns = ["login", "business_type"]

    sub["login_key"] = sub["login"].map(normalize_login)
    sub["business_type"] = sub["business_type"].astype(str).map(clean_spaces)

    sub = sub[sub["login_key"] != ""].drop_duplicates("login_key", keep="last")
    mp = dict(zip(sub["login_key"], sub["business_type"]))

    log.info(f"BK: загружено login → тип бизнеса: {len(mp)}")
    return mp


def resolve_business_type(project_name: str, creator_username: str, maintainers: list[str], bk_map: dict[str, str]) -> str:
    creator_key = normalize_login(creator_username)
    if creator_key:
        bt = bk_map.get(creator_key, "")
        if bt:
            log.info(f'BT project="{project_name}" source=creator bt="{bt}"')
            return bt
        log.info(f'BT project="{project_name}" creator="{creator_username}" not_found')

    found = []
    for u in maintainers:
        k = normalize_login(u)
        t = bk_map.get(k, "")
        if t:
            found.append((u, t))

    if not found:
        log.info(f'BT project="{project_name}" source=none bt=""')
        return ""

    counts = {}
    for _, t in found:
        counts[t] = counts.get(t, 0) + 1

    max_cnt = max(counts.values())
    winners = sorted([k for k, v in counts.items() if v == max_cnt])
    bt = winners[0]
    log.info(f'BT project="{project_name}" source=maintainers chosen="{bt}"')
    return bt


def autosize_columns(ws, max_width=80):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        best = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        ws.column_dimensions[letter].width = min(max(10, best + 2), max_width)


def main():
    log.info("Старт детального отчета GitLab проектов (без агрегации)")
    gl = connect()

    log.info("Загрузка BK (login → business_type)")
    bk_map = load_bk_login_business_type_map(BK_FILE)

    out_path = str(Path(OUTPUT_XLSX).resolve())
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    headers = [
        "project_id",
        "project",
        "web_url",
        "creator",
        "maintainers",
        "business_type",
        "repo_size",
        "job_artifacts_size",
        "total_size",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    user_cache: dict[int, str] = {}
    errors = 0
    start_ts = time.time()

    log.info("Начинаем обход ВСЕХ проектов (без лимита)")

    for i, p in enumerate(gl.projects.list(all=True, iterator=True), start=1):
        proj_id = getattr(p, "id", None)
        proj_name = getattr(p, "path_with_namespace", None) or getattr(p, "name", None) or str(proj_id)

        try:
            full = gl.projects.get(proj_id, statistics=True)

            # creator
            creator_username = ""
            creator_id = getattr(full, "creator_id", None)
            if creator_id:
                if creator_id in user_cache:
                    creator_username = user_cache[creator_id]
                else:
                    u = gl.users.get(creator_id)
                    creator_username = (getattr(u, "username", "") or "").strip()
                    user_cache[creator_id] = creator_username

            # maintainers (access >= 40)
            maintainers = []
            try:
                members = full.members_all.list(all=True)
            except Exception:
                members = full.members.list(all=True)

            for m in members:
                lvl = int(getattr(m, "access_level", 0) or 0)
                if lvl >= 40:
                    uname = (getattr(m, "username", "") or "").strip()
                    if uname:
                        maintainers.append(uname)

            maintainers_unique = sorted(set(maintainers))
            maintainers_str = ", ".join(maintainers_unique)

            # business type (опционально, но раз ты сказал "данные что мы получили" — оставил)
            bt = clean_spaces(resolve_business_type(proj_name, creator_username, maintainers_unique, bk_map))

            # sizes
            stats = getattr(full, "statistics", {}) or {}
            repo_bytes = int(stats.get("repository_size", 0) or 0)
            job_bytes = int(stats.get("job_artifacts_size", 0) or 0)
            total_bytes = repo_bytes + job_bytes

            ws.append(
                [
                    proj_id,
                    proj_name,
                    getattr(full, "web_url", "") or "",
                    creator_username,
                    maintainers_str,
                    bt,
                    humanize.naturalsize(repo_bytes, binary=True),
                    humanize.naturalsize(job_bytes, binary=True) if job_bytes else "",
                    humanize.naturalsize(total_bytes, binary=True),
                ]
            )

        except Exception as e:
            errors += 1
            log.warning(f'FAIL project="{proj_name}" err={e}')

        if LOG_EVERY and i % LOG_EVERY == 0:
            elapsed = time.time() - start_ts
            rate = i / elapsed if elapsed > 0 else 0
            log.info(f"PROGRESS i={i} rate={rate:.2f} proj/s errors={errors}")

        if SLEEP_SEC > 0:
            time.sleep(SLEEP_SEC)

    autosize_columns(ws)
    wb.save(out_path)
    log.info(f"Saved: {out_path} | rows={ws.max_row - 1} errors={errors}")
    log.info("Готово")


if __name__ == "__main__":
    main()
